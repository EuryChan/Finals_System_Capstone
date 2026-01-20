from django.shortcuts import render, redirect, get_object_or_404
from .models import Employee
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
import json
from django.contrib.auth.models import User
from .models import UserProfile
from django.core.exceptions import ValidationError
from django.contrib.auth.password_validation import validate_password
from .decorators import role_required
from django.db.models import Q, Count, Avg, Sum
from django.db import transaction, connection
from django.core.paginator import Paginator
from django.core.cache import cache
from django.http import HttpResponse, HttpResponseForbidden
import csv
import openpyxl
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Q, Count, Avg, Sum  
from .models import EligibilityRequest, Barangay, Requirement, RequirementSubmission, RequirementAttachment, Notification, Announcement, FileCategory, MonitoringFile, BarangayOfficial
from django.views.decorators.http import require_POST
import os
from datetime import date
from .models import CategorizedFile
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import timedelta
import traceback
from PIL import Image
import pytesseract, PyPDF2
from app.tasks import send_email_task 




try:
    from .models import AuditLog
except ImportError:
    class AuditLog:
        @staticmethod
        def objects():
            return None
        
        @staticmethod
        def create(**kwargs):
            pass

#------------USER ACCESS LOGIN/LOGOUT VIEW------------#
def get_client_ip(request):
    """Helper function to get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def landing_page(request):
    return render(request, 'landing.html')    

def logout_view(request):
    if request.user.is_authenticated:
        try:
            AuditLog.objects.create(
                user=request.user,
                action='LOGOUT',
                ip_address=get_client_ip(request),
                description=f"User {request.user.username} logged out"
            )
        except:
            pass  # Skip audit logging if AuditLog doesn't exist
    
    logout(request)
    messages.success(request, "You have successfully logged out.")
    return redirect('landing_page')

def login_page(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        print(f"\n{'='*60}")
        print(f"🔐 LOGIN ATTEMPT")
        print(f"{'='*60}")
        print(f"Username: {username}")
        print(f"Password length: {len(password) if password else 0}")
        
        user = authenticate(request, username=username, password=password)
        print(f"Authentication result: {user}")
        
        if user is not None:
            print(f"✅ User authenticated successfully")
            
            # Check if user has profile
            try:
                profile = user.userprofile
                print(f"✅ Profile found")
                print(f"   Role: {profile.role}")
                print(f"   Approved: {profile.is_approved}")
                print(f"   Barangay: {profile.barangay}")
            except UserProfile.DoesNotExist:
                print(f"❌ ERROR: No UserProfile found for {username}")
                profile, created = UserProfile.objects.get_or_create(user=user)
                print(f"   Created new profile: {created}")
            
            # Check approval
            if not profile.is_approved:
                print(f"❌ User not approved - blocking login")
                messages.warning(request, '⏳ Your account is pending approval. Please wait for administrator approval.')
                return redirect('login_page')
            
            print(f"✅ User is approved - proceeding with login")
            
            # Log the user in
            login(request, user)
            print(f"✅ Django login() called")
            
            profile.update_login_info(get_client_ip(request))
            
            # Get redirect URL
            redirect_url = profile.get_redirect_url()
            print(f"🔍 Redirect URL: {redirect_url}")
            print(f"{'='*60}\n")
            
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect(redirect_url)
        
        else:
            print(f"❌ Authentication FAILED")
            print(f"{'='*60}\n")
            messages.error(request, "Invalid username or password.")

    return render(request, 'login_page.html')

@login_required
def debug_user_role(request):
    """Temporary debug view"""
    profile = request.user.userprofile
    return JsonResponse({
        'raw_role': repr(profile.role),  # Shows exact string including whitespace
        'role_length': len(profile.role),
        'expected_roles': ['barangay official', 'municipal officer', 'dilg staff'],
        'matches': {
            'barangay official': profile.role.strip().lower() == 'barangay official',
            'municipal officer': profile.role.strip().lower() == 'municipal officer',
            'dilg staff': profile.role.strip().lower() == 'dilg staff',
        }
    })

def landing_menu(request):
    return render(request, 'landing_menu.html')

@login_required 
def debug_employee(request):
    """
    Debug view to inspect Employee model
    """
    try:
        # Get first employee
        emp = Employee.objects.first()
        
        # Get all field names
        field_names = [f.name for f in Employee._meta.fields]
        
        # Get sample data safely
        sample_data = {}
        if emp:
            for field_name in field_names:
                try:
                    value = getattr(emp, field_name, 'N/A')
                    sample_data[field_name] = str(value)[:100]  # Limit length
                except:
                    sample_data[field_name] = 'Error reading'
        
        return JsonResponse({
            'total_employees': Employee.objects.count(),
            'field_names': field_names,
            'sample_data': sample_data
        }, indent=2)
        
    except Exception as e:
        return JsonResponse({'error': str(e)})


#------------APPLICATION REQUEST VIEW------------#
@login_required
@require_http_methods(["GET"])
def api_get_eligibility_request(request, request_id):
    """
    API endpoint to get eligibility request details including documents
    """
    try:
        eligibility_request = get_object_or_404(EligibilityRequest, id=request_id)
        
        data = {
            'id': eligibility_request.id,
            'full_name': eligibility_request.full_name,
            'email': eligibility_request.email,
            'barangay': eligibility_request.barangay,
            'position_type': eligibility_request.position_type,
            'certifier': eligibility_request.get_certifier_display(),
            'status': eligibility_request.status,
            'date_submitted': eligibility_request.date_submitted.strftime('%B %d, %Y') if eligibility_request.date_submitted else 'N/A',
            'date_processed': eligibility_request.date_processed.strftime('%B %d, %Y') if eligibility_request.date_processed else 'N/A',
            'approved_by': eligibility_request.approved_by.get_full_name() if eligibility_request.approved_by else 'N/A',
            'rejection_reason': eligibility_request.rejection_reason if hasattr(eligibility_request, 'rejection_reason') else None,

            # ✅ Document URLs
            'id_front': eligibility_request.id_front.url if eligibility_request.id_front else None,
            'id_back': eligibility_request.id_back.url if eligibility_request.id_back else None,
            'signature': eligibility_request.signature.url if eligibility_request.signature else None,
        }
        
        return JsonResponse({
            'success': True,
            'request': data
        })
        
    except Exception as e:
        import traceback
        print(f"Error in api_get_eligibility_request: {str(e)}")
        print(traceback.format_exc())
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def application_request(request):
    # Get all eligibility requests
    requests = EligibilityRequest.objects.all().order_by('-date_submitted')
    
     # Count monitoring files
    monitoring_count = CategorizedFile.objects.filter(
        category__name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually'],
        is_archived=False
    ).count()
    
    # Count certification files
    certification_count = CategorizedFile.objects.filter(
        category__name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures'],
        is_archived=False
    ).count()
    
    # Count pending users
    pending_count = User.objects.filter(
        userprofile__is_approved=False
    ).count()
    
    # Count pending applications
    pending_applications_count = EligibilityRequest.objects.filter(
        status='pending',
        archived=False
    ).count()
    
    context = {
        'requests': requests,
        'pending_applications_count': pending_applications_count,
        'pending_count': pending_count,  # for User Approvals badge
        'monitoring_count': monitoring_count,
        'certification_count': certification_count,
    }
    return render(request, 'application_request.html', context)

from django.views.decorators.http import require_http_methods
from django.http import JsonResponse

@require_http_methods(["POST"])
def archive_application(request, application_id):
    try:
        app = get_object_or_404(EligibilityRequest, id=application_id)
        app.archived = True
        app.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def restore_application(request, application_id):
    try:
        app = get_object_or_404(EligibilityRequest, id=application_id)
        app.archived = False
        app.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


#------------APPLICATION REQUEST VIEW END------------#


#------------EMPLOYEES PROFILE VIEW------------#

def employees_profile(request):
    # Handle POST - Add new employee with validation
    if request.method == 'POST':
        try:
            with transaction.atomic():  # Ensure data consistency
                name = request.POST.get('name', '').strip()
                id_no = request.POST.get('id_no', '').strip()
                task = request.POST.get('task', '').strip()
                email = request.POST.get('email', '').strip()
                phone = request.POST.get('phone', '').strip()
                department = request.POST.get('department', '')
                position = request.POST.get('position', '').strip()
                hire_date = request.POST.get('hire_date')
                supervisor_id = request.POST.get('supervisor')
                
                # Validation
                if not all([name, id_no, task]):
                    return JsonResponse({
                        'success': False, 
                        'error': 'Name, ID number, and task are required'
                    }, status=400)
                
                # Check for duplicate ID
                if Employee.objects.filter(id_no=id_no).exists():
                    return JsonResponse({
                        'success': False, 
                        'error': 'Employee ID already exists'
                    }, status=400)
                # Count monitoring files

                # Create employee
                employee_data = {
                    'name': name,
                    'id_no': id_no,
                    'task': task,
                    'department': department if department else None,
                    'position': position,
                }
                
                if email:
                    employee_data['email'] = email
                if phone:
                    employee_data['phone'] = phone
                if hire_date:
                    employee_data['hire_date'] = hire_date
                if supervisor_id:
                    employee_data['supervisor_id'] = supervisor_id
                
                employee = Employee.objects.create(**employee_data)
                
                messages.success(request, f'Employee {name} added successfully!')
                
                # Clear cache
                cache.delete('employee_stats')
    
                
                return redirect('employees_profile')
                
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'error': f'Error creating employee: {str(e)}'
            }, status=400)
    
    # Handle GET - Display employees with advanced filtering and pagination
    search_query = request.GET.get('search', '').strip()
    department_filter = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort', 'name')
    
    # Base queryset with optimizations
    employees = Employee.objects.select_related('supervisor').prefetch_related('subordinates')
    
    # Apply filters
    if search_query:
        employees = employees.filter(
            Q(name__icontains=search_query) |
            Q(id_no__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(task__icontains=search_query) |
            Q(position__icontains=search_query)
        )
    
    if department_filter:
        employees = employees.filter(department=department_filter)
    
    if status_filter:
        employees = employees.filter(status=status_filter)
    
    # Sorting
    valid_sort_fields = ['name', '-name', 'id_no', '-id_no', 'created_at', '-created_at', 'department']
    if sort_by in valid_sort_fields:
        employees = employees.order_by(sort_by)
    
    # Pagination
    paginator = Paginator(employees, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get statistics (cached for performance)
    stats = cache.get('employee_stats')
    if not stats:
        try:
            stats = Employee.get_statistics()
        except AttributeError:
            # If get_statistics method doesn't exist, create basic stats
            stats = {
                'total': Employee.objects.count(),
                'active': Employee.objects.filter(status='active').count() if hasattr(Employee, 'status') else 0,
                'departments': Employee.objects.values('department').distinct().count()
            }
        cache.set('employee_stats', stats, 300)  # Cache for 5 minutes
    
    # Get choices for dropdowns
    supervisors = Employee.objects.filter(status='active') if hasattr(Employee, 'status') else Employee.objects.all()
    
    # Get department and status choices if they exist
    department_choices = getattr(Employee, 'DEPARTMENT_CHOICES', [])
    status_choices = getattr(Employee, 'STATUS_CHOICES', [])

    monitoring_count = CategorizedFile.objects.filter(
        category__name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually'],
        is_archived=False
    ).count()
    
    # Count certification files
    certification_count = CategorizedFile.objects.filter(
        category__name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures'],
        is_archived=False
    ).count()
    
    # Count pending users
    pending_count = User.objects.filter(
        userprofile__is_approved=False
    ).count()
    
    # Count pending applications
    pending_applications_count = EligibilityRequest.objects.filter(
        status='pending',
        archived=False
    ).count()

    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'department_filter': department_filter,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'stats': stats,
        'supervisors': supervisors,
        'department_choices': department_choices,
        'status_choices': status_choices,
        'monitoring_count': monitoring_count,
        'certification_count': certification_count,
        'pending_count': pending_count,
        'pending_applications_count': pending_applications_count,
                
    }
    
    return render(request, 'employees_profile.html', context)


@require_http_methods(["POST"])
def edit_employee(request, employee_id):
    """Handle AJAX edit employee requests"""
    try:
        employee = get_object_or_404(Employee, pk=employee_id)
        
        # Parse JSON data from request body
        data = json.loads(request.body)
        
        name = data.get('name', '').strip()
        id_no = data.get('id_no', '').strip()
        task = data.get('task', '').strip()
        
        # Validate required fields
        if not name:
            return JsonResponse({
                'success': False, 
                'error': 'Name is required'
            }, status=400)
        
        if not id_no:
            return JsonResponse({
                'success': False, 
                'error': 'ID number is required'
            }, status=400)
        
        if not task:
            return JsonResponse({
                'success': False, 
                'error': 'Task assignment is required'
            }, status=400)
        
        # Check if ID number is already taken by another employee
        existing_employee = Employee.objects.filter(id_no=id_no).exclude(pk=employee_id).first()
        if existing_employee:
            return JsonResponse({
                'success': False, 
                'error': 'This ID number is already assigned to another employee'
            }, status=400)
        
        # Update employee
        employee.name = name
        employee.id_no = id_no
        employee.task = task
        employee.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Employee updated successfully'
        })
        
    except Employee.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Employee not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False, 
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred: {str(e)}'
        }, status=500)


@require_http_methods(["DELETE"])
def delete_employee(request, employee_id):
    try:
        employee = Employee.objects.get(id=employee_id)
        employee_name = employee.name  # Store name for response
        employee.delete()
        return JsonResponse({
            'success': True,
            'message': f'Employee {employee_name} deleted successfully'
        })
    except Employee.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Employee not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'An error occurred: {str(e)}'
        }, status=500)


# In your views.py

from django.views.decorators.http import require_http_methods
from django.http import JsonResponse
from django.shortcuts import get_object_or_404

@require_http_methods(["POST"])
def archive_employee(request, employee_id):
    try:
        employee = get_object_or_404(Employee, id=employee_id)
        employee.archived = True
        employee.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@require_http_methods(["POST"])
def restore_employee(request, employee_id):
    try:
        employee = get_object_or_404(Employee, id=employee_id)
        employee.archived = False
        employee.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
def employee_search_api(request):
    """AJAX endpoint for advanced employee search"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'employees': []})
    
    # Complex search across multiple fields
    try:
        employees = Employee.objects.filter(
            Q(name__icontains=query) |
            Q(id_no__icontains=query) |
            Q(email__icontains=query) |
            Q(position__icontains=query) |
            Q(department__icontains=query)
        ).values('id', 'name', 'id_no', 'email', 'department', 'status')[:10]
    except:
        employees = Employee.objects.filter(
            Q(name__icontains=query) |
            Q(id_no__icontains=query)
        ).values('id', 'name', 'id_no')[:10]
    
    return JsonResponse({
        'employees': list(employees)
    })


# Bulk operations
@login_required
@require_http_methods(["POST"])
def bulk_employee_operations(request):
    """Handle bulk operations on employees"""
    try:
        data = json.loads(request.body)
        action = data.get('action')
        employee_ids = data.get('employee_ids', [])
        
        if not employee_ids:
            return JsonResponse({'success': False, 'error': 'No employees selected'})
        
        with transaction.atomic():
            employees = Employee.objects.filter(id__in=employee_ids)
            
            if action == 'delete':
                count = employees.count()
                employees.delete()
                message = f'{count} employees deleted successfully'
                
            elif action == 'activate':
                if hasattr(Employee, 'status'):
                    count = employees.update(status='active')
                    message = f'{count} employees activated'
                else:
                    return JsonResponse({'success': False, 'error': 'Status field not available'})
                
            elif action == 'deactivate':
                if hasattr(Employee, 'status'):
                    count = employees.update(status='inactive')
                    message = f'{count} employees deactivated'
                else:
                    return JsonResponse({'success': False, 'error': 'Status field not available'})
                
            elif action == 'update_department':
                department = data.get('department')
                if not department:
                    return JsonResponse({'success': False, 'error': 'Department required'})
                if hasattr(Employee, 'department'):
                    count = employees.update(department=department)
                    message = f'{count} employees moved to {department}'
                else:
                    return JsonResponse({'success': False, 'error': 'Department field not available'})
                
            else:
                return JsonResponse({'success': False, 'error': 'Invalid action'})
            
            # Log bulk operation
            try:
                AuditLog.objects.create(
                    user=request.user,
                    action='UPDATE',
                    description=f"Bulk operation: {action} on {len(employee_ids)} employees"
                )
            except:
                pass  # Skip audit logging if AuditLog doesn't exist
            
            # Clear cache
            cache.delete('employee_stats')
            
            return JsonResponse({'success': True, 'message': message})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


#------------EMPLOYEES PROFILE VIEW END------------#


#-----------SETTINGS VIEW-----------#
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
import json


# views.py

@login_required
def settings(request):
    """Settings page with notification counts for sidebar"""
    from .models import CategorizedFile
    
    # Count monitoring files
    monitoring_count = CategorizedFile.objects.filter(
        category__name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually'],
        is_archived=False
    ).count()
    
    # Count certification files
    certification_count = CategorizedFile.objects.filter(
        category__name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures'],
        is_archived=False
    ).count()
    
    # Count pending users
    pending_count = User.objects.filter(
        userprofile__is_approved=False
    ).count()
    
    # Count pending applications
    pending_applications_count = EligibilityRequest.objects.filter(
        status='pending',
        archived=False
    ).count()
    
    context = {
        'monitoring_count': monitoring_count,
        'certification_count': certification_count,
        'pending_count': pending_count,
        'pending_applications_count': pending_applications_count,
    }
    
    return render(request, 'settings.html', context)


@login_required
def update_profile(request):
    """Update user profile information"""
    if request.method == 'POST':
        try:
            user = request.user
            user.first_name = request.POST.get('firstName', '').strip()
            user.last_name = request.POST.get('lastName', '').strip()
            user.email = request.POST.get('email', '').strip()
            user.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Profile updated successfully'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def update_account(request):
    """Update user account settings (role)"""
    if request.method == 'POST':
        try:
            user = request.user
            role = request.POST.get('role', '').strip().lower()
            
            # Validate role against ROLE_CHOICES
            valid_roles = ['barangay official', 'municipal officer', 'dilg staff']
            
            if role not in valid_roles:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid role selected'
                })
            
            # Update role in profile
            if hasattr(user, 'userprofile'):
                user.userprofile.role = role
                user.userprofile.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Account updated successfully',
                    'role': role
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'User profile not found'
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
  
    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
@require_http_methods(["POST"])
def change_password(request):
    try:

        current_password = request.POST.get('currentPassword', '')
        new_password = request.POST.get('newPassword', '')
        confirm_password = request.POST.get('confirmPassword', '')
        
        print(f"🔐 Password Change Request")
        print(f"   User: {request.user.username}")
        print(f"   Current password provided: {'Yes' if current_password else 'No'}")
        print(f"   New password length: {len(new_password)}")
        print(f"   Confirm password length: {len(confirm_password)}")

        if not all([current_password, new_password, confirm_password]):
            return JsonResponse({
                'success': False,
                'message': 'All fields are required'
            }, status=400)
        
        if new_password != confirm_password:
            return JsonResponse({
                'success': False,
                'message': 'New passwords do not match'
            }, status=400)
        

        if len(new_password) < 8:
            return JsonResponse({
                'success': False,
                'message': 'Password must be at least 8 characters long'
            }, status=400)
        

        user = request.user
        if not user.check_password(current_password):
            print(f"   ❌ Current password is incorrect")
            return JsonResponse({
                'success': False,
                'message': 'Current password is incorrect'
            }, status=400)
        
        print(f"    Current password verified")
        user.set_password(new_password)
        user.save()      
        print(f"    New password saved to database")
        update_session_auth_hash(request, user)
        
        print(f"    Session updated - user remains logged in")
        print(f" Password changed successfully for user: {user.username}")
        
        return JsonResponse({
            'success': True,
            'message': 'Password changed successfully'
        })
        
    except Exception as e:
        print(f" Error changing password: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while changing password'
        }, status=500)
    

@login_required
def get_notification_preferences(request):
    try:
        profile = request.user.userprofile
        prefs = profile.notification_preferences or {
            'emailApplications': True,
            'emailCertificates': True,
            'emailAnnouncements': True,
            'pushDesktop': False,
            'pushMobile': False
        }
        
        return JsonResponse({'success': True, 'preferences': prefs})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
def update_notifications(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            profile = request.user.userprofile
            profile.notification_preferences = data
            profile.save()
            
            return JsonResponse({'success': True, 'message': 'Preferences saved'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

@login_required
def toggle_2fa(request):
    """Toggle two-factor authentication"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            enabled = data.get('enabled', False)
            
            # Store in session (or add field to UserProfile later)
            request.session['two_factor_enabled'] = enabled
            
            return JsonResponse({
                'success': True,
                'message': '2FA ' + ('enabled' if enabled else 'disabled')
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def delete_account(request):
    """Delete user account (soft delete)"""
    if request.method == 'POST':
        try:
            user = request.user
            # Soft delete - deactivate account
            user.is_active = False
            user.save()
            
            # Or hard delete (uncomment if you want permanent deletion):
            # user.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Account deleted successfully'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def get_user_stats(request):
    """Get user statistics for dashboard"""
    try:
        user = request.user
        profile = user.userprofile
        
        # You can add logic to count actual requests from your models
        # For now, returning basic stats
        stats = {
            'loginCount': profile.login_count,
            'requestCount': 0,  # Add your logic to count requests
            'memberSince': user.date_joined.year
        }
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })


#-----------SETTINGS VIEW END-----------#

def civil_service_certification(request):
    """Render the public certification form"""
    return render(request, 'civil_service_certification.html')


def application_letter(request):
    return render(request, 'application_letter.html')


def monitoring_filess(request):
    return render(request, 'monitoring_files.html')


def certification_filess(request):
    return render(request, 'certification_files.html')


def signup_page(request):
    if request.method == 'POST':
        # ADD DEBUG PRINTS
        print("=" * 50)
        print("🔍 SIGNUP DEBUG")
        print("=" * 50)
        
        username = request.POST.get('username')
        email = request.POST.get('email')
        role = request.POST.get('role')
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        contact_number = request.POST.get('contact_number')
        barangay = request.POST.get('barangay')
        city = request.POST.get('city')
        
        # Print what we received
        print(f"Username: {username}")
        print(f"Email: {email}")
        print(f"Role: {role}")
        print(f"Password1: {'***' if password1 else None}")
        print(f"Password2: {'***' if password2 else None}")
        print(f"Contact Number: {contact_number}")
        print(f"Barangay: {barangay}")
        print(f"City: {city}")
        print(f"All fields filled: {all([username, email, role, password1, password2, contact_number, barangay, city])}")
        print("=" * 50)

        if not all([username, email, role, password1, password2, contact_number, barangay, city]):
            print("❌ FAILED: Missing fields")
            messages.error(request, 'Please fill in all fields.')
            return render(request, 'signup_page.html')

        if password1 != password2:
            print("❌ FAILED: Passwords don't match")
            messages.error(request, 'Passwords do not match.')
            return render(request, 'signup_page.html')

        if User.objects.filter(username=username).exists():
            print("❌ FAILED: Username exists")
            messages.error(request, 'Username already exists.')
            return render(request, 'signup_page.html')

        if User.objects.filter(email=email).exists():
            print("❌ FAILED: Email exists")
            messages.error(request, 'Email already registered.')
            return render(request, 'signup_page.html')

        try:
            validate_password(password1)
            print("✅ Password validation passed")
        except ValidationError as e:
            print(f"❌ FAILED: Password validation - {e}")
            for error in e:
                messages.error(request, error)
            return render(request, 'signup_page.html')

        # Create the user (set is_active to False for admin approval)
        print("🚀 Creating user...")
        user = User.objects.create_user(username=username, email=email, password=password1)
        user.is_active = False  # User must be approved by admin
        user.save()
        print(f"✅ User created: {user.id}")

        # Create or get profile with all fields
        UserProfile.objects.get_or_create(
            user=user, 
            defaults={
                'role': role.strip(),
                'is_approved': False,
                'contact_number': contact_number,
                'barangay': barangay,
                'city': city
            }
        )
        print("✅ Profile created with additional info")

        messages.success(request, 'Your account has been created and is waiting for admin verification.')
        print("✅ Redirecting to pending approval page...")
        return redirect('signup_pending')  # Change this from 'login_page' to 'signup_pending'

    return render(request, 'signup_page.html')


def signup_pending(request):
    """Show 'waiting for approval' message after signup"""
    return render(request, 'signup_pending.html')

def signup_message(request):
    return render(request, 'signup_message.html')


@login_required
def pending_users(request):
    """Admin dashboard to view and approve pending users"""
    print("🔍 PENDING USERS VIEW CALLED!")  # ADD THIS
    print(f"User: {request.user.username}")  # ADD THIS
    print(f"Is staff: {request.user.is_staff}")  # ADD THIS
    
    # Check if user is admin/staff
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'Access denied. Admin only.')
        return redirect('dashboard')
    
    # Get all pending users (not approved yet)
    pending = UserProfile.objects.filter(
        is_approved=False
    ).select_related('user').order_by('-submitted_at')

     # Count monitoring files
    monitoring_count = CategorizedFile.objects.filter(
        category__name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually'],
        is_archived=False
    ).count()
    
    # Count certification files
    certification_count = CategorizedFile.objects.filter(
        category__name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures'],
        is_archived=False
    ).count()
    
    # Count pending users
    pending_count = User.objects.filter(
        userprofile__is_approved=False
    ).count()
    
    # Count pending applications
    pending_applications_count = EligibilityRequest.objects.filter(
        status='pending',
        archived=False
    ).count()
    
    # Get recently approved users (last 10)
    approved = UserProfile.objects.filter(
        is_approved=True,
        approved_at__isnull=False
    ).select_related('user', 'approved_by').order_by('-approved_at')[:10]
    
    context = {
        'pending_users': pending,
        'approved_users': approved,
        'pending_count': pending.count(),
        'monitoring_count': monitoring_count,
        'certification_count': certification_count,
        'pending_count': pending_count,
        'pending_applications_count': pending_applications_count,
    }
    return render(request, 'pending_users.html', context)


@login_required
def approve_user(request, user_id):
    """Approve a pending user"""
    if not request.user.is_staff and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        profile = get_object_or_404(UserProfile, user_id=user_id)
        
        profile.is_approved = True
        profile.approved_at = timezone.now()
        profile.approved_by = request.user
        profile.save()
        
        # Log the approval
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            description=f"Approved user: {profile.user.username} ({profile.role})"
        )
        
        messages.success(request, f'✅ User {profile.user.username} has been approved!')
        return redirect('pending_users')
        
    except Exception as e:
        messages.error(request, f'Error approving user: {str(e)}')
        return redirect('pending_users')

@login_required
def reject_user(request, user_id):
    """Reject and delete a pending user"""
    if not request.user.is_staff and not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    
    try:
        profile = get_object_or_404(UserProfile, user_id=user_id)
        username = profile.user.username
        
        # Log before deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            description=f"Rejected and deleted user: {username} ({profile.role})"
        )
        
        # Delete the user (profile will be deleted via cascade)
        profile.user.delete()
        
        messages.success(request, f'❌ User {username} has been rejected and deleted.')
        return redirect('pending_users')
        
    except Exception as e:
        messages.error(request, f'Error rejecting user: {str(e)}')
        return redirect('pending_users')

@login_required
def export_employees(request):
    """Export employees data"""
    format_type = request.GET.get('format', 'csv')
    
    employees = Employee.objects.select_related('supervisor').all()
    
    if format_type == 'excel':
        # Create Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="employees_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Employees'
        
        # Headers
        headers = ['ID', 'Name', 'Employee ID', 'Email', 'Department', 'Position', 'Status', 'Task', 'Supervisor', 'Hire Date', 'Created At']
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Data
        for row, employee in enumerate(employees, 2):
            worksheet.cell(row=row, column=1, value=employee.id)
            worksheet.cell(row=row, column=2, value=employee.name)
            worksheet.cell(row=row, column=3, value=employee.id_no)
            worksheet.cell(row=row, column=4, value=getattr(employee, 'email', '') or '')
            
            # Handle department display
            if hasattr(employee, 'get_department_display'):
                dept = employee.get_department_display() or ''
            else:
                dept = getattr(employee, 'department', '') or ''
            worksheet.cell(row=row, column=5, value=dept)
            
            worksheet.cell(row=row, column=6, value=getattr(employee, 'position', '') or '')
            
            # Handle status display
            if hasattr(employee, 'get_status_display'):
                status = employee.get_status_display()
            else:
                status = getattr(employee, 'status', '') or ''
            worksheet.cell(row=row, column=7, value=status)
            
            worksheet.cell(row=row, column=8, value=getattr(employee, 'task', '') or '')
            worksheet.cell(row=row, column=9, value=employee.supervisor.name if employee.supervisor else '')
            worksheet.cell(row=row, column=10, value=getattr(employee, 'hire_date', ''))
            worksheet.cell(row=row, column=11, value=getattr(employee, 'created_at', ''))
        
        workbook.save(response)
        
    else:  # CSV format
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="employees_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Name', 'Employee ID', 'Email', 'Department', 'Position', 'Status', 'Task', 'Supervisor', 'Hire Date', 'Created At'])
        
        for employee in employees:
            # Handle department display
            if hasattr(employee, 'get_department_display'):
                dept = employee.get_department_display() or ''
            else:
                dept = getattr(employee, 'department', '') or ''
            
            # Handle status display
            if hasattr(employee, 'get_status_display'):
                status = employee.get_status_display()
            else:
                status = getattr(employee, 'status', '') or ''
            
            writer.writerow([
                employee.id,
                employee.name,
                employee.id_no,
                getattr(employee, 'email', '') or '',
                dept,
                getattr(employee, 'position', '') or '',
                status,
                getattr(employee, 'task', '') or '',
                employee.supervisor.name if employee.supervisor else '',
                getattr(employee, 'hire_date', ''),
                getattr(employee, 'created_at', '')
            ])
    
    # Log export action
    try:
        AuditLog.objects.create(
            user=request.user,
            action='CREATE',
            description=f"Exported employees data as {format_type.upper()}"
        )
    except:
        pass  # Skip audit logging if AuditLog doesn't exist
    
    return response



#---------------ANALYTICS DASHBOARD VIEW---------------#
@login_required
@role_required('dilg staff')
def dashboard(request):
    """
    Main Analytics Dashboard - Real-Time with Chart Data
    Shows: Employee stats, certification trends, barangay performance
    """
    try:
        from datetime import timedelta
        current_year = timezone.now().year
        current_month = timezone.now().month
        
        print(f"🔍 Loading dashboard for year={current_year}, month={current_month}")
        
        # ============================================
        # STATS CARDS
        # ============================================
        total_employees = Employee.objects.count()
        active_employees = Employee.objects.filter(
            Q(status='active') | Q(status__isnull=True) | Q(status='')
        ).count()
        
        # Fallback: if no active status, assume all are active
        if active_employees == 0 and total_employees > 0:
            active_employees = total_employees
        
        approved_certificates = EligibilityRequest.objects.filter(
            status='approved',
            date_processed__year=current_year
        ).count()
        
        # ============================================
        # SECONDARY STATS
        # ============================================
        pending_requests = EligibilityRequest.objects.filter(
            status='pending'
        ).count()
        
        total_requests = EligibilityRequest.objects.count()
        processing_rate = round((approved_certificates / total_requests * 100), 1) if total_requests > 0 else 0
        
        # Calculate average processing time
        processed_requests = EligibilityRequest.objects.filter(
            status='approved',
            date_processed__isnull=False,
            date_submitted__isnull=False
        )
        
        avg_processing_days = 0
        if processed_requests.exists():
            total_days = 0
            count = 0
            for req in processed_requests:
                delta = req.date_processed - req.date_submitted
                total_days += delta.days
                count += 1
            avg_processing_days = round(total_days / count) if count > 0 else 0
        
        # This month's certifications - FIX: Use current_month variable
        this_month_certs = EligibilityRequest.objects.filter(
            status='approved',
            date_processed__year=current_year,
            date_processed__month=current_month
        ).count()
        
        print(f"📅 This month (month {current_month}): {this_month_certs} certificates")
        print(f"   Query: year={current_year}, month={current_month}, status='approved'")
        
        # ============================================
        # CHART 1: Certifications by Month
        # ============================================
        certifications_by_month = []
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        for month_num in range(1, 13):
            elective_count = EligibilityRequest.objects.filter(
                status='approved',
                date_processed__year=current_year,
                date_processed__month=month_num,
                position_type='elective'
            ).count()
            
            appointive_count = EligibilityRequest.objects.filter(
                status='approved',
                date_processed__year=current_year,
                date_processed__month=month_num,
                position_type='appointive'
            ).count()
            
            certifications_by_month.append({
                'month': months[month_num - 1],
                'elective': elective_count,
                'appointive': appointive_count
            })

        
        # ============================================
        # CHART 2: Top Performing Barangays
        # ============================================
        barangay_stats = []
        
        for barangay in Barangay.objects.all():
            total_submissions = RequirementSubmission.objects.filter(
                barangay=barangay
            ).count()
            
            # Only include barangays with submissions
            if total_submissions > 0:
                approved = RequirementSubmission.objects.filter(
                    barangay=barangay,
                    status='approved'
                ).count()
                
                accomplished = RequirementSubmission.objects.filter(
                    barangay=barangay,
                    status='accomplished'
                ).count()
                
                # Compliance = completed tasks / total submissions
                completed = approved + accomplished
                compliance_rate = round((completed / total_submissions) * 100, 1)
                
                barangay_stats.append({
                    'name': barangay.name,
                    'total_submissions': total_submissions,
                    'approved': approved,
                    'accomplished': accomplished,
                    'compliance_rate': compliance_rate
                })
        
        # Sort by compliance rate, take top 10
        top_barangays = sorted(
            barangay_stats, 
            key=lambda x: (x['compliance_rate'], x['accomplished']), 
            reverse=True
        )[:10]

        monitoring_count = CategorizedFile.objects.filter(
            category__name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually'],
            is_archived=False
        ).count()
        
        certification_count = CategorizedFile.objects.filter(
            category__name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures'],
            is_archived=False
        ).count()
        
        pending_count = User.objects.filter(
            userprofile__is_approved=False
        ).count()
        
        pending_applications_count = EligibilityRequest.objects.filter(
            status='pending',
            archived=False
        ).count()
            
        # ============================================
        # BUILD CONTEXT
        # ============================================
        context = {
            'total_employees': total_employees,
            'active_employees': active_employees,
            'approved_certificates': approved_certificates,
            'current_year': current_year,
            'certifications_by_month': certifications_by_month,
            'top_barangays': top_barangays,
            # Secondary stats
            'pending_requests': pending_requests,
            'processing_rate': processing_rate,
            'avg_processing_days': avg_processing_days,
            'this_month_certs': this_month_certs,
            'monitoring_count': monitoring_count,
            'certification_count': certification_count,
            'pending_count': pending_count,
            'pending_applications_count': pending_applications_count,
        }
        
        print(f"✅ Dashboard loaded: Employees={total_employees}, "
              f"Certs={approved_certificates}, "
              f"Pending={pending_requests}, "
              f"This Month={this_month_certs}, "
              f"Barangays={len(top_barangays)}")
        
        return render(request, 'dashboard.html', context)
        
    except Exception as e:
        print(f"❌ Dashboard error: {e}")
        import traceback
        traceback.print_exc()
        
        # Safe fallback context
        context = {
            'total_employees': 0,
            'active_employees': 0,
            'approved_certificates': 0,
            'current_year': timezone.now().year,
            'certifications_by_month': [],
            'top_barangays': [],
            'pending_requests': 0,
            'processing_rate': 0,
            'avg_processing_days': 0,
            'this_month_certs': 0,
        }
        return render(request, 'dashboard.html', context)


# ===============================================
# API ENDPOINTS - Real-Time Data Refresh
# ===============================================

@login_required
@require_http_methods(["GET"])
def refresh_analytics(request):
    """
    API: Refresh stats cards data
    Returns: total_employees, active_employees, approved_certificates, timestamp
    """
    try:
        from datetime import timedelta
        current_year = timezone.now().year
        current_month = timezone.now().month
        
        total_employees = Employee.objects.count()
        
        active_employees = Employee.objects.filter(
            Q(status='active') | Q(status__isnull=True) | Q(status='')
        ).count()
        
        if active_employees == 0 and total_employees > 0:
            active_employees = total_employees
        
        approved_certificates = EligibilityRequest.objects.filter(
            status='approved',
            date_processed__year=current_year
        ).count()
        
        # Additional metrics for secondary stats
        pending_requests = EligibilityRequest.objects.filter(
            status='pending'
        ).count()
        
        total_requests = EligibilityRequest.objects.count()
        processing_rate = round((approved_certificates / total_requests * 100), 1) if total_requests > 0 else 0
        
        # Calculate average processing time
        processed_requests = EligibilityRequest.objects.filter(
            status='approved',
            date_processed__isnull=False
        )
        
        avg_processing_days = 0
        if processed_requests.exists():
            total_days = 0
            count = 0
            for req in processed_requests:
                if req.date_submitted and req.date_processed:
                    delta = req.date_processed - req.date_submitted
                    total_days += delta.days
                    count += 1
            avg_processing_days = round(total_days / count) if count > 0 else 0
        
        # This month's certifications
        this_month_certs = EligibilityRequest.objects.filter(
            status='approved',
            date_processed__year=current_year,
            date_processed__month=current_month
        ).count()
        
        return JsonResponse({
            'success': True,
            'data': {
                'total_employees': total_employees,
                'active_employees': active_employees,
                'approved_certificates': approved_certificates,
                'pending_requests': pending_requests,
                'processing_rate': processing_rate,
                'avg_processing_days': avg_processing_days,
                'this_month_certs': this_month_certs,
                'timestamp': timezone.now().strftime('%B %d, %Y • %I:%M %p')
            }
        })
    except Exception as e:
        print(f"❌ Error refreshing analytics: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["GET"])
def certifications_data(request):
    """
    API: Get monthly certification data (elective vs appointive)
    Returns: labels (months), elective counts, appointive counts
    """
    try:
        current_year = timezone.now().year
        months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                  'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        elective_data = []
        appointive_data = []
        
        for month_num in range(1, 13):
            elective = EligibilityRequest.objects.filter(
                status='approved',
                date_processed__year=current_year,
                date_processed__month=month_num,
                position_type='elective'
            ).count()
            
            appointive = EligibilityRequest.objects.filter(
                status='approved',
                date_processed__year=current_year,
                date_processed__month=month_num,
                position_type='appointive'
            ).count()
            
            elective_data.append(elective)
            appointive_data.append(appointive)
        
        return JsonResponse({
            'success': True,
            'certifications': {
                'labels': months,
                'elective': elective_data,
                'appointive': appointive_data
            }
        })
    except Exception as e:
        print(f"❌ Error fetching certifications: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
@login_required
@require_http_methods(["GET"])
def barangays_data(request):
    """
    API: Get top performing barangays by compliance rate
    Returns: labels (names), compliance rates, detailed stats
    
    ✅ FIXED: Now checks if ANY submissions exist before calculating
    """
    try:
        print(f"\n{'='*70}")
        print(f"📊 BARANGAYS DATA API - START")
        print(f"{'='*70}")
        
        # ✅ CRITICAL FIX: Check if ANY RequirementSubmissions exist
        total_submissions_in_db = RequirementSubmission.objects.count()
        print(f"Total RequirementSubmissions in database: {total_submissions_in_db}")
        
        if total_submissions_in_db == 0:
            print(f"⚠️ NO SUBMISSIONS FOUND - Returning empty data")
            print(f"{'='*70}\n")
            return JsonResponse({
                'success': True,
                'barangays': {
                    'labels': [],
                    'data': [],
                    'details': []
                },
                'message': 'No submission data available yet'
            })
        
        barangay_stats = []
        
        for barangay in Barangay.objects.all():
            total_submissions = RequirementSubmission.objects.filter(
                barangay=barangay
            ).count()
            
            print(f"\n🏘️ {barangay.name}: {total_submissions} total submissions")
            
            # ✅ FIXED: Skip barangays with NO submissions
            if total_submissions == 0:
                print(f"   ⏭️ Skipping - no submissions")
                continue
            
            # Only include barangays with submissions
            approved = RequirementSubmission.objects.filter(
                barangay=barangay,
                status='approved'
            ).count()
            
            accomplished = RequirementSubmission.objects.filter(
                barangay=barangay,
                status='accomplished'
            ).count()
            
            # Compliance = completed tasks / total submissions
            completed = approved + accomplished
            compliance_rate = round((completed / total_submissions) * 100, 1)
            
            print(f"   Approved: {approved}")
            print(f"   Accomplished: {accomplished}")
            print(f"   Compliance Rate: {compliance_rate}%")
            
            barangay_stats.append({
                'name': barangay.name,
                'total': total_submissions,
                'approved': approved,
                'accomplished': accomplished,
                'compliance_rate': compliance_rate
            })
        
        # Sort by compliance rate and take top 10
        top_barangays = sorted(
            barangay_stats,
            key=lambda x: (x['compliance_rate'], x['accomplished']),
            reverse=True
        )[:10]
        
        print(f"\n📈 Top {len(top_barangays)} performing barangays:")
        for i, b in enumerate(top_barangays, 1):
            print(f"   {i}. {b['name']}: {b['compliance_rate']}%")
        
        print(f"{'='*70}\n")
        
        return JsonResponse({
            'success': True,
            'barangays': {
                'labels': [b['name'] for b in top_barangays],
                'data': [b['compliance_rate'] for b in top_barangays],
                'details': top_barangays
            }
        })
    except Exception as e:
        print(f"❌ Error fetching barangays: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
#--------------END OF ANALYTICS DASHBOARD VIEW--------------#


    
#----------------CATEGORIZATION VIEW----------------#
from PIL import Image, ImageOps
import io
from django.core.files.uploadedfile import InMemoryUploadedFile

def process_signature_image(uploaded_file):
    """
    Process signature image to ensure it has white background
    and black signature (fixes black signature display issue)
    """
    try:
        print(f"🖊️ Processing signature: {uploaded_file.name}")
        print(f"   Original size: {uploaded_file.size} bytes")
        
        # Store original filename
        original_name = uploaded_file.name
        
        # Read the uploaded file
        uploaded_file.seek(0)
        image = Image.open(uploaded_file)
        
        print(f"   Image mode: {image.mode}")
        print(f"   Image size: {image.size}")
        
        # Convert to RGBA first if needed
        if image.mode != 'RGBA':
            image = image.convert('RGBA')
        
        # Create a white background
        white_bg = Image.new('RGB', image.size, (255, 255, 255))
        
        # Paste the signature onto white background
        # This converts transparent areas to white
        white_bg.paste(image, mask=image.split()[-1])  # Use alpha channel as mask
        
        # Save to BytesIO
        output = io.BytesIO()
        white_bg.save(output, format='PNG', quality=95)
        output.seek(0)
        
        processed_size = output.getbuffer().nbytes
        print(f"   ✓ Processed size: {processed_size} bytes")
        
        # Create new InMemoryUploadedFile with ORIGINAL FILENAME
        processed_file = InMemoryUploadedFile(
            output,
            'ImageField',
            original_name,  # ✅ Keep the original filename
            'image/png',
            processed_size,
            None
        )
        
        return processed_file
        
    except Exception as e:
        print(f"   ⚠️ Error processing signature: {e}")
        import traceback
        print(traceback.format_exc())
        # Return original file if processing fails
        uploaded_file.seek(0)
        return uploaded_file

from .models import send_certificate_notification_async
@require_http_methods(["POST"])
def submit_eligibility_request(request):
    """
    FIXED: Handle eligibility form submission
    Email is sent ONLY after successful database save
    """
    try:
        print(f"\n{'='*80}")
        print(f"📋 NEW ELIGIBILITY REQUEST SUBMISSION")
        print(f"{'='*80}")
        
        # ============================================
        # STEP 1: EXTRACT AND VALIDATE FORM DATA
        # ============================================
        last_name = request.POST.get('last_name', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        middle_initial = request.POST.get('middle_initial', '').strip()
        barangay = request.POST.get('barangay', '').strip()
        email = request.POST.get('email', '').strip()
        position_type = request.POST.get('position_type', '').strip()
        certifier = request.POST.get('certifier', '').strip()
        
        print(f"📋 Applicant: {first_name} {last_name}")
        print(f"📧 Email: {email}")
        print(f"🏘️ Barangay: {barangay}")
        print(f"💼 Position Type: {position_type}")
        
        # Basic validation
        if not all([last_name, first_name, email, barangay, position_type, certifier]):
            return JsonResponse({
                'success': False,
                'error': 'All required fields must be filled'
            }, status=400)
        
        # Validate certifier choice
        from app.models import EligibilityRequest
        valid_certifiers = [choice[0] for choice in EligibilityRequest.CERTIFIER_CHOICES]
        if certifier not in valid_certifiers:
            return JsonResponse({
                'success': False,
                'error': f'Invalid certifier selection'
            }, status=400)
        
        # Get uploaded files
        id_front = request.FILES.get('id_front')
        id_back = request.FILES.get('id_back')
        signature = request.FILES.get('signature')
        
        if not all([id_front, id_back, signature]):
            return JsonResponse({
                'success': False,
                'error': 'All files (ID front, ID back, signature) are required'
            }, status=400)
        
        print(f"📎 Files received: ID Front, ID Back, Signature")
        
        # ============================================
        # STEP 2: CREATE ELIGIBILITY REQUEST OBJECT
        # ============================================
        eligibility_request = EligibilityRequest(
            first_name=first_name,
            last_name=last_name,
            middle_initial=middle_initial if middle_initial else None,
            barangay=barangay,
            email=email,
            position_type=position_type,
            certifier=certifier,
            status='pending',
            date_submitted=timezone.now()
        )
        
        # ============================================
        # STEP 3: ADD POSITION-SPECIFIC DATA
        # ============================================
        if position_type == 'appointive':
            # ✅ CRITICAL FIX: Clean years fields to remove " yrs" suffix
            years_in_service_raw = request.POST.get('years_in_service', '').strip()
            pb_years_service_raw = request.POST.get('pb_years_service', '').strip()
            
            # Remove " yrs" suffix if present
            years_in_service_clean = years_in_service_raw.replace(' yrs', '').replace('yrs', '').strip()
            pb_years_service_clean = pb_years_service_raw.replace(' yrs', '').replace('yrs', '').strip()
            
            print(f"🔧 Years cleaning:")
            print(f"   Raw: '{years_in_service_raw}' → Clean: '{years_in_service_clean}'")
            print(f"   Raw: '{pb_years_service_raw}' → Clean: '{pb_years_service_clean}'")
            
           # NEW CODE (CORRECT):
            from decimal import Decimal, InvalidOperation  # ← Add this import

            try:
                if years_in_service_clean:
                    # ✅ Use Decimal, not float!
                    years_in_service = Decimal(years_in_service_clean).quantize(Decimal('0.1'))
                else:
                    years_in_service = None
                    
                if pb_years_service_clean:
                    # ✅ Use Decimal, not float!
                    pb_years_service = Decimal(pb_years_service_clean).quantize(Decimal('0.1'))
                else:
                    pb_years_service = None
                    
                print(f"   ✅ Converted to Decimal: years_in_service={years_in_service}, pb_years_service={pb_years_service}")
                print(f"   ✅ Types: {type(years_in_service)}, {type(pb_years_service)}")
                
            except (ValueError, InvalidOperation) as e:
                print(f"   ❌ Conversion error: {e}")
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid years format. Please use numbers only (e.g., 0.1, 2.5, 10.0)'
                }, status=400)
            
            eligibility_request.appointing_authority = request.POST.get('appointing_authority', '')
            eligibility_request.appointment_from = request.POST.get('appointment_from')
            eligibility_request.appointment_to = request.POST.get('appointment_to')
            eligibility_request.years_in_service = years_in_service
            eligibility_request.appointing_punong_barangay = request.POST.get('appointing_punong_barangay', '')
            eligibility_request.pb_date_elected = request.POST.get('pb_date_elected')
            eligibility_request.pb_years_service = pb_years_service
            
        elif position_type == 'elective':
            eligibility_request.position_held = request.POST.get('position_held', '')
            eligibility_request.election_from = request.POST.get('election_from')
            eligibility_request.election_to = request.POST.get('election_to')
            eligibility_request.term_office = request.POST.get('term_office', '')
            eligibility_request.completed_term = request.POST.get('completed_term', '')
            eligibility_request.incomplete_reason = request.POST.get('incomplete_reason', '')
        
        # ============================================
        # STEP 4: PROCESS AND SAVE FILES
        # ============================================
        from app.views import save_categorized_eligibility_file, process_signature_image
        
        print(f"\n📁 Processing files...")
        
        # Process ID Front
        id_front_path = save_categorized_eligibility_file(
            file=id_front,
            category='ids',
            user_name=f"{first_name}_{last_name}",
            file_type='id_front',
            request_id=None  # Will update after save
        )
        
        # Process ID Back
        id_back_path = save_categorized_eligibility_file(
            file=id_back,
            category='ids',
            user_name=f"{first_name}_{last_name}",
            file_type='id_back',
            request_id=None
        )
        
        # Process Signature (with white background fix)
        processed_signature = process_signature_image(signature)
        signature_path = save_categorized_eligibility_file(
            file=processed_signature,
            category='signatures',
            user_name=f"{first_name}_{last_name}",
            file_type='signature',
            request_id=None
        )
        
        eligibility_request.id_front = id_front_path
        eligibility_request.id_back = id_back_path
        eligibility_request.signature = signature_path
        
        print(f"✅ Files processed successfully")
        
        # ============================================
        # STEP 5: VALIDATE MODEL (BEFORE SAVING)
        # ============================================
        print(f"\n🔍 Validating model...")
        try:
            eligibility_request.full_clean()  # ✅ CRITICAL: Validate BEFORE save
            print(f"✅ Validation passed")
        except ValidationError as ve:
            print(f"❌ Validation failed: {ve}")
            return JsonResponse({
                'success': False,
                'error': f'Validation failed: {str(ve)}'
            }, status=400)
        
        # ============================================
        # STEP 6: SAVE TO DATABASE
        # ============================================
        print(f"\n💾 Saving to database...")
        try:
            eligibility_request.save()
            print(f"✅ Saved to database - ID: {eligibility_request.id}")
        except Exception as save_error:
            print(f"❌ Database save failed: {save_error}")
            return JsonResponse({
                'success': False,
                'error': f'Database error: {str(save_error)}'
            }, status=400)
        
        # ============================================
        # STEP 7: SEND EMAIL (ONLY AFTER SUCCESSFUL SAVE)
        # ============================================
        print(f"\n📧 Sending confirmation email...")
        email_sent = False
        try:
            send_mail(
                subject=f'Eligibility Certificate Request - {position_type.title()}',
                message=f'''Dear {first_name} {last_name},

Your eligibility certificate request has been successfully submitted.

Request Details:
- Request ID: EC-{timezone.now().year}-{eligibility_request.id:05d}
- Position Type: {position_type.title()}
- Barangay: {barangay}
- Certifier: {certifier}
- Submitted: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}

We will process your request and send the certificate to this email address once completed.

Thank you,
DILG Lucena - Barangay Certification System
''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
            email_sent = True
            print(f"✅ Confirmation email sent to {email}")
            
        except Exception as email_error:
            # ✅ Email failure should NOT fail the entire request
            print(f"⚠️ Email sending failed: {str(email_error)}")
            print(f"   But request was saved successfully!")
            # Don't return error - request was saved successfully
        
        # ============================================
        # STEP 8: RETURN SUCCESS
        # ============================================
        reference_number = f"EC-{timezone.now().year}-{eligibility_request.id:05d}"
        
        print(f"\n{'='*80}")
        print(f"✅ SUBMISSION COMPLETE")
        print(f"   Request ID: {eligibility_request.id}")
        print(f"   Reference: {reference_number}")
        print(f"   Email sent: {'Yes' if email_sent else 'No (but saved successfully)'}")
        print(f"{'='*80}\n")
        
        return JsonResponse({
            'success': True,
            'message': 'Application submitted successfully!' + (' Check your email for confirmation.' if email_sent else ''),
            'id_number': reference_number,
            'request_id': eligibility_request.id,
            'email_sent': email_sent
        })
        
    except Exception as e:
        print(f"\n{'='*80}")
        print(f"❌ SUBMISSION ERROR")
        print(f"{'='*80}")
        print(f"Error: {str(e)}")
        print(traceback.format_exc())
        print(f"{'='*80}\n")
        
        return JsonResponse({
            'success': False,
            'error': f'Submission failed: {str(e)}'
        }, status=400)



def smart_categorize_file(file, file_type_hint):
    """
    🧠 SMART CATEGORIZATION ENGINE
    Analyzes file content to determine correct folder
    
    Returns: 'appointive_certificates', 'elective_certificates', 'ids', or 'signatures'
    """
    try:
        print(f"🔍 Analyzing: {file.name}")
        
        # Force categorization based on file type hint
        if file_type_hint in ['id_front', 'id_back']:
            print(f"   ✅ Category: ids (based on file type)")
            return 'ids'
        
        if file_type_hint == 'signature':
            print(f"   ✅ Category: signatures (based on file type)")
            return 'signatures'
        
        # For other files, analyze content
        file_extension = file.name.lower().split('.')[-1]
        text_content = ""
        
        # Extract text based on file type
        if file_extension == 'pdf':
            text_content = extract_text_from_pdf(file)
        elif file_extension in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
            text_content = extract_text_from_image(file)
        
        # If we have text, analyze it
        if text_content:
            category = analyze_text_for_category(text_content, file.name)
            print(f"   ✅ Category: {category} (based on content analysis)")
            return category
        
        # Fallback to filename analysis
        category = categorize_by_filename(file.name)
        print(f"   ✅ Category: {category} (based on filename)")
        return category
        
    except Exception as e:
        print(f"   ⚠️ Categorization error: {e}")
        return 'ids'  # Default fallback


def extract_text_from_pdf(file):
    """Extract text from PDF file"""
    try:
        file.seek(0)
        pdf_reader = PyPDF2.PdfReader(file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"   ⚠️ PDF extraction error: {e}")
        return ""


def extract_text_from_image(file):
    """Extract text from image using OCR"""
    try:
        file.seek(0)
        image = Image.open(file)
        text = pytesseract.image_to_string(image)
        return text
    except Exception as e:
        print(f"   ⚠️ OCR extraction error: {e}")
        return ""


def analyze_text_for_category(text, filename):
    """
    Analyze extracted text to determine certificate type
    """
    text_lower = text.lower()
    filename_lower = filename.lower()
    
    # Keywords for different certificate types
    appointive_keywords = [
        'appointive official',
        'appointive position',
        'date of appointment',
        'appointing authority',
        'appointing punong barangay',
        'barangay secretary',
        'barangay treasurer',
        'appointment',
        'csc-erpo boe form 1(b)'
    ]
    
    elective_keywords = [
        'elective official',
        'elective position',
        'date of election',
        'term of office',
        'punong barangay',
        'sanguniang barangay member',
        'elected',
        'election',
        'csc-erpo boe form 1(a)'
    ]
    
    id_keywords = [
        'identification',
        'id card',
        'government issued id',
        'driver',
        'passport',
        'sss',
        'philhealth',
        'tin',
        'voter'
    ]
    
    signature_keywords = [
        'signature',
        'sign here',
        'e-signature'
    ]
    
    # Count keyword matches
    appointive_score = sum(1 for keyword in appointive_keywords if keyword in text_lower)
    elective_score = sum(1 for keyword in elective_keywords if keyword in text_lower)
    id_score = sum(1 for keyword in id_keywords if keyword in text_lower)
    signature_score = sum(1 for keyword in signature_keywords if keyword in text_lower)
    
    # Check filename hints
    if any(word in filename_lower for word in ['id_front', 'id_back', 'identification']):
        id_score += 5
    if 'signature' in filename_lower or 'sign' in filename_lower:
        signature_score += 5
    if 'appointive' in filename_lower:
        appointive_score += 3
    if 'elective' in filename_lower:
        elective_score += 3
    
    print(f"    Scores - Appointive:{appointive_score}, Elective:{elective_score}, ID:{id_score}, Signature:{signature_score}")
    
    # Determine category based on highest score
    scores = {
        'appointive_certificates': appointive_score,
        'elective_certificates': elective_score,
        'ids': id_score,
        'signatures': signature_score
    }
    
    max_score = max(scores.values())
    
    if max_score == 0:
        return categorize_by_filename(filename)
    
    # Return category with highest score
    for category, score in scores.items():
        if score == max_score:
            return category
    
    return 'ids'


def categorize_by_filename(filename):
    """Fallback categorization based on filename"""
    filename_lower = filename.lower()
    
    if any(word in filename_lower for word in ['id_front', 'id_back', 'identification', '_id_']):
        return 'ids'
    elif any(word in filename_lower for word in ['signature', 'sign', 'esign']):
        return 'signatures'
    elif 'appointive' in filename_lower:
        return 'appointive_certificates'
    elif 'elective' in filename_lower:
        return 'elective_certificates'
    else:
        return 'ids'
    
def save_categorized_eligibility_file(file, category, user_name, file_type, request_id):
    """
    Save file to storage AND create CategorizedFile database entry
    FIXED: Prevents duplicates by checking if file already exists
    """
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    from .models import FileCategory, CategorizedFile, EligibilityRequest

    file_extension = os.path.splitext(file.name)[1]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{user_name}_{file_type}_{timestamp}{file_extension}"

    year = datetime.now().strftime('%Y')
    month = datetime.now().strftime('%m')
    folder_path = f"certification_files/{category}/{year}/{month}"

    # Ensure .gitkeep exists
    gitkeep_path = f"{folder_path}/.gitkeep"
    if not default_storage.exists(gitkeep_path):
        default_storage.save(gitkeep_path, ContentFile(b''))
    
    # Save file
    file_path = os.path.join(folder_path, filename)
    file.seek(0)  
    path = default_storage.save(file_path, ContentFile(file.read()))
    print(f"    ✓ Saved to: {path}")
    
    # ✅ FIX: Check if CategorizedFile already exists for this file
    existing_file = CategorizedFile.objects.filter(
        file=path,
        original_filename=filename
    ).first()
    
    if existing_file:
        print(f"    ⚠️ File already exists in database (ID: {existing_file.id}), skipping duplicate")
        return path
    
    # Create/get FileCategory
    file_category, _ = FileCategory.objects.get_or_create(
        name=category,
        defaults={
            'display_name': category.replace('_', ' ').title(),
            'folder_path': f'certification_files/{category}/'
        }
    )
    
    # Get the eligibility request
    try:
        eligibility_request = EligibilityRequest.objects.get(id=request_id) if request_id else None
    except EligibilityRequest.DoesNotExist:
        eligibility_request = None
    
    # Create CategorizedFile entry (only if doesn't exist)
    categorized = CategorizedFile.objects.create(
        file=path,
        original_filename=filename,
        file_type='image' if file_extension.lower() in ['.jpg', '.jpeg', '.png', '.gif'] else 'document',
        file_size=file.size if hasattr(file, 'size') else 0,
        mime_type=getattr(file, 'content_type', 'application/octet-stream'),
        category=file_category,
        source='eligibility',
        detected_content=f'ID {file_type}' if 'id' in file_type else 'Signature',
        eligibility_request=eligibility_request,
        tags=f"{user_name}, {file_type}, Eligibility Request {request_id}"
    )
    
    # Update category file count
    file_category.update_file_count()
    
    print(f"    ✓ Created NEW CategorizedFile ID: {categorized.id}")
    
    return path


# Add this helper function for generating certificates (when admin approves)
def generate_certificate_pdf(eligibility_request):
    """
    Generate certificate PDF with DILG logos and proper formatting
    Fixed: Logo placement, Director name, Table formatting
    """
    try:
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        from io import BytesIO
        import os
        from django.conf import settings
        
        print(f"\n{'='*70}")
        print(f"📄 GENERATING OFFICIAL DILG CERTIFICATE")
        print(f"{'='*70}")
        print(f"Request ID: {eligibility_request.id}")
        print(f"Name: {eligibility_request.full_name}")
        print(f"Position: {eligibility_request.position_type}")
        print(f"Barangay: {eligibility_request.barangay}")
        
        # Determine folder
        if eligibility_request.position_type == 'appointive':
            folder = 'appointive_certificates'
            form_ref = "CSC-ERPO BOE Form 1(b). April 2012"
            position_label = "(Appointive Official)"
        else:
            folder = 'elective_certificates'
            form_ref = "CSC-ERPO BOE Form 1(a) (Revised, June 2017)"
            position_label = "(Elective Official)"
        
        # Get or create category
        from .models import FileCategory, CategorizedFile
        category, _ = FileCategory.objects.get_or_create(
            name=folder,
            defaults={
                'display_name': folder.replace('_', ' ').title(),
                'folder_path': f'certification_files/{folder}/',
            }
        )
        
        # Create PDF
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # === OUTER BORDER ===
        c.setStrokeColor(colors.HexColor('#1A237E'))
        c.setLineWidth(2)
        c.rect(0.4*inch, 0.4*inch, width - 0.8*inch, height - 0.8*inch)
        
        # === INNER BORDER ===
        c.setLineWidth(0.5)
        c.rect(0.5*inch, 0.5*inch, width - 1*inch, height - 1*inch)
        
        # === DILG LOGOS (IMPROVED) ===
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'Pictures', 'logo1.png')
        
        if os.path.exists(logo_path):
            try:
                logo_size = 0.7*inch  # Increased size
                logo_y = height - 1.3*inch
                
                # Left logo
                c.drawImage(logo_path, 0.75*inch, logo_y, 
                           width=logo_size, height=logo_size, 
                           preserveAspectRatio=True, mask='auto')
                
                # Right logo
                c.drawImage(logo_path, width - 0.75*inch - logo_size, logo_y, 
                           width=logo_size, height=logo_size, 
                           preserveAspectRatio=True, mask='auto')
                print(f"✓ DILG logos added")
            except Exception as logo_err:
                print(f"⚠️ Logo error: {logo_err}")
        else:
            print(f"⚠️ Logo not found at: {logo_path}")
        
        # === HEADER ===
        y_pos = height - 1*inch
        
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "Republic of the Philippines")
        
        y_pos -= 0.2*inch
        c.setFillColor(colors.HexColor('#1A237E'))
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width/2, y_pos, "DEPARTMENT OF THE INTERIOR AND")
        y_pos -= 0.18*inch
        c.drawCentredString(width/2, y_pos, "LOCAL GOVERNMENT")
        
        y_pos -= 0.2*inch
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "REGION IV-A CALABARZON")
        
        y_pos -= 0.15*inch
        c.drawCentredString(width/2, y_pos, "CITY OF LUCENA")
        
        # Form reference (right aligned)
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.gray)
        c.drawRightString(width - 0.6*inch, y_pos - 0.3*inch, form_ref)
        
        # === HORIZONTAL LINE ===
        y_pos -= 0.5*inch
        c.setStrokeColor(colors.black)
        c.setLineWidth(1)
        c.line(0.75*inch, y_pos, width - 0.75*inch, y_pos)
        
        # === TITLE SECTION ===
        y_pos -= 0.5*inch
        c.setFillColor(colors.HexColor('#1A237E'))
        c.setFont("Helvetica-Bold", 18)
        c.drawCentredString(width/2, y_pos, "CERTIFICATION")
        
        y_pos -= 0.25*inch
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 10)
        c.drawCentredString(width/2, y_pos, "on Services Rendered in the Barangay*")
        
        y_pos -= 0.2*inch
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width/2, y_pos, position_label)
        
        # === HORIZONTAL LINE ===
        y_pos -= 0.3*inch
        c.line(0.75*inch, y_pos, width - 0.75*inch, y_pos)
        
        # === BODY TEXT ===
        y_pos -= 0.4*inch
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        
        text_line = f"This is to certify that "
        c.drawString(0.75*inch, y_pos, text_line)
        
        name_x = 0.75*inch + c.stringWidth(text_line, "Helvetica", 10)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(name_x, y_pos, eligibility_request.full_name.upper())
        
        after_name_x = name_x + c.stringWidth(eligibility_request.full_name.upper(), "Helvetica-Bold", 10)
        c.setFont("Helvetica", 10)
        c.drawString(after_name_x, y_pos, " has rendered services in")
        
        y_pos -= 0.18*inch
        barangay_text = f"Barangay {eligibility_request.barangay}, with the following details:"
        c.drawString(0.75*inch, y_pos, barangay_text)
        
        # === TABLE (FIXED FORMATTING) ===
        y_pos -= 0.5*inch
        
        if eligibility_request.position_type == 'appointive':
            # Appointive table with proper spacing
            table_data = [
                ['Position\nHeld', 
                 'Date of\nAppointment', 
                 'Inclusive Dates\nFrom', 
                 'Inclusive Dates\nTo',
                 'No. of Years\nServed', 
                 'Appointing Punong\nBarangay Name',
                 'Date Elected',
                 'Term of Office\n(years)'],
                [
                    'Barangay\nSecretary',
                    eligibility_request.appointment_from.strftime('%m/%d/%Y') if eligibility_request.appointment_from else 'N/A',
                    eligibility_request.appointment_from.strftime('%m/%d/%Y') if eligibility_request.appointment_from else 'N/A',
                    eligibility_request.appointment_to.strftime('%m/%d/%Y') if eligibility_request.appointment_to else 'N/A',
                    f"{float(eligibility_request.years_in_service)} yrs" if eligibility_request.years_in_service else '0.0 yrs',
                    eligibility_request.appointing_punong_barangay or 'N/A',
                    eligibility_request.pb_date_elected.strftime('%m/%d/%Y') if eligibility_request.pb_date_elected else 'N/A',
                    f"{float(eligibility_request.pb_years_service)} yrs" if eligibility_request.pb_years_service else '0.0 yrs'
                ]
            ]
            
            # Adjusted column widths for better spacing
            col_widths = [0.8*inch, 0.75*inch, 0.75*inch, 0.75*inch, 0.7*inch, 1.1*inch, 0.7*inch, 0.7*inch]
            row_heights = [0.6*inch, 0.5*inch]
            
        else:
            # Elective table - matching official format exactly
            table_data = [
                ['Position Held', 'Date of Election\n(mm/dd/yyyy)', 'Term of Office\n(no. of years)', 
                 'Inclusive Dates\nFrom (mm/dd/yyyy)', 'Inclusive Dates\nTo (mm/dd/yyyy)'],
                [
                    eligibility_request.position_held or 'Punong Barangay',
                    eligibility_request.election_from.strftime('%m/%d/%Y') if eligibility_request.election_from else '',
                    eligibility_request.term_office or 'November 2025 -\nNovember 2025',
                    eligibility_request.election_from.strftime('%m/%d/%Y') if eligibility_request.election_from else '',
                    eligibility_request.election_to.strftime('%m/%d/%Y') if eligibility_request.election_to else ''
                ]
            ]
            
            col_widths = [1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.3*inch]
            row_heights = [0.6*inch, 0.5*inch]
        
        # Create table with improved styling
        table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
        
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            
            # Borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            
            # Padding
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        table.wrapOn(c, width, height)
        table_height = table._height
        table.drawOn(c, 0.75*inch, y_pos - table_height)
        
        y_pos -= (table_height + 0.4*inch)
        
        # === COMPLETED TERM SECTION (ELECTIVE ONLY) ===
        if eligibility_request.position_type == 'elective':
            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(colors.black)
            c.drawString(0.75*inch, y_pos, "Completed Term of Office?")
            c.setFont("Helvetica", 9)
            c.drawString(2.5*inch, y_pos, "(Please check (√) appropriate box)")
            
            y_pos -= 0.25*inch
            
            completed_term = eligibility_request.completed_term
            checkbox_size = 0.12*inch
            
            # YES checkbox
            c.rect(0.95*inch, y_pos - checkbox_size/2, checkbox_size, checkbox_size)
            if completed_term and completed_term.lower() == 'yes':
                # Draw checkmark
                c.setLineWidth(2)
                c.line(0.97*inch, y_pos - checkbox_size/4, 
                       1.0*inch, y_pos - checkbox_size/1.5)
                c.line(1.0*inch, y_pos - checkbox_size/1.5,
                       1.05*inch, y_pos)
                c.setLineWidth(0.5)
            
            c.drawString(1.15*inch, y_pos - 0.05*inch, "YES")
            
            # NO checkbox
            c.rect(1.8*inch, y_pos - checkbox_size/2, checkbox_size, checkbox_size)
            if completed_term and completed_term.lower() == 'no':
                # Draw checkmark
                c.setLineWidth(2)
                c.line(1.82*inch, y_pos - checkbox_size/4, 
                       1.85*inch, y_pos - checkbox_size/1.5)
                c.line(1.85*inch, y_pos - checkbox_size/1.5,
                       1.90*inch, y_pos)
                c.setLineWidth(0.5)
            
            c.drawString(2.0*inch, y_pos - 0.05*inch, "NO. Specify total number of days not served")
            
            y_pos -= 0.25*inch
            
            # If NO was selected, show reason box
            if completed_term and completed_term.lower() == 'no':
                c.setFont("Helvetica-Bold", 9)
                c.drawString(0.95*inch, y_pos, "Reason for non-completion:")
                
                y_pos -= 0.25*inch
                
                # Draw reason box
                reason_text = eligibility_request.incomplete_reason or 'Not specified'
                c.setFont("Helvetica", 9)
                c.setFillColor(colors.HexColor('#f5f5f5'))
                c.rect(0.95*inch, y_pos - 0.35*inch, 5.5*inch, 0.4*inch, fill=1)
                
                c.setFillColor(colors.black)
                c.drawString(1.05*inch, y_pos - 0.15*inch, reason_text)
                
                y_pos -= 0.45*inch
            
            # ASSUMED checkbox (always show)
            c.rect(0.95*inch, y_pos - checkbox_size/2, checkbox_size, checkbox_size)
            c.drawString(1.15*inch, y_pos - 0.05*inch, "Assumed under rule on succession.")
            
            y_pos -= 0.35*inch
        
        # === FOOTER TEXT ===
        c.setFont("Helvetica", 9)
        c.setFillColor(colors.black)
        
        if eligibility_request.position_type == 'elective':
            # Elective footer text with CSC Resolutions
            footer_text = f"This Certification is issued in support of the evaluation/processing of the application of {eligibility_request.full_name.upper()}"
            c.drawString(0.75*inch, y_pos, footer_text)
            
            y_pos -= 0.15*inch
            footer_text2 = "for the grant of Barangay Official Eligibility pursuant to Republic Act No. 7160, in accordance"
            c.drawString(0.75*inch, y_pos, footer_text2)
            
            y_pos -= 0.15*inch
            footer_text3 = "with CSC Resolution No. 1200865 dated June 14, 2012 and CSC Resolution No."
            c.drawString(0.75*inch, y_pos, footer_text3)
            
            y_pos -= 0.15*inch
            footer_text4 = "1601257 dated November 21, 2016."
            c.drawString(0.75*inch, y_pos, footer_text4)
        else:
            # Appointive footer text
            footer_text = f"This Certification is issued in support of the evaluation/processing of the application of {eligibility_request.full_name.upper()}"
            c.drawString(0.75*inch, y_pos, footer_text)
            
            y_pos -= 0.15*inch
            footer_text2 = "for the grant of Barangay Official Eligibility pursuant to Republic Act No. 7160, in accordance"
            c.drawString(0.75*inch, y_pos, footer_text2)
            
            y_pos -= 0.15*inch
            footer_text3 = "with CSC Resolution No. 13 series of 2012."
            c.drawString(0.75*inch, y_pos, footer_text3)
        
        # === SIGNATURE SECTION ===
        y_pos -= 0.5*inch
        
        from django.utils import timezone
        date_text = f"Lucena City, Quezon, {timezone.now().strftime('%B %d, %Y')}."
        c.drawString(0.75*inch, y_pos, date_text)
        
        # === SIGNATURE SECTION (UPDATED DIRECTOR) ===
        y_pos -= 0.8*inch
        
        # Signature line (right side)
        sig_x = width - 2.8*inch
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(sig_x, y_pos, width - 0.75*inch, y_pos)
        
        # Director name and title
        y_pos -= 0.2*inch
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.black)
        c.drawCentredString(sig_x + 1.025*inch, y_pos, "LEANDRO SIPOY GIGANTOCA, CESE")
        
        y_pos -= 0.18*inch
        c.setFont("Helvetica", 9)
        c.drawCentredString(sig_x + 1.025*inch, y_pos, "OIC-HUC Director, Lucena City")
        
        # Save PDF
        c.showPage()
        c.save()
        
        pdf_data = buffer.getvalue()
        buffer.close()
        
        print(f"✓ PDF generated successfully ({len(pdf_data)} bytes)")
        
        # Create filename
        safe_name = eligibility_request.full_name.replace(' ', '_').replace('.', '')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{safe_name}_{eligibility_request.position_type}_Certificate_{timestamp}.pdf"
        
        # Save to storage
        file_path = f"certification_files/{folder}/{filename}"
        saved_path = default_storage.save(file_path, ContentFile(pdf_data))
        
        print(f"✓ PDF saved: {saved_path}")
        
        # Create CategorizedFile entry
        categorized = CategorizedFile.objects.create(
            file=saved_path,
            original_filename=filename,
            file_type='pdf',
            file_size=len(pdf_data),
            mime_type='application/pdf',
            category=category,
            source='eligibility',
            detected_content=f'{eligibility_request.get_position_type_display()} Certificate',
            eligibility_request=eligibility_request,
            uploaded_by=eligibility_request.approved_by,
            tags=f"{eligibility_request.full_name}, {eligibility_request.position_type}, Certificate, Approved"
        )
        
        print(f"✓ CategorizedFile created: ID {categorized.id}")
        category.update_file_count()
        print(f"✓ Updated folder file count: {category.file_count}")
        
        print(f"{'='*70}\n")
        
        return saved_path
        
    except Exception as e:
        print(f"\n{'='*70}")
        print(f"❌ CERTIFICATE GENERATION ERROR")
        print(f"{'='*70}")
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc())
        print(f"{'='*70}\n")
        return None


# Add these API endpoints for the certificate files page

@require_http_methods(["GET"])
def get_certificate_files_by_category(request, category):
    """
    FIXED: Get certificate files by category with REAL database IDs
    """
    try:
        from django.core.files.storage import default_storage
        from .models import CategorizedFile, FileCategory
        
        print(f"\n{'='*70}")
        print(f"📁 GET CERTIFICATE FILES BY CATEGORY")
        print(f"{'='*70}")
        print(f"Category requested: {category}")
        
        # Map category names to folder paths
        category_folders = {
            'certificates': ['appointive_certificates', 'elective_certificates'],
            'appointive_certificates': ['appointive_certificates'],
            'elective_certificates': ['elective_certificates'],
            'ids': ['ids'],
            'signatures': ['signatures']
        }
        
        if category not in category_folders:
            return JsonResponse({
                'success': False,
                'error': f'Invalid category: {category}'
            }, status=400)
        
        folders_to_scan = category_folders[category]
        all_files = []
        
        # Get files from CategorizedFile database
        for folder_name in folders_to_scan:
            print(f"\n🔍 Querying CategorizedFile for: {folder_name}")
            
            # Get or create category
            try:
                file_category = FileCategory.objects.get(name=folder_name)
            except FileCategory.DoesNotExist:
                print(f"   ⚠️ Category '{folder_name}' not found in database")
                continue
            
            # Query CategorizedFile with REAL IDs
            files = CategorizedFile.objects.filter(
                category=file_category
            ).select_related('uploaded_by', 'barangay')
            
            print(f"   ✓ Found {files.count()} files in database")
            
            for file_obj in files:
                try:
                    file_info = {
                        'id': file_obj.id,  # ✅ REAL DATABASE ID
                        'filename': file_obj.original_filename,
                        'file_url': file_obj.file.url if file_obj.file else '',
                        'file_type': file_obj.file_type,
                        'file_size': file_obj.file_size_mb,
                        'uploaded_at': file_obj.uploaded_at.strftime('%B %d, %Y %I:%M %p'),
                        'category': file_category.display_name,
                        'folder': folder_name,
                        'uploaded_by': (
                            file_obj.uploaded_by.get_full_name() 
                            if file_obj.uploaded_by 
                            else 'System'
                        ),
                        'barangay': (
                            file_obj.barangay.name 
                            if file_obj.barangay 
                            else None
                        )
                    }
                    
                    all_files.append(file_info)
                    print(f"    ✓ {file_obj.original_filename} (ID: {file_obj.id})")
                    
                except Exception as file_err:
                    print(f"    ✗ Error processing file {file_obj.id}: {file_err}")
                    continue
        
        print(f"\n📊 Total files found: {len(all_files)}")
        print(f"{'='*70}\n")
        
        return JsonResponse({
            'success': True,
            'files': all_files,
            'total_count': len(all_files),
            'category': category
        })
        
    except Exception as e:
        print(f"\n{'='*70}")
        print(f"❌ ERROR: {str(e)}")
        print(f"{'='*70}\n")
        import traceback
        print(traceback.format_exc())
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'category': category
        }, status=500)


def save_categorized_eligibility_file(file, category, user_name, file_type, request_id):
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    from .models import FileCategory, CategorizedFile

    file_extension = os.path.splitext(file.name)[1]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{user_name}_{file_type}_{timestamp}{file_extension}"

    year = datetime.now().strftime('%Y')
    month = datetime.now().strftime('%m')
    folder_path = f"certification_files/{category}/{year}/{month}"

    # Ensure .gitkeep exists
    gitkeep_path = f"{folder_path}/.gitkeep"
    if not default_storage.exists(gitkeep_path):
        default_storage.save(gitkeep_path, ContentFile(b''))
    
    # Save file
    file_path = os.path.join(folder_path, filename)
    file.seek(0)  
    path = default_storage.save(file_path, ContentFile(file.read()))
    print(f"    ✓ Saved to: {path}")
    
    # ✅ FIX: Create/get FileCategory and CategorizedFile
    file_category, _ = FileCategory.objects.get_or_create(
        name=category,  # 'ids' or 'signatures'
        defaults={
            'display_name': category.replace('_', ' ').title(),
            'folder_path': f'certification_files/{category}/'
        }
    )
    
    # Create CategorizedFile entry
    categorized = CategorizedFile.objects.create(
        file=path,
        original_filename=filename,
        file_type='image' if file_extension.lower() in ['.jpg', '.jpeg', '.png', '.gif'] else 'document',
        file_size=file.size,
        mime_type=file.content_type,
        category=file_category,
        source='eligibility',
        detected_content=f'ID {file_type}' if 'id' in file_type else 'Signature',
        tags=f"{user_name}, {file_type}, Eligibility Request {request_id}"
    )
    
    # Update category file count
    file_category.update_file_count()
    
    print(f"    ✓ Created CategorizedFile ID: {categorized.id}")
    
    return path

@require_http_methods(["GET"])
def debug_certificate_categories(request):
    """Debug view to see what's in the database"""
    from .models import FileCategory, CategorizedFile
    
    result = {
        'file_categories': [],
        'files_by_folder': {},
        'total_files': CategorizedFile.objects.count()
    }
    
    # Get all FileCategory objects
    for cat in FileCategory.objects.all():
        result['file_categories'].append({
            'id': cat.id,
            'name': cat.name,
            'display_name': cat.display_name,
            'file_count': CategorizedFile.objects.filter(category=cat).count()
        })
    
    # Get all CategorizedFile objects grouped by their source
    files = CategorizedFile.objects.select_related('category').all()
    
    for file_obj in files:
        folder = file_obj.category.name if file_obj.category else 'No Category'
        
        if folder not in result['files_by_folder']:
            result['files_by_folder'][folder] = []
        
        result['files_by_folder'][folder].append({
            'id': file_obj.id,
            'filename': file_obj.original_filename,
            'source': file_obj.source,
            'file_type': file_obj.file_type,
        })
    
    return JsonResponse(result, json_dumps_params={'indent': 2})


@require_http_methods(["GET"])
def debug_certificate_files(request):
    from django.core.files.storage import default_storage
    import os
    
    try:
        result = {
            'base_path': 'certification_files',
            'folders': {},
            'all_files_found': []
        }
        
        # Check each category folder
        for category in ['appointive_certificates', 'elective_certificates', 'ids', 'signatures']:
            folder_path = f"certification_files/{category}/"
            
            folder_info = {
                'exists': default_storage.exists(folder_path),
                'files_found': []
            }
            
            if folder_info['exists']:
                try:
                    # Recursively find all files
                    def scan_directory(path):
                        files_in_dir = []
                        try:
                            dirs, files = default_storage.listdir(path)
                            
                            # Add files in current directory
                            for f in files:
                                if not f.startswith('.'):
                                    full_path = os.path.join(path, f)
                                    files_in_dir.append({
                                        'path': full_path,
                                        'name': f,
                                        'url': default_storage.url(full_path),
                                        'size': default_storage.size(full_path)
                                    })
                            
                            # Recursively scan subdirectories
                            for d in dirs:
                                subdir_path = os.path.join(path, d)
                                files_in_dir.extend(scan_directory(subdir_path))
                        
                        except Exception as e:
                            print(f"Error scanning {path}: {e}")
                        
                        return files_in_dir
                    
                    folder_info['files_found'] = scan_directory(folder_path)
                    result['all_files_found'].extend(folder_info['files_found'])
                    
                except Exception as e:
                    folder_info['error'] = str(e)
            
            result['folders'][category] = folder_info
        
        # Summary
        result['summary'] = {
            'total_files': len(result['all_files_found']),
            'files_by_category': {
                cat: len(info['files_found']) 
                for cat, info in result['folders'].items()
            }
        }
        
        return JsonResponse(result, json_dumps_params={'indent': 2})
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

def test_certificate_setup(request):

    try:
        from django.core.files.storage import default_storage
        import os
        
        results = {
            'status': 'Testing Certificate Files Setup',
            'media_root': settings.MEDIA_ROOT if hasattr(settings, 'MEDIA_ROOT') else 'Not configured',
            'media_url': settings.MEDIA_URL if hasattr(settings, 'MEDIA_URL') else 'Not configured',
            'folders': {},
            'sample_structure': {}
        }
        
        # Check each category folder
        categories = ['appointive_certificates', 'elective_certificates', 'ids', 'signatures']
        
        for category in categories:
            folder_path = f'certification_files/{category}/'
            
            folder_info = {
                'exists': default_storage.exists(folder_path),
                'full_path': os.path.join(settings.MEDIA_ROOT, folder_path) if hasattr(settings, 'MEDIA_ROOT') else 'Unknown',
                'file_count': 0,
                'sample_files': []
            }
            
            if folder_info['exists']:
                try:
                    # Try to count files
                    year_dirs, direct_files = default_storage.listdir(folder_path)
                    folder_info['file_count'] = len(direct_files)
                    folder_info['sample_files'] = direct_files[:3]  # First 3 files
                    folder_info['year_folders'] = year_dirs
                except Exception as e:
                    folder_info['error'] = str(e)
            
            results['folders'][category] = folder_info
        
        # Show expected structure
        results['sample_structure'] = {
            'certification_files/': {
                'appointive_certificates/': {
                    '2024/': {
                        '11/': ['John_Doe_certificate_20241106.pdf']
                    }
                },
                'elective_certificates/': {
                    '2024/': {
                        '11/': ['Jane_Smith_certificate_20241106.pdf']
                    }
                },
                'ids/': {
                    '2024/': {
                        '11/': ['John_Doe_id_front_20241106.jpg', 'John_Doe_id_back_20241106.jpg']
                    }
                },
                'signatures/': {
                    '2024/': {
                        '11/': ['John_Doe_signature_20241106.png']
                    }
                }
            }
        }
        
        # Test URL patterns
        from django.urls import reverse, NoReverseMatch
        
        url_tests = {}
        for category in categories:
            try:
                url = reverse('get_certificate_files_by_category', kwargs={'category': category})
                url_tests[category] = {
                    'url': url,
                    'status': 'URL pattern exists ✓'
                }
            except NoReverseMatch:
                url_tests[category] = {
                    'status': 'URL pattern MISSING ✗',
                    'fix': f"Add path('api/certificate-files/category/<str:category>/', views.get_certificate_files_by_category) to urls.py"
                }
        
        results['url_patterns'] = url_tests
        
        return JsonResponse(results, json_dumps_params={'indent': 2})
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500, json_dumps_params={'indent': 2})




@login_required
@require_http_methods(["POST"])
def update_application_status(request):
    """Update status AND send email with rejection reason"""
    try:
        data = json.loads(request.body)
        request_id = data.get('id')
        new_status = data.get('status')
        rejection_reason = data.get('rejection_reason')  # Get rejection reason
        
        eligibility_request = get_object_or_404(EligibilityRequest, id=request_id)
        
        # Update status
        old_status = eligibility_request.status
        eligibility_request.status = new_status
        
        if new_status in ['approved', 'rejected']:
            eligibility_request.approved_by = request.user
            eligibility_request.date_processed = timezone.now()
        
        # ✅ Store rejection reason if provided
        if new_status == 'rejected' and rejection_reason:
            eligibility_request.rejection_reason = rejection_reason
        
        eligibility_request.save()
        
        print(f"\n🔄 Status changed for request #{request_id}: {old_status} → {new_status}")
        
        # Send email notification
        if new_status in ['approved', 'rejected']:
            print(f"   Sending email to applicant: {eligibility_request.email}")
            try:
                # ✅ Pass rejection_reason to email function
                send_email_task(
                    eligibility_request, 
                    new_status,
                    rejection_reason=rejection_reason  # Pass the reason
                )
                print(f"✅ Email sent successfully")
            except Exception as email_error:
                print(f"❌ Email error: {str(email_error)}")
        
        # Generate certificate when approved
        certificate_path = None
        if new_status == 'approved' and old_status != 'approved':
            print(f"\n✅ APPROVAL DETECTED - Generating certificate...")
            certificate_path = generate_certificate_pdf(eligibility_request)
        
        response_data = {
            'success': True,
            'message': f'Status updated to {new_status.capitalize()}',
            'new_status': new_status,
            'approved_by': request.user.get_full_name() if request.user.get_full_name() else request.user.username
        }
        
        if certificate_path:
            import os
            response_data['certificate_generated'] = True
            response_data['certificate_filename'] = os.path.basename(certificate_path)
        
        return JsonResponse(response_data)
        
    except Exception as e:
        import traceback
        print(f"❌ Error: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def generate_certificate_pdf(eligibility_request):
    """
    Generate certificate PDF - FIXED for elective incomplete term
    """
    try:
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        from io import BytesIO
        import os
        from django.conf import settings
        
        print(f"\n{'='*70}")
        print(f"📄 GENERATING CERTIFICATE")
        print(f"Position: {eligibility_request.position_type}")
        if eligibility_request.position_type == 'elective':
            print(f"Completed Term: {eligibility_request.completed_term}")
            print(f"Days Not Served: {eligibility_request.days_not_served}")
            print(f"Reason: {eligibility_request.incomplete_reason}")
        print(f"{'='*70}")
        
        # Determine folder based on position type AND completion status
        if eligibility_request.position_type == 'appointive':
            folder = 'appointive_certificates'
            form_ref = "CSC-ERPO BOE Form 1(b). April 2012"
            position_label = "(Appointive Official)"
        else:
            folder = 'elective_certificates'
            form_ref = "CSC-ERPO BOE Form 1(a) (Revised, June 2017)"
            position_label = "(Elective Official)"
        
        from .models import FileCategory, CategorizedFile
        category, _ = FileCategory.objects.get_or_create(
            name=folder,
            defaults={
                'display_name': folder.replace('_', ' ').title(),
                'folder_path': f'certification_files/{folder}/',
            }
        )
        
        # Create PDF
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # === BORDERS ===
        c.setStrokeColor(colors.HexColor('#1A237E'))
        c.setLineWidth(2)
        c.rect(0.4*inch, 0.4*inch, width - 0.8*inch, height - 0.8*inch)
        
        c.setLineWidth(0.5)
        c.rect(0.5*inch, 0.5*inch, width - 1*inch, height - 1*inch)
        
        # === LOGOS ===
        from reportlab.lib.utils import ImageReader
        
        # Try multiple possible logo paths
        possible_logo_paths = [
            os.path.join(settings.BASE_DIR, 'static', 'Pictures', 'logo1.png'),
            os.path.join(settings.BASE_DIR, 'app', 'static', 'Pictures', 'logo1.png'),
            os.path.join(settings.BASE_DIR, 'static', 'pictures', 'logo1.png'),
            os.path.join(settings.STATIC_ROOT, 'Pictures', 'logo1.png') if hasattr(settings, 'STATIC_ROOT') and settings.STATIC_ROOT else None,
        ]
        
        logo_path = None
        for path in possible_logo_paths:
            if path and os.path.exists(path):
                logo_path = path
                print(f"✓ Logo found at: {path}")
                break
        
        if logo_path:
            try:
                # Use ImageReader for better compatibility
                img = ImageReader(logo_path)
                logo_size = 0.7*inch
                logo_y = height - 1.3*inch
                
                # Draw left logo
                c.drawImage(img, 0.75*inch, logo_y, 
                           width=logo_size, height=logo_size, 
                           preserveAspectRatio=True, mask='auto')
                
                # Draw right logo (reload for second instance)
                img2 = ImageReader(logo_path)
                c.drawImage(img2, width - 0.75*inch - logo_size, logo_y, 
                           width=logo_size, height=logo_size, 
                           preserveAspectRatio=True, mask='auto')
                
                print(f"✓ Logos rendered successfully on both sides")
            except Exception as e:
                print(f"⚠️ Logo rendering error: {e}")
                import traceback
                print(traceback.format_exc())
        else:
            print(f"⚠️ Logo file not found. Searched:")
            for path in possible_logo_paths:
                if path:
                    print(f"   - {path}")
        
        # === HEADER ===
        y_pos = height - 1*inch
        
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "Republic of the Philippines")
        
        y_pos -= 0.2*inch
        c.setFillColor(colors.HexColor('#1A237E'))
        c.setFont("Helvetica-Bold", 11)
        c.drawCentredString(width/2, y_pos, "DEPARTMENT OF THE INTERIOR AND")
        y_pos -= 0.18*inch
        c.drawCentredString(width/2, y_pos, "LOCAL GOVERNMENT")
        
        y_pos -= 0.2*inch
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 9)
        c.drawCentredString(width/2, y_pos, "REGION IV-A CALABARZON")
        
        y_pos -= 0.15*inch
        c.drawCentredString(width/2, y_pos, "CITY OF LUCENA")
        
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.gray)
        c.drawRightString(width - 0.6*inch, y_pos - 0.3*inch, form_ref)
        
        # === LINE ===
        y_pos -= 0.5*inch
        c.setStrokeColor(colors.black)
        c.setLineWidth(1)
        c.line(0.75*inch, y_pos, width - 0.75*inch, y_pos)
        
        # === TITLE ===
        y_pos -= 0.5*inch
        c.setFillColor(colors.HexColor('#1A237E'))
        c.setFont("Helvetica-Bold", 18)
        c.drawCentredString(width/2, y_pos, "CERTIFICATION")
        
        y_pos -= 0.25*inch
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 10)
        c.drawCentredString(width/2, y_pos, "on Services Rendered in the Barangay*")
        
        y_pos -= 0.2*inch
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width/2, y_pos, position_label)
        
        # === LINE ===
        y_pos -= 0.3*inch
        c.line(0.75*inch, y_pos, width - 0.75*inch, y_pos)
        
        # === BODY ===
        y_pos -= 0.4*inch
        c.setFont("Helvetica", 10)
        c.setFillColor(colors.black)
        
        # First line with indentation (0.5 inch indent)
        indent = 0.5*inch
        text_line = f"This is to certify that "
        c.drawString(0.75*inch + indent, y_pos, text_line)
        
        name_x = 0.75*inch + indent + c.stringWidth(text_line, "Helvetica", 10)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(name_x, y_pos, eligibility_request.full_name.upper())
        
        after_name_x = name_x + c.stringWidth(eligibility_request.full_name.upper(), "Helvetica-Bold", 10)
        c.setFont("Helvetica", 10)
        c.drawString(after_name_x, y_pos, " has rendered services in")
        
        # Second line - no indent
        y_pos -= 0.18*inch
        barangay_text = f"Barangay {eligibility_request.barangay}, with the following details:"
        c.drawString(0.75*inch, y_pos, barangay_text)
        
        # === TABLE ===
        y_pos -= 0.5*inch
        
        if eligibility_request.position_type == 'elective':
            # ✅ ELECTIVE TABLE
            table_data = [
                ['Position Held', 'Date of Election\n(mm/dd/yyyy)', 'Term of Office\n(no. of years)', 
                 'From\n(mm/dd/yyyy)', 'To\n(mm/dd/yyyy)'],
                [
                    eligibility_request.position_held or 'N/A',
                    eligibility_request.election_from.strftime('%m/%d/%Y') if eligibility_request.election_from else 'N/A',
                    eligibility_request.term_office or 'N/A',
                    eligibility_request.election_from.strftime('%m/%d/%Y') if eligibility_request.election_from else 'N/A',
                    eligibility_request.election_to.strftime('%m/%d/%Y') if eligibility_request.election_to else 'N/A'
                ]
            ]
            
            col_widths = [1.4*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch]
            row_heights = [0.5*inch, 0.4*inch]
            
        else:
            # APPOINTIVE TABLE
            table_data = [
                ['Position\nHeld', 'Date of\nAppointment', 'Inclusive Dates\nFrom', 'Inclusive Dates\nTo',
                 'No. of Years\nServed', 'Appointing Punong\nBarangay Name', 'Date Elected', 'Term of Office\n(years)'],
                [
                    'Barangay\nSecretary',
                    eligibility_request.appointment_from.strftime('%m/%d/%Y') if eligibility_request.appointment_from else 'N/A',
                    eligibility_request.appointment_from.strftime('%m/%d/%Y') if eligibility_request.appointment_from else 'N/A',
                    eligibility_request.appointment_to.strftime('%m/%d/%Y') if eligibility_request.appointment_to else 'N/A',
                    f"{float(eligibility_request.years_in_service)} yrs" if eligibility_request.years_in_service else '0.0 yrs',
                    eligibility_request.appointing_punong_barangay or 'N/A',
                    eligibility_request.pb_date_elected.strftime('%m/%d/%Y') if eligibility_request.pb_date_elected else 'N/A',
                    f"{float(eligibility_request.pb_years_service)} yrs" if eligibility_request.pb_years_service else '0.0 yrs'
                ]
            ]
            col_widths = [0.8*inch, 0.75*inch, 0.75*inch, 0.75*inch, 0.7*inch, 1.1*inch, 0.7*inch, 0.7*inch]
            row_heights = [0.6*inch, 0.5*inch]
        
        table = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
        
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        table.wrapOn(c, width, height)
        table_height = table._height
        table.drawOn(c, 0.75*inch, y_pos - table_height)
        
        y_pos -= (table_height + 0.3*inch)
        
        # === ✅ COMPLETED TERM SECTION (ELECTIVE ONLY) ===
        if eligibility_request.position_type == 'elective':
            c.setFont("Helvetica-Bold", 10)
            c.setFillColor(colors.black)
            c.drawString(0.75*inch, y_pos, "Completed Term of Office?")
            
            c.setFont("Helvetica", 9)
            c.drawString(2.9*inch, y_pos, "(Please check (√) appropriate box)")
            
            y_pos -= 0.3*inch
            
            completed_term = eligibility_request.completed_term
            checkbox_size = 0.15*inch
            
            # YES checkbox
            c.setStrokeColor(colors.black)
            c.setLineWidth(1)
            c.rect(1.0*inch, y_pos - 0.05*inch, checkbox_size, checkbox_size)
            
            if completed_term and completed_term.lower() == 'yes':
                c.setFont("Helvetica-Bold", 14)
                c.drawString(1.03*inch, y_pos - 0.02*inch, "✓")
            
            c.setFont("Helvetica", 10)
            c.drawString(1.25*inch, y_pos, "YES")
            
            # NO checkbox
            c.rect(2.1*inch, y_pos - 0.05*inch, checkbox_size, checkbox_size)
            
            if completed_term and completed_term.lower() == 'no':
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2.13*inch, y_pos - 0.02*inch, "✓")
            
            c.setFont("Helvetica", 10)
            c.drawString(2.35*inch, y_pos, "NO, Specify total number of days not served")
            
            y_pos -= 0.3*inch
            
            # REASON BOX - ONLY SHOWS WHEN completed_term == 'no'
            if completed_term and completed_term.lower() == 'no':
                c.setFont("Helvetica-Bold", 9)
                c.drawString(1.0*inch, y_pos, "Reason for non-completion:")
                
                y_pos -= 0.25*inch
                
                c.setFillColor(colors.HexColor('#f5f5f5'))
                c.setStrokeColor(colors.HexColor('#cccccc'))
                c.setLineWidth(0.5)
                reason_box_height = 0.7*inch
                c.rect(1.0*inch, y_pos - reason_box_height, 5.5*inch, reason_box_height, fill=1, stroke=1)
                
                c.setFillColor(colors.black)
                c.setFont("Helvetica", 9)
                
                reason_text = eligibility_request.incomplete_reason or 'Not specified'
                max_width = 5.2*inch
                
                words = reason_text.split()
                lines = []
                current_line = []
                
                for word in words:
                    test_line = ' '.join(current_line + [word])
                    if c.stringWidth(test_line, "Helvetica", 9) < max_width:
                        current_line.append(word)
                    else:
                        if current_line:
                            lines.append(' '.join(current_line))
                        current_line = [word]
                
                if current_line:
                    lines.append(' '.join(current_line))
                
                text_y = y_pos - 0.2*inch
                for i, line in enumerate(lines[:4]):
                    c.drawString(1.1*inch, text_y, line)
                    text_y -= 0.13*inch
                
                y_pos -= (reason_box_height + 0.2*inch)
            
            # "Assumed under rule on succession" checkbox
            y_pos -= 0.25*inch
            c.setStrokeColor(colors.black)
            c.rect(1.0*inch, y_pos - 0.05*inch, checkbox_size, checkbox_size)
            c.setFont("Helvetica", 9)
            c.drawString(1.25*inch, y_pos, "Assumed under rule on succession.")
            
            y_pos -= 0.35*inch
        
        # === FOOTER TEXT ===
        c.setFont("Helvetica", 9)
        c.setFillColor(colors.black)
        
        # Create properly wrapped footer text with indentation
        left_margin = 0.75*inch
        indent = 0.5*inch  # Same indent as opening paragraph
        max_width = width - 1.5*inch - indent  # Account for indent
        
        if eligibility_request.position_type == 'elective':
            footer_text = f"This Certification is issued in support of the evaluation/processing of the application of {eligibility_request.full_name.upper()} for the grant of Barangay Official Eligibility pursuant to Republic Act No. 7160, in accordance with CSC Resolution No. 1200865 dated June 14, 2012 and CSC Resolution No. 1601257 dated November 21, 2016."
        else:
            footer_text = f"This Certification is issued in support of the evaluation/processing of the application of {eligibility_request.full_name.upper()} for the grant of Barangay Official Eligibility pursuant to Republic Act No. 7160, in accordance with CSC Resolution No. 13 series of 2012."
        
        # Word wrap the footer text
        words = footer_text.split()
        lines = []
        current_line = []
        
        for word in words:
            test_line = ' '.join(current_line + [word])
            if c.stringWidth(test_line, "Helvetica", 9) < max_width:
                current_line.append(word)
            else:
                if current_line:
                    lines.append(' '.join(current_line))
                current_line = [word]
        
        if current_line:
            lines.append(' '.join(current_line))
        
        # Draw wrapped footer text with indentation on first line only
        for i, line in enumerate(lines):
            if i == 0:
                # First line gets indentation
                c.drawString(left_margin + indent, y_pos, line)
            else:
                # Subsequent lines align with left margin
                c.drawString(left_margin, y_pos, line)
            y_pos -= 0.15*inch
        
        # === DATE ===
        y_pos -= 0.2*inch
        from django.utils import timezone
        date_text = f"Lucena City, Quezon, {timezone.now().strftime('%B %d, %Y')}."
        c.drawString(0.75*inch, y_pos, date_text)
        
        # === SIGNATURES ===
        y_pos -= 0.8*inch
        
        # Director signature (RIGHT SIDE - properly aligned)
        sig_line_width = 2.5*inch
        sig_right_margin = 0.75*inch
        sig_line_x_start = width - sig_right_margin - sig_line_width
        sig_line_x_end = width - sig_right_margin
        
        # Draw signature line
        c.setStrokeColor(colors.black)
        c.setLineWidth(0.5)
        c.line(sig_line_x_start, y_pos, sig_line_x_end, y_pos)
        
        # Director name (centered above line)
        y_pos -= 0.18*inch
        sig_center_x = sig_line_x_start + (sig_line_width / 2)
        
        c.setFont("Helvetica-Bold", 10)
        name_width = c.stringWidth("LEANDRO SIPOY GIGANTOCA, CESE", "Helvetica-Bold", 10)
        c.drawString(sig_center_x - (name_width / 2), y_pos, "LEANDRO SIPOY GIGANTOCA, CESE")
        
        # Director title (centered below name)
        y_pos -= 0.15*inch
        c.setFont("Helvetica", 9)
        title_width = c.stringWidth("OIC-HUC Director, Lucena City", "Helvetica", 9)
        c.drawString(sig_center_x - (title_width / 2), y_pos, "OIC-HUC Director, Lucena City")
        
        # Save PDF
        c.showPage()
        c.save()
        
        pdf_data = buffer.getvalue()
        buffer.close()
        
        print(f"✓ PDF generated ({len(pdf_data)} bytes)")
        
        # ✅ FIXED FILENAME GENERATION
        safe_name = eligibility_request.full_name.replace(' ', '_').replace('.', '')
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        
        if eligibility_request.position_type == 'elective':
            completion_status = 'Completed' if eligibility_request.completed_term and eligibility_request.completed_term.lower() == 'yes' else 'Incomplete'
            filename = f"{safe_name}_Elective_{completion_status}_Certificate_{timestamp}.pdf"
        else:
            filename = f"{safe_name}_Appointive_Certificate_{timestamp}.pdf"
        
        # Save to storage
        file_path = f"certification_files/{folder}/{filename}"
        saved_path = default_storage.save(file_path, ContentFile(pdf_data))
        
        print(f"✓ Saved: {saved_path}")
        
        # Create CategorizedFile record
        completion_tag = 'Completed' if eligibility_request.position_type == 'elective' and eligibility_request.completed_term and eligibility_request.completed_term.lower() == 'yes' else 'Incomplete'
        
        categorized = CategorizedFile.objects.create(
            file=saved_path,
            original_filename=filename,
            file_type='pdf',
            file_size=len(pdf_data),
            mime_type='application/pdf',
            category=category,
            source='eligibility',
            detected_content=f'{eligibility_request.get_position_type_display()} Certificate - {completion_tag}',
            eligibility_request=eligibility_request,
            uploaded_by=eligibility_request.approved_by,
            tags=f"{eligibility_request.full_name}, {eligibility_request.position_type}, {completion_tag}"
        )
        
        category.update_file_count()
        print(f"{'='*70}\n")
        
        return saved_path
        
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return None


# Add to views.py
from PIL import Image, ImageOps
import io

def process_signature_image(uploaded_file):
    """
    Process signature image to ensure it has white background
    and black signature (fixes black signature display issue)
    """
    try:
        print(f"🖊️ Processing signature: {uploaded_file.name}")
        print(f"   Original size: {uploaded_file.size} bytes")
        
        # Store original filename and content type
        original_name = uploaded_file.name
        original_content_type = uploaded_file.content_type
        
        # Read the uploaded file
        uploaded_file.seek(0)
        image = Image.open(uploaded_file)
        
        print(f"   Image mode: {image.mode}")
        print(f"   Image size: {image.size}")
        
        # Convert to RGBA first if needed
        if image.mode != 'RGBA':
            image = image.convert('RGBA')
        
        # Create a white background
        white_bg = Image.new('RGB', image.size, (255, 255, 255))
        
        # Paste the signature onto white background
        # This converts transparent areas to white
        white_bg.paste(image, mask=image.split()[-1])  # Use alpha channel as mask
        
        # Save to BytesIO
        output = io.BytesIO()
        white_bg.save(output, format='PNG', quality=95)
        output.seek(0)
        
        processed_size = output.getbuffer().nbytes
        print(f"   ✓ Processed size: {processed_size} bytes")
        
        # Create new InMemoryUploadedFile with ORIGINAL FILENAME
        processed_file = InMemoryUploadedFile(
            output,
            'ImageField',
            original_name,  # ✅ Keep the original filename
            'image/png',
            processed_size,
            None
        )
        
        # ✅ FIX: Add the 'name' attribute that save_categorized_eligibility_file expects
        processed_file.name = original_name
        
        return processed_file
        
    except Exception as e:
        print(f"   ⚠️ Error processing signature: {e}")
        import traceback
        print(traceback.format_exc())
        # Return original file if processing fails
        uploaded_file.seek(0)
        return uploaded_file

@require_http_methods(["GET"])
def setup_certificate_folders(request):
    """
    Create all required certificate folder structures
    """
    from django.core.files.storage import default_storage
    import os
    
    categories = ['appointive_certificates', 'elective_certificates', 'ids', 'signatures']
    results = {}
    
    for category in categories:
        base_path = f'certification_files/{category}/'
        
        try:
            # Create year/month structure for current year
            from datetime import datetime
            current_year = datetime.now().strftime('%Y')
            current_month = datetime.now().strftime('%m')
            
            full_path = f'{base_path}{current_year}/{current_month}/'
            
            # Create a dummy file to ensure folder exists
            dummy_file = f'{full_path}.gitkeep'
            
            if not default_storage.exists(dummy_file):
                from django.core.files.base import ContentFile
                default_storage.save(dummy_file, ContentFile(b''))
                results[category] = f'✓ Created: {full_path}'
            else:
                results[category] = f'✓ Already exists: {full_path}'
                
        except Exception as e:
            results[category] = f'✗ Error: {str(e)}'
    
    return JsonResponse({
        'success': True,
        'message': 'Certificate folders setup complete',
        'details': results
    }, json_dumps_params={'indent': 2})




#-----------------REQUIREMENTS_MONITORING--------------
@login_required
@require_http_methods(["GET", "POST"])
def user_settings_api(request):
    """API endpoint to get and update user settings"""
    
    if request.method == "GET":
        # Get current user settings
        try:
            user = request.user
            
            # Get settings from UserProfile or create default settings
            settings_data = {
                "success": True,
                "settings": {
                    "email": user.email or "",
                    "email_notifications": getattr(user.userprofile, 'email_notifications', True) if hasattr(user, 'userprofile') else True,
                    "deadline_reminders": getattr(user.userprofile, 'deadline_reminders', True) if hasattr(user, 'userprofile') else True,
                    "announcements": getattr(user.userprofile, 'announcements', True) if hasattr(user, 'userprofile') else True,
                    "compact_view": getattr(user.userprofile, 'compact_view', False) if hasattr(user, 'userprofile') else False,
                }
            }
            
            return JsonResponse(settings_data)
            
        except Exception as e:
            import traceback
            print(f"❌ Error loading settings: {e}")
            print(traceback.format_exc())
            return JsonResponse({
                "success": False,
                "error": str(e)
            }, status=500)
    
    elif request.method == "POST":
        # Save user settings
        try:
            user = request.user
            data = json.loads(request.body)
            
            # Update email if changed
            new_email = data.get('email', '').strip()
            if new_email and new_email != user.email:
                user.email = new_email
                user.save()
            
            # Update password if provided
            current_password = data.get('current_password', '')
            new_password = data.get('new_password', '')
            
            if current_password and new_password:
                # Verify current password
                if not user.check_password(current_password):
                    return JsonResponse({
                        "success": False,
                        "error": "Current password is incorrect"
                    }, status=400)
                
                # Set new password
                user.set_password(new_password)
                user.save()
                
                # Keep user logged in after password change
                update_session_auth_hash(request, user)
            
            # Update notification preferences in UserProfile
            if hasattr(user, 'userprofile'):
                profile = user.userprofile
                
                # Update notification settings
                profile.email_notifications = data.get('email_notifications', True)
                profile.deadline_reminders = data.get('deadline_reminders', True)
                profile.announcements = data.get('announcements', True)
                profile.compact_view = data.get('compact_view', False)
                
                profile.save()
            
            return JsonResponse({
                "success": True,
                "message": "Settings saved successfully"
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                "success": False,
                "error": "Invalid JSON data"
            }, status=400)
        except Exception as e:
            import traceback
            print(f"❌ Error saving settings: {e}")
            print(traceback.format_exc())
            return JsonResponse({
                "success": False,
                "error": str(e)
            }, status=500)


@login_required
@require_http_methods(["GET"])
def user_profile_api(request):
    """API endpoint to get current user's profile information"""
    try:
        user = request.user
        
        # Get barangay name from UserProfile
        barangay_name = "N/A"
        try:
            if hasattr(user, 'userprofile') and user.userprofile and user.userprofile.barangay:
                barangay_name = user.userprofile.barangay.name
        except Exception as e:
            print(f"Error getting barangay: {e}")
        
        # Get full name
        full_name = user.get_full_name().strip()
        if not full_name:
            # Parse from username: "punong_barangay_barangay_i" -> "Punong Barangay Barangay I"
            full_name = user.username.replace('_', ' ').title()
        
        # Get role display from UserProfile
        role_display = "Barangay User"
        try:
            if hasattr(user, 'userprofile') and user.userprofile:
                role_display = user.userprofile.get_role_display()
        except Exception as e:
            print(f"Error getting role: {e}")
            if user.is_superuser:
                role_display = "Administrator"
            elif user.is_staff:
                role_display = "Staff"
        
        # Format date joined
        date_joined = user.date_joined.strftime("%B %d, %Y") if user.date_joined else "N/A"
        
        profile_data = {
            "success": True,
            "profile": {
                "username": user.username,
                "email": user.email or "Not provided",
                "full_name": full_name,
                "first_name": user.first_name or "",
                "last_name": user.last_name or "",
                "barangay_name": barangay_name,
                "role_display": role_display,
                "date_joined": date_joined,
                "is_active": user.is_active,
            }
        }
        
        return JsonResponse(profile_data)
        
    except Exception as e:
        import traceback
        print(f"❌ Error in user_profile_api: {e}")
        print(traceback.format_exc())
        return JsonResponse({
            "success": False,
            "error": str(e)
        }, status=500)


@require_http_methods(["GET"])
def user_profile(request):
    return JsonResponse({
        'success': True,
        'user': {
            'full_name': request.user.get_full_name(),
            'email': request.user.email,
            'username': request.user.username,
            'role': request.user.groups.first().name if request.user.groups.exists() else 'User',
            'barangay': request.user.profile.barangay.name if hasattr(request.user, 'profile') else 'Not assigned',
            'member_since': request.user.date_joined.strftime('%B %d, %Y')
        }
    })

@login_required
def requirements_monitoring(request):
    """
    Requirements submission page - BARANGAY OFFICIALS ONLY
    Each official only sees their assigned barangay
    """
    user_profile = request.user.userprofile
    
    # 🔒 STRICT ACCESS CONTROL
    if user_profile.role == 'dilg staff':
        messages.warning(request, '⚠️ DILG Admin should use the Admin Submissions page.')
        return redirect('admin_submissions_page')
    
    if user_profile.role != 'barangay user':
        messages.error(request, '🚫 Access Denied')
        return redirect('dashboard')
    
    barangays = Barangay.objects.all().order_by('name')
    context = {
        'barangays': barangays,
        'user_role': user_profile.role,
        'page_title': 'Submit Requirements',
        'is_submitter': True,
        'can_submit': user_profile.barangay is not None,
    }
    return render(request, 'requirements_monitoring.html', context)


import logging
logger = logging.getLogger(__name__)
@login_required
def get_barangay_status(request, barangay_id):

    try:
        barangay = Barangay.objects.get(id=barangay_id)
        submissions = RequirementSubmission.objects.filter(barangay=barangay)
        today = date.today()
        
        if not submissions.exists():
            return JsonResponse({
                'status': 'no_data',
                'color': 'gray',
                'tooltip': f'{barangay.name}: No requirements assigned',
                'counts': {
                    'total': 0,
                    'overdue': 0,
                    'pending': 0,
                    'in_progress': 0,
                    'accomplished': 0,
                    'approved': 0,
                    'rejected': 0
                }
            })
        
        total = submissions.count()

        overdue = submissions.filter(
            status__in=['pending', 'in_progress', 'accomplished'],
            due_date__lt=today
        ).count()
        
        pending = submissions.filter(status='pending').count()
        in_progress = submissions.filter(status='in_progress').count()
        accomplished = submissions.filter(status='accomplished').count()
        approved = submissions.filter(status='approved').count()
        rejected = submissions.filter(status='rejected').count()
        
        if overdue > 0:
            return JsonResponse({
                'status': 'overdue',
                'color': 'red',
                'tooltip': f'{barangay.name}: {overdue} overdue requirement(s) ⚠️',
                'counts': {
                    'total': total,
                    'overdue': overdue,
                    'pending': pending,
                    'in_progress': in_progress,
                    'accomplished': accomplished,
                    'approved': approved,
                    'rejected': rejected
                }
            })
        

        elif approved == total:
            return JsonResponse({
                'status': 'completed',
                'color': 'green',
                'tooltip': f'{barangay.name}: All {total} requirements approved! ✓',
                'counts': {
                    'total': total,
                    'overdue': 0,
                    'pending': 0,
                    'in_progress': 0,
                    'accomplished': 0,
                    'approved': approved,
                    'rejected': rejected
                }
            })

        elif in_progress > 0 or accomplished > 0:
            return JsonResponse({
                'status': 'in_progress',
                'color': 'yellow',
                'tooltip': f'{barangay.name}: {in_progress} in progress, {accomplished} awaiting review',
                'counts': {
                    'total': total,
                    'overdue': 0,
                    'pending': pending,
                    'in_progress': in_progress,
                    'accomplished': accomplished,
                    'approved': approved,
                    'rejected': rejected
                }
            })
        
        elif pending > 0:
            return JsonResponse({
                'status': 'pending',
                'color': 'blue',
                'tooltip': f'{barangay.name}: {pending} pending requirements',
                'counts': {
                    'total': total,
                    'overdue': 0,
                    'pending': pending,
                    'in_progress': 0,
                    'accomplished': 0,
                    'approved': approved,
                    'rejected': rejected
                }
            })

        else:
            return JsonResponse({
                'status': 'partial',
                'color': 'blue',
                'tooltip': f'{barangay.name}: {approved}/{total} approved',
                'counts': {
                    'total': total,
                    'overdue': 0,
                    'pending': pending,
                    'in_progress': in_progress,
                    'accomplished': accomplished,
                    'approved': approved,
                    'rejected': rejected
                }
            })
            
    except Barangay.DoesNotExist:
        return JsonResponse({
            'error': 'Barangay not found',
            'status': 'error',
            'color': 'gray'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'status': 'error',
            'color': 'gray'
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_barangay_requirements(request):
    """API endpoint to list requirements for a barangay"""
    try:
        from django.utils import timezone
        
        # Get barangay from logged-in user instead of parameter
        user_profile = request.user.userprofile
        
        if user_profile.role != 'barangay official':
            return JsonResponse({
                'success': False, 
                'error': 'Only barangay officials can access this'
            }, status=403)
        
        barangay = user_profile.barangay
        
        if not barangay:
            return JsonResponse({
                'success': False, 
                'error': 'No barangay assigned to your account'
            }, status=400)
        
        period = request.GET.get('period', 'weekly')
        week = request.GET.get('week', 1)
        search = request.GET.get('search', '').strip()
        
        # Get submissions for this barangay and period
        submissions = RequirementSubmission.objects.filter(
            barangay=barangay,
            requirement__period=period,
            requirement__is_active=True
        ).select_related('requirement')
        
        # Filter by week if period is weekly
        if period == 'weekly':
            submissions = submissions.filter(week_number=week)
        
        # Search filter
        if search:
            submissions = submissions.filter(
                Q(requirement__title__icontains=search) |
                Q(requirement__description__icontains=search)
            )
        
        # Prepare response data
        submissions_data = []
        for sub in submissions:
            # Calculate is_overdue safely
            is_overdue = False
            if sub.due_date and sub.status not in ['approved', 'accomplished']:
                is_overdue = timezone.now().date() > sub.due_date
            
            # Calculate last_update safely
            last_update = 'Never'
            if sub.updated_at:
                diff = timezone.now() - sub.updated_at
                if diff.days > 0:
                    last_update = f"{diff.days} day{'s' if diff.days > 1 else ''} ago"
                else:
                    hours = diff.seconds // 3600
                    if hours > 0:
                        last_update = f"{hours} hour{'s' if hours > 1 else ''} ago"
                    else:
                        minutes = (diff.seconds % 3600) // 60
                        last_update = f"{minutes} minute{'s' if minutes > 1 else ''} ago"
            
            submissions_data.append({
                'id': sub.id,
                'title': sub.requirement.title,
                'description': sub.requirement.description,
                'status': sub.status,
                'status_display': sub.get_status_display(),
                'due_date': sub.due_date.strftime('%B %d, %Y') if sub.due_date else 'N/A',
                'is_overdue': is_overdue,
                'last_update': last_update,
                'update_text': sub.update_text or '',
            })
        
        return JsonResponse({
            'success': True,
            'submissions': submissions_data,
            'barangay_name': barangay.name
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error in api_barangay_requirements:")
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def api_requirements_list(request):
    try:
        user_profile = request.user.userprofile
        
        # Get filter parameters
        period = request.GET.get('period', 'weekly')
        week = request.GET.get('week', 1)  # This is used for ALL periods as subdivision
        search = request.GET.get('search', '').strip()
        
        print(f"📋 Requirements list request:")
        print(f"   User: {request.user.username}")
        print(f"   Role: {user_profile.role}")
        print(f"   Period: {period}, Week/Subdivision: {week}")
        
        # ============================================
        # BARANGAY USER - Show their barangay only
        # ============================================
        if 'barangay' in user_profile.role.lower():
            barangay = user_profile.barangay
            
            if not barangay:
                return JsonResponse({
                    'success': False,
                    'error': 'No barangay assigned to your account'
                }, status=400)
            
            print(f"   Barangay: {barangay.name}")

            submissions = RequirementSubmission.objects.filter(
                barangay=barangay,
                requirement__period=period,
                requirement__is_active=True
            ).select_related('requirement', 'barangay')
            
            if period == 'weekly':
                submissions = submissions.filter(week_number=int(week))
                print(f"   After weekly filter (week {week}): {submissions.count()}")
            elif period == 'annually':
                # ✅ FIX: Annually should ALWAYS show week_number=1
                submissions = submissions.filter(week_number=1)
                print(f"   After annually filter (week_number=1): {submissions.count()}")
            elif period in ['monthly', 'quarterly', 'semestral']:
                # For other periods, also filter by week
                submissions = submissions.filter(week_number=int(week))
                print(f"   After {period} filter (week {week}): {submissions.count()}")
            
            # Apply search
            if search:
                submissions = submissions.filter(
                    Q(requirement__title__icontains=search) |
                    Q(requirement__description__icontains=search) |
                    Q(update_text__icontains=search)
                )
            
            # Order by due date
            submissions = submissions.order_by('due_date', '-created_at')

            print(f"   Final count: {submissions.count()}")
            if submissions.count() > 0:
                for sub in submissions[:3]:
                    print(f"      - {sub.requirement.title} (week_number={sub.week_number})")
            
            # Format data
            submissions_data = []
            for sub in submissions:
                # Get attachments count
                attachment_count = sub.attachments.count()
                
                # Calculate last update
                last_update = 'Never'
                if sub.updated_at:
                    from django.utils import timezone
                    diff = timezone.now() - sub.updated_at
                    if diff.days > 0:
                        last_update = f"{diff.days} day{'s' if diff.days > 1 else ''} ago"
                    else:
                        hours = diff.seconds // 3600
                        if hours > 0:
                            last_update = f"{hours} hour{'s' if hours > 1 else ''} ago"
                        else:
                            minutes = (diff.seconds % 3600) // 60
                            last_update = f"{minutes} minute{'s' if minutes > 1 else ''} ago"
                
                submissions_data.append({
                    'id': sub.id,
                    'title': sub.requirement.title,
                    'description': sub.requirement.description,
                    'status': sub.status,
                    'status_display': sub.get_status_display(),
                    'due_date': sub.due_date.strftime('%B %d, %Y') if sub.due_date else 'N/A',
                    'is_overdue': sub.is_overdue if hasattr(sub, 'is_overdue') else False,
                    'last_update': last_update,
                    'update_text': sub.update_text or '',
                    'attachment_count': attachment_count
                })
            
            print(f"   ✅ Found {len(submissions_data)} submissions (after filtering)")
            
            return JsonResponse({
                'success': True,
                'submissions': submissions_data,
                'barangay_name': barangay.name,
                'barangayId': barangay.id,  
                'period': period,  
                'week': int(week),  
                'count': len(submissions_data)
            })
    
        # ============================================
        # DILG STAFF - Show all requirements (admin)
        # ============================================
        elif 'dilg' in user_profile.role.lower():
            # Get filter parameters
            barangay_id = request.GET.get('barangay_id')
            status_filter = request.GET.get('status')
            
            # Base query - all submissions
            submissions = RequirementSubmission.objects.select_related(
                'requirement',
                'barangay',
                'submitted_by'
            ).filter(
                requirement__is_active=True
            )
            
            # Apply filters
            if barangay_id:
                submissions = submissions.filter(barangay_id=barangay_id)
            
            if status_filter:
                submissions = submissions.filter(status=status_filter)
            
            if period:
                submissions = submissions.filter(requirement__period=period)
                
                # ✅ OPTION 1: Only filter by week for Weekly period (admin view)
                if period == 'weekly':
                    submissions = submissions.filter(week_number=int(week))
                elif period == 'annually':
                    submissions = submissions.filter(week_number=1)
            
            if search:
                submissions = submissions.filter(
                    Q(requirement__title__icontains=search) |
                    Q(barangay__name__icontains=search) |
                    Q(update_text__icontains=search)
                )
            
            # Order by most recent
            submissions = submissions.order_by('-submitted_at', '-created_at')
            
            # Format for admin view
            submissions_data = []
            for sub in submissions:
                attachments = []
                for attachment in sub.attachments.all():
                    attachments.append({
                        'id': attachment.id,
                        'file_name': attachment.file.name.split('/')[-1],
                        'file_url': attachment.file.url if attachment.file else '#',
                        'file_size': round(attachment.file.size / 1024) if attachment.file else 0,
                        'uploaded_at': attachment.uploaded_at.strftime('%B %d, %Y')
                    })
                
                submissions_data.append({
                    'id': sub.id,
                    'title': sub.requirement.title,
                    'barangay_name': sub.barangay.name,
                    'status': sub.status,
                    'status_display': sub.get_status_display(),
                    'due_date': sub.due_date.strftime('%B %d, %Y') if sub.due_date else 'N/A',
                    'submitted_at': sub.submitted_at.strftime('%B %d, %Y') if sub.submitted_at else 'Not submitted',
                    'submitted_by': sub.submitted_by.get_full_name() if sub.submitted_by else 'Unknown',
                    'period': sub.requirement.get_period_display(),
                    'update_text': sub.update_text or '',
                    'attachments': attachments,
                    'is_overdue': sub.is_overdue if hasattr(sub, 'is_overdue') else False
                })
            
            print(f"   ✅ Found {len(submissions_data)} submissions (admin view)")
            
            return JsonResponse({
                'success': True,
                'submissions': submissions_data,
                'count': len(submissions_data)
            })
        
        else:
            return JsonResponse({
                'success': False,
                'error': f'Invalid role: {user_profile.role}'
            }, status=403)
        
    except Exception as e:
        print(f"❌ Error in api_requirements_list: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)



@login_required
def debug_requirement_issue(request):
    """
    Debug endpoint to check requirement and notification setup
    Access at: /api/debug-requirement-issue/
    """
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        debug_info = {
            'current_user': {
                'username': request.user.username,
                'role': request.user.userprofile.role if hasattr(request.user, 'userprofile') else 'No profile',
                'barangay': request.user.userprofile.barangay.name if hasattr(request.user, 'userprofile') and request.user.userprofile.barangay else None,
            },
            'requirements': {
                'total': Requirement.objects.count(),
                'active': Requirement.objects.filter(is_active=True).count(),
                'recent': []
            },
            'submissions': {
                'total': RequirementSubmission.objects.count(),
                'by_barangay': {}
            },
            'barangay_users': [],
            'notifications': {
                'total': Notification.objects.count(),
                'recent_requirement_notifications': []
            }
        }
        
        # Get recent requirements
        recent_reqs = Requirement.objects.order_by('-created_at')[:5]
        for req in recent_reqs:
            debug_info['requirements']['recent'].append({
                'id': req.id,
                'title': req.title,
                'period': req.period,
                'created_at': str(req.created_at),
                'is_active': req.is_active
            })
        
        # Count submissions by barangay
        for barangay in Barangay.objects.all():
            count = RequirementSubmission.objects.filter(barangay=barangay).count()
            debug_info['submissions']['by_barangay'][barangay.name] = count
        
        # Get all barangay users
        barangay_users = User.objects.filter(
            userprofile__barangay__isnull=False
        ).values('username', 'userprofile__role', 'userprofile__barangay__name', 'is_active')
        
        debug_info['barangay_users'] = list(barangay_users)
        
        # Get recent requirement notifications
        recent_notifs = Notification.objects.filter(
            notification_type='new_requirement'
        ).order_by('-created_at')[:10]
        
        for notif in recent_notifs:
            debug_info['notifications']['recent_requirement_notifications'].append({
                'id': notif.id,
                'user': notif.user.username,
                'title': notif.title,
                'created_at': str(notif.created_at),
                'is_read': notif.is_read
            })
        
        return JsonResponse(debug_info, json_dumps_params={'indent': 2})
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from .models import Requirement
@login_required
@require_http_methods(["GET"])
def requirements_list_api(request):
    """
    DEBUG VERSION: Shows every step of what's happening
    """
    print(f"\n{'='*80}")
    print(f"🔍 REQUIREMENTS_LIST_API DEBUG START")
    print(f"{'='*80}")
    
    try:
        # STEP 1: Check user authentication
        print(f"\n📌 STEP 1: User Authentication")
        print(f"   User: {request.user.username}")
        print(f"   Is authenticated: {request.user.is_authenticated}")
        print(f"   Is staff: {request.user.is_staff}")
        
        # STEP 2: Check user profile
        print(f"\n📌 STEP 2: User Profile")
        try:
            user_profile = request.user.userprofile
            print(f"   ✅ Profile found")
            print(f"   Role: '{user_profile.role}'")
            print(f"   Role (repr): {repr(user_profile.role)}")
            print(f"   Barangay: {user_profile.barangay}")
        except Exception as profile_error:
            print(f"   ❌ Profile error: {str(profile_error)}")
            return JsonResponse({
                'success': False,
                'error': 'User profile not found'
            }, status=403)
        
        # STEP 3: Check authorization
        print(f"\n📌 STEP 3: Authorization Check")
        if user_profile.role != 'dilg staff':
            print(f"   ❌ UNAUTHORIZED")
            print(f"   Expected: 'dilg staff'")
            print(f"   Got: '{user_profile.role}'")
            return JsonResponse({
                'success': False,
                'error': f'Unauthorized access. Role: {user_profile.role}'
            }, status=403)
        print(f"   ✅ Authorized (DILG Staff)")
        
        # STEP 4: Get filter parameters
        print(f"\n📌 STEP 4: Request Parameters")
        period = request.GET.get('period', '')
        priority = request.GET.get('priority', '')
        status = request.GET.get('status', 'active')
        sort = request.GET.get('sort', '-created_at')
        
        print(f"   period: '{period}' (empty={not period})")
        print(f"   priority: '{priority}' (empty={not priority})")
        print(f"   status: '{status}'")
        print(f"   sort: '{sort}'")
        print(f"   Full query string: {request.GET.urlencode()}")
        
        # STEP 5: Import Requirement model
        print(f"\n📌 STEP 5: Import Requirement Model")
        try:
            from .models import Requirement
            print(f"   ✅ Requirement model imported")
        except ImportError as import_error:
            print(f"   ❌ Import error: {str(import_error)}")
            return JsonResponse({
                'success': False,
                'error': 'Could not import Requirement model'
            }, status=500)
        
        # STEP 6: Check database connection
        print(f"\n📌 STEP 6: Database Query")
        print(f"   Querying Requirement.objects.all()...")
        
        try:
            # Get ALL requirements first (no filters)
            all_requirements = Requirement.objects.all()
            total_count = all_requirements.count()
            print(f"   ✅ Total requirements in database: {total_count}")
            
            # Show each requirement
            if total_count > 0:
                print(f"\n   📋 ALL Requirements in Database:")
                for req in all_requirements:
                    print(f"      ID:{req.id:3d} | {req.title:40s} | {req.period:10s} | Active:{req.is_active} | Priority:{req.priority}")
            else:
                print(f"   ⚠️ NO REQUIREMENTS FOUND IN DATABASE!")
                print(f"   This means the Requirement table is empty.")
                
        except Exception as db_error:
            print(f"   ❌ Database error: {str(db_error)}")
            import traceback
            print(traceback.format_exc())
            return JsonResponse({
                'success': False,
                'error': f'Database error: {str(db_error)}'
            }, status=500)
        
        # STEP 7: Apply filters
        print(f"\n📌 STEP 7: Applying Filters")
        requirements = Requirement.objects.all()
        print(f"   Starting with: {requirements.count()} requirements")
        
        # Status filter
        if status == 'active':
            requirements = requirements.filter(is_active=True)
            print(f"   After is_active=True filter: {requirements.count()}")
        elif status == 'archived':
            requirements = requirements.filter(is_active=False)
            print(f"   After is_active=False filter: {requirements.count()}")
        else:
            print(f"   No status filter applied (showing all)")
        
        # Period filter
        if period:
            requirements = requirements.filter(period=period)
            print(f"   After period='{period}' filter: {requirements.count()}")
        else:
            print(f"   No period filter applied")
        
        # Priority filter
        if priority:
            requirements = requirements.filter(priority=priority)
            print(f"   After priority='{priority}' filter: {requirements.count()}")
        else:
            print(f"   No priority filter applied")
        
        # Sorting
        print(f"   Applying sort: {sort}")
        requirements = requirements.order_by(sort)
        
        final_count = requirements.count()
        print(f"   ✅ Final filtered count: {final_count}")
        
        # STEP 8: Build response
        print(f"\n📌 STEP 8: Building Response")
        requirements_data = []
        
        if final_count > 0:
            print(f"   📋 Filtered Requirements:")
            for req in requirements:
                print(f"      ID:{req.id:3d} | {req.title:40s} | {req.period:10s}")
                
                requirements_data.append({
                    'id': req.id,
                    'title': req.title,
                    'description': req.description,
                    'period': req.period,
                    'period_display': req.get_period_display(),
                    'priority': req.priority,
                    'due_date': req.due_date.strftime('%B %d, %Y') if req.due_date else 'N/A',
                    'is_active': req.is_active,
                    'created_by': req.created_by.username if req.created_by else 'System',
                    'created_at': req.created_at.strftime('%B %d, %Y %I:%M %p'),
                })
        else:
            print(f"   ⚠️ No requirements match the filters!")
        
        # STEP 9: Return response
        print(f"\n📌 STEP 9: Returning Response")
        response_data = {
            'success': True,
            'requirements': requirements_data,
            'count': len(requirements_data)
        }
        print(f"   Response count: {len(requirements_data)}")
        print(f"   Response size: ~{len(str(response_data))} bytes")
        
        print(f"\n{'='*80}")
        print(f"✅ REQUIREMENTS_LIST_API DEBUG END - SUCCESS")
        print(f"{'='*80}\n")
        
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"\n{'='*80}")
        print(f"❌ REQUIREMENTS_LIST_API DEBUG END - ERROR")
        print(f"{'='*80}")
        print(f"Error type: {type(e).__name__}")
        print(f"Error message: {str(e)}")
        
        import traceback
        print(f"\nFull traceback:")
        print(traceback.format_exc())
        print(f"{'='*80}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'error_type': type(e).__name__
        }, status=500)



@login_required
@require_http_methods(["GET"])
def debug_requirements_count(request):
    """
    Simple debug endpoint to check database
    Call this from: /api/debug/requirements/
    """
    try:
        from .models import Requirement
        
        all_reqs = Requirement.objects.all()
        active_reqs = Requirement.objects.filter(is_active=True)
        archived_reqs = Requirement.objects.filter(is_active=False)
        
        requirements_list = []
        for req in all_reqs:
            requirements_list.append({
                'id': req.id,
                'title': req.title,
                'period': req.period,
                'priority': req.priority,
                'is_active': req.is_active,
                'due_date': str(req.due_date) if req.due_date else None,
            })
        
        return JsonResponse({
            'success': True,
            'total_count': all_reqs.count(),
            'active_count': active_reqs.count(),
            'archived_count': archived_reqs.count(),
            'requirements': requirements_list,
            'message': 'This is what is in your database'
        }, json_dumps_params={'indent': 2})
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500, json_dumps_params={'indent': 2})



@login_required
@require_http_methods(["GET"])
def requirement_detail_api(request, requirement_id):
    """
    Get single requirement details for editing
    URL: /api/admin/requirements/<id>/
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized'
            }, status=403)
        
        req = Requirement.objects.get(id=requirement_id)
        
        return JsonResponse({
            'success': True,
            'requirement': {
                'id': req.id,
                'title': req.title,
                'description': req.description,
                'period': req.period,
                'due_date': req.due_date.strftime('%Y-%m-%d') if req.due_date else '',
                'priority': req.priority,
                'is_active': req.is_active,
            }
        })
        
    except Requirement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Requirement not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_requirement_detail(request, requirement_id):
    """
    API endpoint to get single requirement details
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized'
            }, status=403)
        
        req = get_object_or_404(Requirement, id=requirement_id)
        
        return JsonResponse({
            'success': True,
            'requirement': {
                'id': req.id,
                'title': req.title,
                'description': req.description,
                'period': req.period,
                'priority': req.priority,
                'due_date': req.due_date.strftime('%Y-%m-%d') if req.due_date else '',
                'is_active': req.is_active,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_update_requirement(request, requirement_id):
    """
    API endpoint to update a requirement
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized'
            }, status=403)
        
        req = get_object_or_404(Requirement, id=requirement_id)
        
        data = json.loads(request.body)
        
        # Update fields
        req.title = data.get('title', req.title)
        req.description = data.get('description', req.description)
        req.period = data.get('period', req.period)
        req.priority = data.get('priority', req.priority)
        
        # Update due_date if provided
        due_date_str = data.get('due_date')
        if due_date_str:
            from datetime import datetime
            req.due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
            
            # Update due_date for all related submissions
            RequirementSubmission.objects.filter(requirement=req).update(due_date=req.due_date)
        
        req.save()
        
        # Log the update
        try:
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                content_object=req,
                description=f"Updated requirement: {req.title}"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement updated successfully'
        })
        
    except Exception as e:
        import traceback
        print(f"=== UPDATE REQUIREMENT ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_archive_requirement(request, requirement_id):
    """
    API endpoint to archive a requirement
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized'
            }, status=403)
        
        req = get_object_or_404(Requirement, id=requirement_id)
        req.is_active = False
        req.save()
        
        # Log the archive
        try:
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                content_object=req,
                description=f"Archived requirement: {req.title}"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement archived successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["POST"])
def api_restore_requirement(request, requirement_id):
    """
    API endpoint to restore an archived requirement
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized'
            }, status=403)
        
        req = get_object_or_404(Requirement, id=requirement_id)
        req.is_active = True
        req.save()
        
        # Log the restore
        try:
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                content_object=req,
                description=f"Restored requirement: {req.title}"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement restored successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["GET"])
def api_user_session(request):
    """Get current user session info"""
    try:
        user_profile = request.user.userprofile
        
        return JsonResponse({
            'success': True,
            'user': {
                'id': request.user.id,
                'username': request.user.username,
                'full_name': request.user.get_full_name() or request.user.username,
                'role': user_profile.role,
            },
            'barangay_id': user_profile.barangay.id if user_profile.barangay else None,
            'barangay_name': user_profile.barangay.name if user_profile.barangay else None,
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["GET"])
def api_submission_detail(request, submission_id):
    """API endpoint to get submission details"""
    try:
        from django.utils import timezone
        import os
        
        submission = get_object_or_404(
            RequirementSubmission.objects.select_related(
                'requirement', 'barangay', 'submitted_by'
            ).prefetch_related('attachments'),
            id=submission_id
        )
        
        # Calculate is_overdue
        is_overdue = False
        if submission.due_date and submission.status not in ['approved', 'accomplished']:
            is_overdue = timezone.now().date() > submission.due_date
        
        # Calculate last_update
        last_update = 'Never'
        if submission.updated_at:
            diff = timezone.now() - submission.updated_at
            if diff.days > 0:
                last_update = f"{diff.days} day{'s' if diff.days > 1 else ''} ago"
            else:
                hours = diff.seconds // 3600
                if hours > 0:
                    last_update = f"{hours} hour{'s' if hours > 1 else ''} ago"
                else:
                    minutes = (diff.seconds % 3600) // 60
                    last_update = f"{minutes} minute{'s' if minutes > 1 else ''} ago"
        
        # Get attachments safely
        attachments = []
        for att in submission.attachments.all():
            try:
                file_name = os.path.basename(att.file.name) if att.file else 'Unknown'
                file_url = att.file.url if att.file else ''
                file_size = round(att.file.size / 1024) if att.file else 0
                
                attachments.append({
                    'id': att.id,
                    'file_name': file_name,
                    'file_size': file_size,
                    'file_url': file_url,
                })
            except Exception as att_error:
                print(f"Error processing attachment {att.id}: {att_error}")
                continue
        
        data = {
            'id': submission.id,
            'title': submission.requirement.title,
            'description': submission.requirement.description,
            'status': submission.status,
            'status_display': submission.get_status_display(),
            'due_date': submission.due_date.strftime('%B %d, %Y') if submission.due_date else 'N/A',
            'is_overdue': is_overdue,
            'update_text': submission.update_text or '',
            'last_update': last_update,
            'attachments': attachments,
        }
        
        return JsonResponse({
            'success': True,
            'submission': data
        })
        
    except RequirementSubmission.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': f'Submission with ID {submission_id} not found'
        }, status=404)
    except Exception as e:
        import traceback
        print(f"=== ERROR in api_submission_detail ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_submission_submit(request, submission_id):
    """
    Submit requirement to admin (mark as accomplished)
    THIS IS THE ACTUAL FUNCTION BEING CALLED!
    """
    try:
        from .models import RequirementSubmission, AuditLog
        
        print(f"\n{'='*60}")
        print(f"🚀 API SUBMISSION SUBMIT")
        print(f"{'='*60}")
        print(f"👤 User: {request.user.username}")
        print(f"📋 Submission ID: {submission_id}")
        
        submission = RequirementSubmission.objects.select_related(
            'requirement', 'barangay'
        ).get(id=submission_id)
        
        # Check permission
        user_profile = request.user.userprofile
        
        if user_profile.role == 'barangay official':
            if submission.barangay != user_profile.barangay:
                print(f"❌ ERROR: Unauthorized - wrong barangay")
                return JsonResponse({
                    'success': False,
                    'error': 'Unauthorized'
                }, status=403)
        
        print(f"📄 Requirement: {submission.requirement.title}")
        print(f"🏘️ Barangay: {submission.barangay.name}")
        
        # Get update text from request
        data = json.loads(request.body) if request.body else {}
        update_text = data.get('update_text', '').strip()
        
        print(f"📝 Update text: {update_text[:50] if update_text else 'None'}...")
        
        # Update submission
        old_status = submission.status
        submission.status = 'accomplished'
        submission.submitted_at = timezone.now()
        submission.submitted_by = request.user
        if update_text:
            submission.update_text = update_text
        submission.save()
        
        print(f"✅ Submission updated:")
        print(f"   Old status: {old_status}")
        print(f"   New status: {submission.status}")
        print(f"   Submitted at: {submission.submitted_at}")
        print(f"   Submitted by: {request.user.username}")
        
        # ========== CREATE NOTIFICATIONS FOR ADMINS ==========
        print(f"\n🔔 Creating admin notifications...")
        notifications_sent = notify_admins(
            title=f"📋 New Submission: {submission.requirement.title}",
            message=f"{submission.barangay.name} submitted '{submission.requirement.title}' for review.",
            notification_type='new_submission',
            submission=submission
        )
        print(f"✅ Notifications sent to {notifications_sent} admins")
        # ====================================================
        
        # Log the submission
        try:
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                content_object=submission,
                description=f"Submitted requirement: {submission.requirement.title} to admin"
            )
            print(f"✅ Audit log created")
        except Exception as audit_error:
            print(f"⚠️ Audit log failed: {str(audit_error)}")
        
        print(f"\n✅ SUBMISSION COMPLETE")
        print(f"   Total notifications sent: {notifications_sent}")
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement submitted successfully! Waiting for admin approval.',
            'notifications_sent': notifications_sent
        })
        
    except RequirementSubmission.DoesNotExist:
        print(f"❌ ERROR: Submission {submission_id} not found")
        return JsonResponse({
            'success': False,
            'error': 'Submission not found'
        }, status=404)
        
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ SUBMIT ERROR")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def debug_submission(request, submission_id):
    try:
        sub = RequirementSubmission.objects.get(id=submission_id)
        return JsonResponse({
            'submission_exists': True,
            'requirement_title': sub.requirement.title,
            'barangay_name': sub.barangay.name,
            'attachment_count': sub.attachments.count(),
            'status': sub.status
        })
    except Exception as e:
        return JsonResponse({'error': str(e)})

@login_required
@require_http_methods(["POST"])
def api_submission_update(request, submission_id):
    """API endpoint to update submission text"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        update_text = request.POST.get('update_text', '').strip()

        if not user_profile.can_access_barangay(submission.barangay):
            return JsonResponse({
                'success': False,
                'error': '🚫 Access Denied'
            }, status=403)
        
        if not update_text:
            return JsonResponse({'success': False, 'error': 'Update text required'}, status=400)
        
        submission.update_text = update_text
        submission.save()
        
        # Log the update
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            content_object=submission,
            description=f"Updated requirement: {submission.requirement.title}"
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Update saved successfully',
            'submission': {
                'id': submission.id,
                'update_text': submission.update_text,
                'last_update': submission.updated_at.strftime('%B %d, %Y at %I:%M %p'),
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def api_attachment_upload(request, submission_id):
    """API endpoint to upload file attachments"""
    try:
        import os
        
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Check access
        user_profile = request.user.userprofile
        if user_profile.role == 'barangay official' and submission.barangay != user_profile.barangay:
            return JsonResponse({
                'success': False,
                'error': 'Access Denied'
            }, status=403)
        
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False, 
                'error': 'No file provided'
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Validate file size (10MB)
        max_size = 10 * 1024 * 1024
        if uploaded_file.size > max_size:
            return JsonResponse({
                'success': False, 
                'error': 'File too large. Max: 10MB'
            }, status=400)
        
        # Create attachment
        attachment = RequirementAttachment.objects.create(
            submission=submission,
            file=uploaded_file,
            file_type=uploaded_file.content_type,
            file_size=uploaded_file.size,
            uploaded_by=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'File uploaded successfully',
            'attachment': {
                'id': attachment.id,
                'file_name': os.path.basename(attachment.file.name),
                'file_size': round(attachment.file.size / 1024),
                'file_url': attachment.file.url,
            }
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Upload error: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False, 
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_attachment_delete(request, attachment_id):
    """API endpoint to delete attachment"""
    try:
        attachment = get_object_or_404(RequirementAttachment, id=attachment_id)
        
        # Check if user has permission to delete
        if attachment.uploaded_by != request.user and not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Log before deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            description=f"Deleted attachment: {os.path.basename(attachment.file.name)}"
        )
        
        attachment.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'File removed successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)




@login_required
@require_http_methods(["POST"])
def api_submission_delete(request, submission_id):
    """API endpoint to delete submission"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Check if user has permission
        if not request.user.is_staff:
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Log before deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            description=f"Deleted submission: {submission.requirement.title} - {submission.barangay.name}"
        )
        
        submission.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def requirements_monitoring(request):
    """
    Requirements submission page - BARANGAY OFFICIALS ONLY
    Each official only sees their assigned barangay
    """
    user_profile = request.user.userprofile
    
    # 🔒 ACCESS CONTROL
    if user_profile.role == 'dilg staff':
        messages.warning(request, '⚠️ DILG Staff should use the Admin Submissions page.')
        return redirect('admin_submissions_page')
    
    if user_profile.role not in ['barangay official', 'municipal officer']:
        messages.error(request, '🚫 Access Denied: This page is only for Barangay Officials.')
        return redirect('dashboard')
    
    # 🆕 FILTER BARANGAYS BASED ON ROLE
    if user_profile.role == 'barangay official':
        # Barangay officials only see their assigned barangay
        if user_profile.barangay:
            barangays = Barangay.objects.filter(id=user_profile.barangay.id)
        else:
            barangays = Barangay.objects.none()
            messages.error(request, '⚠️ You are not assigned to any barangay. Contact DILG Admin.')
    else:
        # Municipal officers see all barangays
        barangays = Barangay.objects.all().order_by('name')
    
    context = {
        'barangays': barangays,
        'user_role': user_profile.role,
        'page_title': 'Submit Requirements',
        'is_submitter': True,
        'assigned_barangay': user_profile.barangay,  # 🆕 Pass to template
    }
    return render(request, 'requirements_monitoring.html', context)

@login_required
@require_http_methods(["GET"])
def get_requirements_list(request):
    """AJAX endpoint to get requirements list for a barangay"""
    try:
        barangay_id = request.GET.get('barangay_id')
        period = request.GET.get('period', 'weekly')
        week = request.GET.get('week', 1)
        search = request.GET.get('search', '')
        
        if not barangay_id:
            return JsonResponse({
                'success': False,
                'error': 'Barangay ID is required'
            }, status=400)
        
        barangay = get_object_or_404(Barangay, id=barangay_id)
        
        # 🔒 VERIFY ACCESS
        user_profile = request.user.userprofile
        if not user_profile.can_access_barangay(barangay):
            return JsonResponse({
                'success': False,
                'error': '🚫 Access Denied: You can only view your assigned barangay.'
            }, status=403)
        
        # Get current year and week
        current_year = timezone.now().year
        current_week = int(week)
        
        # 🆕 Get requirements for this period that apply to this barangay
        # Either: no specific barangays (applies to all) OR includes this barangay
        requirements = Requirement.objects.filter(
            period=period,
            is_active=True
        ).filter(
            Q(applicable_barangays__isnull=True) | 
            Q(applicable_barangays=barangay)
        ).distinct()
        
        # Apply search filter
        if search:
            requirements = requirements.filter(
                Q(title__icontains=search) | Q(description__icontains=search)
            )
        
        # 🆕 Get or create submissions for these requirements
        submissions_data = []
        for req in requirements:
            # Get or create submission for this week/period
            submission, created = RequirementSubmission.objects.get_or_create(
                requirement=req,
                barangay=barangay,
                week_number=current_week if period == 'weekly' else None,
                year=current_year,
                defaults={
                    'due_date': calculate_due_date(period, current_week, current_year),
                    'status': 'pending'
                }
            )
            
            submissions_data.append({
                'id': submission.id,
                'requirement_id': req.id,
                'title': req.title,
                'description': req.description,
                'status': submission.status,
                'status_display': submission.get_status_display(),
                'due_date': submission.due_date.strftime('%B %d, %Y'),
                'last_update': submission.updated_at.strftime('%B %d, %Y'),
                'is_overdue': submission.is_overdue,
                'has_attachments': submission.attachments.exists(),
                'attachment_count': submission.attachments.count(),
            })
        
        return JsonResponse({
            'success': True,
            'submissions': submissions_data,
            'barangay_name': barangay.name,
            'period': period,
            'week': current_week
        })
        
    except Exception as e:
        import traceback
        print(f"Error in get_requirements_list: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["GET"])
def get_submission_detail(request, submission_id):
    """Get detailed information about a specific submission"""
    try:
        submission = get_object_or_404(
            RequirementSubmission.objects.select_related(
                'requirement', 'barangay', 'submitted_by', 'reviewed_by'
            ).prefetch_related('attachments'),
            id=submission_id
        )
        
        # 🆕 VERIFY ACCESS
        user_profile = request.user.userprofile
        if not user_profile.can_access_barangay(submission.barangay):
            return JsonResponse({
                'success': False,
                'error': '🚫 Access Denied: You cannot view this submission.'
            }, status=403)
        
        # Get attachments
        attachments = []
        for attachment in submission.attachments.all():
            attachments.append({
                'id': attachment.id,
                'file_url': attachment.file.url,
                'file_name': attachment.file.name.split('/')[-1],
                'file_size': attachment.file_size_kb,
                'file_type': attachment.file_type,
                'uploaded_at': attachment.uploaded_at.strftime('%B %d, %Y %I:%M %p')
            })
        
        data = {
            'id': submission.id,
            'requirement': {
                'title': submission.requirement.title,
                'description': submission.requirement.description,
                'period': submission.requirement.get_period_display()
            },
            'barangay': {
                'name': submission.barangay.name,
                'code': submission.barangay.code
            },
            'status': submission.status,
            'status_display': submission.get_status_display(),
            'due_date': submission.due_date.strftime('%B %d, %Y'),
            'week_number': submission.week_number,
            'year': submission.year,
            'update_text': submission.update_text or '',
            'is_overdue': submission.is_overdue,
            'submitted_by': submission.submitted_by.get_full_name() if submission.submitted_by else None,
            'submitted_at': submission.submitted_at.strftime('%B %d, %Y %I:%M %p') if submission.submitted_at else None,
            'reviewed_by': submission.reviewed_by.get_full_name() if submission.reviewed_by else None,
            'reviewed_at': submission.reviewed_at.strftime('%B %d, %Y %I:%M %p') if submission.reviewed_at else None,
            'review_notes': submission.review_notes or '',
            'attachments': attachments,
            'last_update': submission.updated_at.strftime('%B %d, %Y %I:%M %p')
        }
        
        return JsonResponse({
            'success': True,
            'submission': data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def update_submission(request, submission_id):
    """Update a requirement submission"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        data = json.loads(request.body)
        update_text = data.get('update_text', '').strip()
        
        if not update_text:
            return JsonResponse({
                'success': False,
                'error': 'Update text is required'
            }, status=400)
        
        # Update submission
        submission.update_text = update_text
        submission.status = 'in_progress'
        submission.updated_at = timezone.now()
        submission.save()
        
        # Log the update
        AuditLog.objects.create(
            user=request.user,
            action='UPDATE',
            content_object=submission,
            description=f"Updated requirement submission: {submission.requirement.title}"
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Submission updated successfully',
            'last_update': submission.updated_at.strftime('%B %d, %Y %I:%M %p')
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def upload_attachment(request, submission_id):
    """Upload file attachments for a submission"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Get uploaded files
        files = request.FILES.getlist('files')
        
        if not files:
            return JsonResponse({
                'success': False,
                'error': 'No files uploaded'
            }, status=400)
        
        uploaded_files = []
        
        for file in files:
            # Validate file size (max 5MB)
            if file.size > 5 * 1024 * 1024:
                return JsonResponse({
                    'success': False,
                    'error': f'File {file.name} is too large (max 5MB)'
                }, status=400)
            
            # Create attachment
            attachment = RequirementAttachment.objects.create(
                submission=submission,
                file=file,
                file_type=file.content_type,
                file_size=file.size,
                uploaded_by=request.user
            )
            
            uploaded_files.append({
                'id': attachment.id,
                'file_name': file.name,
                'file_size': attachment.file_size_kb,
                'file_url': attachment.file.url
            })
        
        # Log the upload
        AuditLog.objects.create(
            user=request.user,
            action='CREATE',
            content_object=submission,
            description=f"Uploaded {len(files)} file(s) for requirement: {submission.requirement.title}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f'{len(files)} file(s) uploaded successfully',
            'files': uploaded_files
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_attachment(request, attachment_id):
    """Delete a file attachment"""
    try:
        attachment = get_object_or_404(RequirementAttachment, id=attachment_id)
        submission = attachment.submission
        
        # Delete the file
        attachment.delete()
        
        # Log the deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            content_object=submission,
            description=f"Deleted attachment for requirement: {submission.requirement.title}"
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Attachment deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def submit_to_admin(request, submission_id):
    """Submit requirement to admin for review"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Validate that there's content
        if not submission.update_text:
            return JsonResponse({
                'success': False,
                'error': 'Please add update details before submitting'
            }, status=400)
        
        # Update submission status
        submission.submit(request.user)
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement submitted to admin successfully',
            'new_status': submission.status,
            'submitted_at': submission.submitted_at.strftime('%B %d, %Y %I:%M %p')
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_submission(request, submission_id):
    """Delete a requirement submission"""
    try:
        submission = get_object_or_404(RequirementSubmission, id=submission_id)
        
        # Only allow deletion if status is pending
        if submission.status not in ['pending', 'in_progress']:
            return JsonResponse({
                'success': False,
                'error': 'Cannot delete submitted or reviewed requirements'
            }, status=400)
        
        requirement_title = submission.requirement.title
        
        # Log before deletion
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            description=f"Deleted requirement submission: {requirement_title}"
        )
        
        # Delete submission (attachments will be deleted via cascade)
        submission.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Requirement "{requirement_title}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


from datetime import datetime, timedelta
from django.utils import timezone

from app.models import Requirement, Barangay, RequirementSubmission
from django.utils import timezone
from datetime import datetime, timedelta



#--------Admin Views and APIs --------#
# ============== NOTIFICATION HELPER FUNCTIONS ==============
# Place these at the top of your views.py file
def create_notification_for_user(user, title, message, notification_type='info', 
                                  submission=None, announcement=None, barangay=None):
    """
    Master notification creator with extensive error handling
    """
    try:
        from .models import Notification
        
        print(f"\n🔔 CREATING NOTIFICATION:")
        print(f"   User: {user.username} (ID: {user.id})")
        print(f"   Title: {title}")
        print(f"   Type: {notification_type}")
        
        notification = Notification.objects.create(
            user=user,
            title=title,
            message=message,
            notification_type=notification_type,
            submission=submission,
            announcement=announcement,
            barangay=barangay,
            is_read=False,
            created_at=timezone.now()
        )
        
        print(f"   ✅ Notification created successfully (ID: {notification.id})")
        return notification
        
    except Exception as e:
        print(f"   ❌ ERROR creating notification: {str(e)}")
        traceback.print_exc()
        return None

def notify_admins(title, message, notification_type='info', submission=None):
    """
    Send notification to ALL admin users (DILG Staff)
    """
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # Get all admin users (DILG staff OR superusers) in a single query
        all_admins = User.objects.filter(
            Q(userprofile__role='dilg staff') | Q(is_superuser=True)
        ).distinct()
        
        print(f"\n📢 NOTIFYING ADMINS:")
        print(f"   Found {all_admins.count()} admin users")
        print(f"   Title: {title}")
        
        notifications_created = 0
        for admin in all_admins:
            notif = create_notification_for_user(
                user=admin,
                title=title,
                message=message,
                notification_type=notification_type,
                submission=submission
            )
            if notif:
                notifications_created += 1
        
        print(f"   ✅ Total: {notifications_created} admin notifications created\n")
        return notifications_created
        
    except Exception as e:
        print(f"   ❌ ERROR notifying admins: {str(e)}")
        import traceback
        traceback.print_exc()
        return 0


def notify_barangay_user(submission, title, message, notification_type='info'):
    """
    Send notification to the barangay user(s) who own the submission
    """
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # Get all users from this barangay
        barangay_users = User.objects.filter(
            userprofile__barangay=submission.barangay,
            userprofile__role='barangay official'
        )
        
        print(f"\n📢 NOTIFYING BARANGAY:")
        print(f"   Barangay: {submission.barangay.name}")
        print(f"   Found {barangay_users.count()} barangay users")
        print(f"   Title: {title}")
        
        if not barangay_users.exists():
            print(f"   ⚠️ WARNING: No barangay users found for {submission.barangay.name}")
            return 0
        
        notifications_created = 0
        for user in barangay_users:
            notif = create_notification_for_user(
                user=user,
                title=title,
                message=message,
                notification_type=notification_type,
                submission=submission,
                barangay=submission.barangay
            )
            if notif:
                notifications_created += 1
        
        print(f"   ✅ Total: {notifications_created} barangay notifications created\n")
        return notifications_created
        
    except Exception as e:
        print(f"   ❌ ERROR notifying barangay: {str(e)}")
        traceback.print_exc()
        return 0


@login_required
@require_POST
def create_announcement(request):
    """Create a new announcement"""
    try:
        data = json.loads(request.body)
        
        print("=" * 60)
        print("📢 CREATE ANNOUNCEMENT")
        print("=" * 60)
        print(f"👤 Admin: {request.user.username}")
        print(f"📋 Title: {data.get('title')}")
        print(f"📅 Date: {data.get('date')}")  # ADD THIS
        print(f"📊 Priority: {data.get('priority')}")
        print(f"🔔 Send notification: {data.get('send_notification')}")
        
        # Create announcement with date
        announcement = Announcement.objects.create(
            title=data.get('title'),
            content=data.get('content'),
            date=data.get('date'),  # ADD THIS LINE
            priority=data.get('priority', 'medium'),
            posted_by=request.user
        )
        
        print(f"✅ Announcement created (ID: {announcement.id})")
        
        notifications_sent = 0
        send_notification = data.get('send_notification', False)
        
        if send_notification:
            barangay_users = User.objects.filter(
                userprofile__role='barangay official',
                is_active=True
            ).distinct()
            
            print(f"🔔 Creating notifications for {barangay_users.count()} barangay users...")
            
            notification_list = []
            for user in barangay_users:
                notification_list.append(
                    Notification(
                        user=user,
                        title=f"📢 New Announcement: {announcement.title}",
                        message=f"{announcement.content[:100]}...",
                        notification_type='info',
                        announcement=announcement,
                        created_at=timezone.now()
                    )
                )
                print(f"   ✅ Notified: {user.username}")
            
            Notification.objects.bulk_create(notification_list)
            notifications_sent = len(notification_list)
            
        print(f"✅ Total notifications sent: {notifications_sent}")
        print("=" * 60)
        
        return JsonResponse({
            'success': True,
            'message': 'Announcement created successfully',
            'announcement_id': announcement.id,
            'notifications_sent': notifications_sent
        })
        
    except Exception as e:
        import traceback
        print(f"❌ CREATE ANNOUNCEMENT ERROR")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

import sys
import logging

# Force console output
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

def force_print(*args, **kwargs):
    """Force print even if stdout is redirected"""
    message = ' '.join(str(arg) for arg in args)
    print(message, **kwargs)
    sys.stdout.flush()
    logger.info(message)

@login_required
@require_http_methods(["POST"])
def api_submit_requirement(request, submission_id):
    """
    EMERGENCY DEBUG VERSION - Shows exactly what's happening
    """
    force_print("\n" + "="*80)
    force_print("🚨 API_SUBMIT_REQUIREMENT CALLED!")
    force_print("="*80)
    force_print(f"Time: {timezone.now()}")
    force_print(f"User: {request.user.username}")
    force_print(f"Submission ID: {submission_id}")
    force_print(f"Request method: {request.method}")
    force_print(f"Request path: {request.path}")
    force_print(f"Request body: {request.body[:200] if request.body else 'Empty'}")
    
    try:
        from .models import RequirementSubmission, Notification
        from django.db import transaction
        
        force_print("\n📦 Models imported successfully")
        
        # Get the submission
        force_print(f"\n🔍 Looking for submission #{submission_id}...")
        submission = RequirementSubmission.objects.select_related(
            'requirement', 'barangay', 'submitted_by'
        ).get(id=submission_id)
        
        force_print(f"✅ Submission found:")
        force_print(f"   Requirement: {submission.requirement.title}")
        force_print(f"   Barangay: {submission.barangay.name}")
        force_print(f"   Current status: {submission.status}")
        
        # Check permission
        if not hasattr(request.user, 'userprofile'):
            force_print("❌ ERROR: User has no profile")
            return JsonResponse({
                'success': False,
                'error': 'User profile not found'
            }, status=403)
        
        force_print(f"\n✅ User profile found:")
        force_print(f"   User barangay: {request.user.userprofile.barangay.name}")
        force_print(f"   Submission barangay: {submission.barangay.name}")
        
        if submission.barangay != request.user.userprofile.barangay:
            force_print("❌ ERROR: Permission denied")
            return JsonResponse({
                'success': False,
                'error': 'Permission denied'
            }, status=403)
        
        # Get update text from request body
        try:
            data = json.loads(request.body) if request.body else {}
            update_text = data.get('update_text', '')
            force_print(f"\n📝 Update text length: {len(update_text)} characters")
        except Exception as e:
            force_print(f"⚠️ Could not parse request body: {str(e)}")
            update_text = ''
        
        # ✅ Use atomic transaction
        force_print("\n💾 Starting database transaction...")
        
        with transaction.atomic():
            # Update submission status
            old_status = submission.status
            submission.status = 'accomplished'
            submission.submitted_at = timezone.now()
            submission.submitted_by = request.user
            if update_text:
                submission.update_text = update_text
            submission.save()
            
            force_print(f"✅ Submission updated:")
            force_print(f"   Old status: {old_status}")
            force_print(f"   New status: {submission.status}")
            force_print(f"   Submitted at: {submission.submitted_at}")
            
            # ========== GET ALL ADMIN USERS ==========
            from django.contrib.auth import get_user_model
            from django.db.models.functions import Lower
            
            User = get_user_model()
            
            force_print(f"\n🔍 FINDING ADMIN USERS...")
            
            # Find admins - case insensitive
            admin_users = User.objects.annotate(
                role_lower=Lower('userprofile__role')
            ).filter(
                role_lower__icontains='dilg'
            ).distinct()
            
            force_print(f"   Found {admin_users.count()} users with 'dilg' in role")
            
            superusers = User.objects.filter(is_superuser=True)
            force_print(f"   Found {superusers.count()} superusers")
            
            all_admins = (admin_users | superusers).distinct()
            
            force_print(f"\n   ✅ TOTAL ADMINS: {all_admins.count()}")
            for admin in all_admins:
                role = getattr(admin.userprofile, 'role', 'superuser') if hasattr(admin, 'userprofile') else 'superuser'
                force_print(f"     - {admin.username} (role: '{role}')")
            
            if all_admins.count() == 0:
                force_print(f"\n   ⚠️ WARNING: NO ADMINS FOUND!")
                force_print(f"   Checking what users exist:")
                all_users = User.objects.filter(userprofile__isnull=False)[:10]
                for u in all_users:
                    force_print(f"     - {u.username}: role='{u.userprofile.role}'")
            
            # ========== CREATE NOTIFICATIONS ==========
            force_print(f"\n🔔 CREATING NOTIFICATIONS...")
            notifications_created = 0
            
            title = f"📋 New Submission: {submission.requirement.title}"
            message = f"{submission.barangay.name} submitted '{submission.requirement.title}' for review."
            
            force_print(f"   Title: {title}")
            force_print(f"   Message: {message}")
            
            for admin in all_admins:
                try:
                    force_print(f"\n   Creating notification for {admin.username}...")
                    
                    # ✅ CRITICAL: Explicitly set is_read=False
                    notification = Notification.objects.create(
                        user=admin,
                        title=title,
                        message=message,
                        notification_type='new_submission',
                        submission=submission,
                        barangay=submission.barangay,
                        is_read=False,
                        created_at=timezone.now()
                    )
                    
                    notifications_created += 1
                    force_print(f"     ✅ Created notification #{notification.id}")
                    force_print(f"        is_read={notification.is_read}")
                    
                    # Double-check it was saved correctly
                    saved_notif = Notification.objects.get(id=notification.id)
                    force_print(f"        Verified from DB: is_read={saved_notif.is_read}")
                    
                    if saved_notif.is_read:
                        force_print(f"        ⚠️ WARNING: Saved as READ! Fixing...")
                        saved_notif.is_read = False
                        saved_notif.save()
                        force_print(f"        Fixed: is_read={saved_notif.is_read}")
                    
                except Exception as e:
                    force_print(f"     ❌ Failed for {admin.username}: {str(e)}")
                    import traceback
                    traceback.print_exc()
        
        # ========== VERIFICATION ==========
        force_print(f"\n📊 VERIFICATION:")
        force_print(f"   Notifications created: {notifications_created}")
        
        # Check what was actually saved
        force_print(f"\n🔍 Checking database for new notifications...")
        recent_notifs = Notification.objects.filter(
            notification_type='new_submission',
            submission=submission
        ).values('id', 'user__username', 'is_read', 'created_at')
        
        force_print(f"   Found {len(recent_notifs)} notifications in database:")
        for notif in recent_notifs:
            force_print(f"     #{notif['id']}: user={notif['user__username']}, is_read={notif['is_read']}")
        
        force_print(f"\n✅ SUBMISSION COMPLETE")
        force_print(f"{'='*80}\n")
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement submitted successfully',
            'notifications_created': notifications_created,
            'debug': {
                'admins_found': all_admins.count(),
                'submission_id': submission.id,
                'submission_status': submission.status
            }
        })
        
    except RequirementSubmission.DoesNotExist:
        force_print(f"❌ ERROR: Submission {submission_id} not found")
        return JsonResponse({
            'success': False,
            'error': 'Submission not found'
        }, status=404)
        
    except Exception as e:
        force_print(f"\n{'='*80}")
        force_print(f"❌ SUBMISSION ERROR")
        force_print(f"{'='*80}")
        force_print(f"Error: {str(e)}")
        import traceback
        force_print(traceback.format_exc())
        force_print(f"{'='*80}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    

@login_required
@require_http_methods(["GET", "POST"])
def test_endpoint(request, submission_id):
    """
    Simple test endpoint to verify routing works
    """
    force_print("\n" + "="*80)
    force_print("🧪 TEST ENDPOINT HIT!")
    force_print("="*80)
    force_print(f"Method: {request.method}")
    force_print(f"Path: {request.path}")
    force_print(f"Submission ID: {submission_id}")
    force_print(f"User: {request.user.username}")
    force_print("="*80 + "\n")
    
    return JsonResponse({
        'success': True,
        'message': 'Test endpoint works!',
        'submission_id': submission_id,
        'user': request.user.username,
        'method': request.method
    })

@login_required
def admin_calendar_view(request):
    """Return all submissions AND announcements for calendar view"""
    print(f"🔍 Calendar Request from user: {request.user.username}")
    
    try:
        user_profile = request.user.userprofile
        print(f"🔍 User role: {user_profile.role}")
        
        if user_profile.role not in ['admin', 'dilg staff']:
            print(f"❌ Access denied - Role is '{user_profile.role}'")
            return JsonResponse({'success': False, 'error': f'Unauthorized - Role: {user_profile.role}'}, status=403)
            
        print(f"✅ Access granted - User role: {user_profile.role}")
        
    except UserProfile.DoesNotExist:
        print(f"❌ UserProfile does not exist for user: {request.user.username}")
        return JsonResponse({'success': False, 'error': 'User profile not found'}, status=403)
    
    try:
        month = int(request.GET.get('month', datetime.now().month))
        year = int(request.GET.get('year', datetime.now().year))
        
        print(f"📅 Loading calendar for: {month}/{year}")
        
        # Get first and last day of the month
        first_day = datetime(year, month, 1)
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)
        
        # Fetch all submissions for the month
        submissions = RequirementSubmission.objects.filter(
            due_date__gte=first_day.date(),
            due_date__lte=last_day.date()
        ).select_related('barangay', 'requirement')
        
        requirements_data = []
        for sub in submissions:
            requirements_data.append({
                'id': sub.id,
                'title': sub.requirement.title,
                'barangay_name': sub.barangay.name,
                'due_date': sub.due_date.strftime('%Y-%m-%d'),
                'status': sub.status,
                'status_display': sub.get_status_display(),
                'type': 'requirement'  # ADD THIS to distinguish from announcements
            })
        
        # ADD THIS: Fetch announcements for the month
        announcements = Announcement.objects.filter(
            date__gte=first_day.date(),
            date__lte=last_day.date()
        ).select_related('posted_by')
        
        for announcement in announcements:
            requirements_data.append({
                'id': f'announcement_{announcement.id}',  # Prefix to avoid ID conflicts
                'title': f"📢 {announcement.title}",
                'barangay_name': 'All Barangays',
                'due_date': announcement.date.strftime('%Y-%m-%d'),
                'status': 'announcement',
                'status_display': announcement.get_priority_display(),
                'type': 'announcement',  # ADD THIS
                'priority': announcement.priority
            })
        
        print(f"✅ Calendar loaded: {len(requirements_data)} items found")
        print(f"   - Requirements: {submissions.count()}")
        print(f"   - Announcements: {announcements.count()}")
        
        return JsonResponse({
            'success': True,
            'requirements': requirements_data
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Calendar Error: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def test_create_notification(request):
    """
    TEST ENDPOINT: Create a test notification for the current user
    Call this from: /api/test-notification/
    """
    try:
        print(f"\n{'='*60}")
        print(f"🧪 TEST NOTIFICATION CREATION")
        print(f"{'='*60}")
        print(f"User: {request.user.username}")
        
        # Create notification
        notification = Notification.objects.create(
            user=request.user,
            title="🧪 Test Notification",
            message="This is a test notification created at " + timezone.now().strftime('%H:%M:%S'),
            notification_type='info',
            is_read=False,
            created_at=timezone.now()
        )
        
        print(f"✅ Created notification ID: {notification.id}")
        print(f"   Title: {notification.title}")
        print(f"   is_read: {notification.is_read}")
        print(f"   created_at: {notification.created_at}")
        
        # Verify it was saved correctly
        saved = Notification.objects.get(id=notification.id)
        print(f"\n🔍 Verification from database:")
        print(f"   ID: {saved.id}")
        print(f"   User: {saved.user.username}")
        print(f"   is_read: {saved.is_read} ← Should be False!")
        print(f"   created_at: {saved.created_at}")
        
        if saved.is_read:
            print(f"   ⚠️ WARNING: Notification was saved as READ!")
            print(f"   This means something is auto-marking notifications as read!")
            print(f"   Check your model's save() method or database triggers!")
        
        # Get current unread count
        unread_count = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).count()
        
        print(f"\n📊 Current unread count for {request.user.username}: {unread_count}")
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': True,
            'message': 'Test notification created',
            'notification': {
                'id': notification.id,
                'title': notification.title,
                'is_read': notification.is_read,
                'created_at': notification.created_at.isoformat()
            },
            'verification': {
                'saved_is_read': saved.is_read,
                'matches': notification.is_read == saved.is_read,
                'unread_count': unread_count
            }
        })
        
    except Exception as e:
        print(f"❌ ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def admin_submissions_page(request):
    """
    Admin review page - DILG STAFF ONLY
    This is where DILG admin reviews all submissions from all barangays
    """
    user_profile = request.user.userprofile
    
    # 🔒 STRICT ACCESS CONTROL - Only DILG Staff
    if user_profile.role == 'barangay official':
        messages.error(request, '🚫 Access Denied: Barangay Officials cannot access the admin review page.')
        return redirect('requirements_monitoring')  # Redirect to submission page
    
    if user_profile.role != 'dilg staff':
        messages.error(request, '🚫 Access Denied: This page is only accessible to DILG Admin.')
        return redirect('dashboard')
    
    barangays = Barangay.objects.all().order_by('name')
    
    context = {
        'barangays': barangays,
        'page_title': 'Review Submissions',
        'is_admin_view': True,  # Flag to show this is admin page
    }
    return render(request, 'admin_submissions.html', context)


@login_required
@require_http_methods(["GET"])
def api_admin_submissions_list(request):
    """
    API endpoint for admin to fetch all submissions with filters
    """
    try:
        # Check if user is admin
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized access'
            }, status=403)
        
        # Get filter parameters
        barangay_id = request.GET.get('barangay_id')
        status = request.GET.get('status')
        period = request.GET.get('period')
        search = request.GET.get('search', '').strip()
        
        # Base query - get all submissions
        submissions = RequirementSubmission.objects.select_related(
            'requirement',
            'barangay',
            'submitted_by'  # ✅ FIXED: was 'created_by'
        ).all()
        
        # Apply filters
        if barangay_id:
            submissions = submissions.filter(barangay_id=barangay_id)
        
        if status:
            submissions = submissions.filter(status=status)
        
        if period:
            submissions = submissions.filter(requirement__period=period)
        
        if search:
            submissions = submissions.filter(
                Q(requirement__title__icontains=search) |
                Q(barangay__name__icontains=search) |
                Q(update_text__icontains=search)
            )
        
        # Order by most recent first
        submissions = submissions.order_by('-submitted_at', '-created_at')
        
        # Format submissions for JSON
        submissions_data = []
        for sub in submissions:
            # Get attachments
            attachments = []
            for attachment in sub.attachments.all():
                attachments.append({
                    'id': attachment.id,
                    'file_name': attachment.file.name.split('/')[-1],
                    'file_url': attachment.file.url if attachment.file else '#',
                    'file_size': round(attachment.file.size / 1024) if attachment.file else 0,
                    'uploaded_at': attachment.uploaded_at.strftime('%B %d, %Y')
                })
            
            submissions_data.append({
            'id': sub.id,
            'title': sub.requirement.title,
            'barangay_name': sub.barangay.name,
            'status': sub.status,
            'status_display': sub.get_status_display(),
            'due_date': sub.due_date.strftime('%B %d, %Y') if sub.due_date else 'N/A',
            'submitted_at': sub.submitted_at.strftime('%B %d, %Y') if sub.submitted_at else 'Not submitted',
            'submitted_by': sub.submitted_by.get_full_name() if sub.submitted_by else 'Unknown',
            'period': sub.requirement.get_period_display(),
            'update_text': sub.update_text or '',
            'attachments': attachments,
            'is_overdue': sub.is_overdue if hasattr(sub, 'is_overdue') else False  # ✅ FIXED: removed ()
        })
        
        return JsonResponse({
            'success': True,
            'submissions': submissions_data,
            'count': len(submissions_data)
        })
        
    except Exception as e:
        print(f"Error in admin_submissions_list: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_admin_review_submission(request, submission_id):
    """
    FIXED: Admin approves/rejects submission and notifies barangay
    """
    try:
        from .models import RequirementSubmission
        
        print(f"\n{'='*60}")
        print(f"⚖️ ADMIN REVIEW SUBMISSION")
        print(f"{'='*60}")
        print(f"👤 Admin: {request.user.username}")
        print(f"📋 Submission ID: {submission_id}")
        
        # Check admin permission
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff' and not request.user.is_superuser:
            print(f"❌ ERROR: User is not admin")
            return JsonResponse({
                'success': False,
                'error': 'Permission denied'
            }, status=403)
        
        # Get submission
        submission = RequirementSubmission.objects.select_related(
            'requirement', 'barangay'
        ).get(id=submission_id)
        
        print(f"📄 Requirement: {submission.requirement.title}")
        print(f"🏘️ Barangay: {submission.barangay.name}")
        
        # Get action
        data = json.loads(request.body)
        action = data.get('action')  # 'approved' or 'rejected'
        review_notes = data.get('review_notes', '')
        
        if action not in ['approved', 'rejected']:
            return JsonResponse({
                'success': False,
                'error': 'Invalid action'
            }, status=400)
        
        print(f"🔍 Action: {action}")
        print(f"📝 Review notes: {review_notes[:50] if review_notes else 'None'}...")
        
        # Update submission
        old_status = submission.status
        submission.status = action
        submission.reviewed_by = request.user
        submission.reviewed_at = timezone.now()
        submission.review_notes = review_notes
        submission.save()
        
        print(f"✅ Submission reviewed:")
        print(f"   Old status: {old_status}")
        print(f"   New status: {submission.status}")
        print(f"   Reviewed by: {request.user.username}")
        
        # ========== NOTIFY BARANGAY USER ==========
        print(f"\n🔔 Creating barangay notification...")
        
        if action == 'approved':
            title = f"✅ Approved: {submission.requirement.title}"
            message = f"Your submission for '{submission.requirement.title}' has been approved!"
            if review_notes:
                message += f" Admin notes: {review_notes}"
            notif_type = 'completed'
        else:  # rejected
            title = f"❌ Needs Revision: {submission.requirement.title}"
            message = f"Your submission for '{submission.requirement.title}' needs revision."
            if review_notes:
                message += f" Admin feedback: {review_notes}"
            notif_type = 'info'
        
        notifications_sent = notify_barangay_user(
            submission=submission,
            title=title,
            message=message,
            notification_type=notif_type
        )
        # ==========================================
        
        print(f"\n✅ REVIEW COMPLETE")
        print(f"   Notifications sent: {notifications_sent}")
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': True,
            'message': f'Submission {action} successfully',
            'notifications_sent': notifications_sent
        })
        
    except RequirementSubmission.DoesNotExist:
        print(f"❌ ERROR: Submission {submission_id} not found")
        return JsonResponse({
            'success': False,
            'error': 'Submission not found'
        }, status=404)
        
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ REVIEW ERROR")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)



def check_page_access(request, allowed_roles):
    """
    Reusable function to check if user has access to a page
    Returns: (has_access: bool, redirect_url: str or None)
    """
    if not request.user.is_authenticated:
        return False, 'login_page'
    
    try:
        user_profile = request.user.userprofile
        if user_profile.role in allowed_roles:
            return True, None
        
        # Determine where to redirect based on role
        if user_profile.role == 'dilg staff':
            return False, 'admin_submissions_page'
        elif user_profile.role == 'barangay official':
            return False, 'requirements_monitoring'
        else:
            return False, 'dashboard'
    except:
        return False, 'dashboard'
    

@login_required
@require_http_methods(["POST"])
def api_create_requirement(request):
    """
    API endpoint for DILG Admin to create new requirements
    ONLY accessible by DILG Staff
    """
    try:
        # STRICT ACCESS CONTROL
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized: Only DILG Admin can create requirements'
            }, status=403)
        
        data = json.loads(request.body)
        
        title = data.get('title', '').strip()
        description = data.get('description', '').strip()
        period = data.get('period', '').strip()
        priority = data.get('priority', 'normal').strip()
        due_date_str = data.get('due_date', '').strip()  # ✅ ADD THIS LINE
        barangay_ids = data.get('barangay_ids', [])
        
        # Validation
        if not title:
            return JsonResponse({'success': False, 'error': 'Title is required'}, status=400)
        
        if not description:
            return JsonResponse({'success': False, 'error': 'Description is required'}, status=400)
        
        if not period:
            return JsonResponse({'success': False, 'error': 'Period is required'}, status=400)
        
        # ✅ ADD THIS: Validate due_date
        if not due_date_str:
            return JsonResponse({'success': False, 'error': 'Due date is required'}, status=400)
        
        # ✅ ADD THIS: Parse due_date
        try:
            from datetime import datetime
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
            print(f"📅 Due date parsed: {due_date}")
        except ValueError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }, status=400)
        
        valid_periods = ['weekly', 'monthly', 'quarterly', 'semestral', 'annually']
        if period not in valid_periods:
            return JsonResponse({
                'success': False,
                'error': f'Invalid period. Must be one of: {valid_periods}'
            }, status=400)
        
        # Create requirement
        requirement = Requirement.objects.create(
            title=title,
            description=description,
            period=period,
            priority=priority,
            due_date=due_date,
            created_by=request.user,
            is_active=True
        )
        
        # Get target barangays
        if barangay_ids:
            target_barangays = Barangay.objects.filter(id__in=barangay_ids)
            requirement.applicable_barangays.set(target_barangays)
        else:
            target_barangays = Barangay.objects.all()
        
        # ============================================
        # CREATE SUBMISSIONS
        # ============================================
        current_year = timezone.now().year
        submissions_created = 0
        
        for barangay in target_barangays:
            if period == 'weekly':
                for week_num in range(1, 5):
                    RequirementSubmission.objects.create(
                        requirement=requirement,
                        barangay=barangay,
                        week_number=week_num,
                        year=current_year,
                        due_date=due_date,
                        status='pending'
                    )
                    submissions_created += 1
            else:
                # ✅ FIX: For all other periods, use week_number=1
                RequirementSubmission.objects.create(
                    requirement=requirement,
                    barangay=barangay,
                    week_number=1,  # ✅ IMPORTANT: Always 1 for non-weekly
                    year=current_year,
                    due_date=due_date,
                    status='pending'
                )
                submissions_created += 1
        
        print(f"✅ Created {submissions_created} submissions")
        
        # ============================================
        # ✅ CREATE NOTIFICATIONS - FIXED VERSION
        # ============================================
        notifications_sent = 0
        priority_emoji = {'normal': '📋', 'important': '⚠️', 'urgent': '🚨'}
        
        print(f"\n🔔 Creating notifications...")
        print(f"   Target barangays: {target_barangays.count()}")
        
        # ✅ Find barangay users with flexible role matching
        barangay_users = User.objects.filter(
            userprofile__barangay__in=target_barangays,
            userprofile__role__icontains='barangay',  # ✅ Flexible matching
            is_active=True
        ).distinct()
        
        print(f"   Found {barangay_users.count()} barangay users")
        
        if barangay_users.count() == 0:
            print(f"   ⚠️ WARNING: No barangay users found!")
        
        # Create notifications
        for user in barangay_users:
            try:
                notification = Notification.objects.create(
                    user=user,
                    notification_type='new_requirement',
                    title=f"{priority_emoji.get(priority, '📋')} New {priority.upper()} Requirement",
                    message=f"A new {period} requirement has been added: {title}. Due: {due_date.strftime('%B %d, %Y')}",
                    is_read=False,
                    created_at=timezone.now()
                )
                notifications_sent += 1
                print(f"   ✓ Notified: {user.username}")
            except Exception as e:
                print(f"   ✗ Failed for {user.username}: {str(e)}")
        
        print(f"   ✅ Total notifications sent: {notifications_sent}\n")
        
        # ============================================
        # RETURN RESPONSE
        # ============================================
        return JsonResponse({
            'success': True,
            'message': f'Requirement created successfully!',
            'requirement': {
                'id': requirement.id,
                'title': requirement.title,
                'period': requirement.period,
                'due_date': due_date.strftime('%Y-%m-%d'),
                'submissions_created': submissions_created,
                'notifications_sent': notifications_sent,
                'barangays_notified': target_barangays.count(),
            }
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@login_required
@require_http_methods(["POST"])
def api_edit_requirement(request, requirement_id):
    """
    API endpoint for DILG Admin to edit existing requirements
    ONLY accessible by DILG Staff
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized: Only DILG Admin can edit requirements'
            }, status=403)
        
        requirement = get_object_or_404(Requirement, id=requirement_id)
        
        data = json.loads(request.body)
        
        title = data.get('title', '').strip()
        description = data.get('description', '').strip()
        period = data.get('period', '').strip()
        is_active = data.get('is_active', True)
        
        if title:
            requirement.title = title
        if description:
            requirement.description = description
        if period:
            requirement.period = period
        
        requirement.is_active = is_active
        requirement.save()
        
        # Log the update
        try:
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                content_object=requirement,
                description=f"DILG Admin updated requirement: {requirement.title}"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@login_required
@require_http_methods(["DELETE"])
def api_delete_requirement(request, requirement_id):
    """
    API endpoint for DILG Admin to delete requirements
    ONLY accessible by DILG Staff
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized: Only DILG Admin can delete requirements'
            }, status=403)
        
        requirement = get_object_or_404(Requirement, id=requirement_id)
        title = requirement.title
        
        # Log before deletion
        try:
            AuditLog.objects.create(
                user=request.user,
                action='DELETE',
                description=f"DILG Admin deleted requirement: {title}"
            )
        except:
            pass
        
        requirement.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Requirement "{title}" deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)



@login_required
@require_http_methods(["GET"])
def api_all_requirements(request):
    """
    API endpoint to get all requirements (for DILG Admin management)
    ONLY accessible by DILG Staff
    """
    try:
        user_profile = request.user.userprofile
        if user_profile.role != 'dilg staff':
            return JsonResponse({
                'success': False,
                'error': 'Unauthorized'
            }, status=403)
        
        requirements = Requirement.objects.all().order_by('-created_at')
        
        requirements_data = []
        for req in requirements:
            applicable_barangays = list(req.applicable_barangays.values_list('name', flat=True))
            
            requirements_data.append({
                'id': req.id,
                'title': req.title,
                'description': req.description,
                'period': req.period,
                'period_display': req.get_period_display(),
                'is_active': req.is_active,
                'applicable_barangays': applicable_barangays if applicable_barangays else ['All Barangays'],
                'created_at': req.created_at.strftime('%B %d, %Y'),
                'created_by': req.created_by.get_full_name() if req.created_by else 'System',
            })
        
        return JsonResponse({
            'success': True,
            'requirements': requirements_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    

#-------------------ANNOUNCEMENTS AND NOTIFICATIONS------------
@login_required
def delete_announcement(request, announcement_id):
    """Delete an announcement"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Only POST requests are allowed'
        }, status=405)
    
    try:
        print(f"Attempting to delete announcement with ID: {announcement_id}")
        
        # Get the announcement
        announcement = Announcement.objects.get(id=announcement_id)
        print(f"Found announcement: {announcement.title}")
        
        # Store info before deletion
        title = announcement.title
        
        # Delete the announcement
        announcement.delete()
        print(f"Successfully deleted announcement: {title}")
        
        return JsonResponse({
            'success': True,
            'message': f'Announcement "{title}" deleted successfully'
        })
        
    except Announcement.DoesNotExist:
        print(f"Announcement with ID {announcement_id} does not exist")
        return JsonResponse({
            'success': False,
            'error': 'Announcement not found'
        }, status=404)
        
    except Exception as e:
        # Log the full error for debugging
        error_details = traceback.format_exc()
        print(f"Error deleting announcement {announcement_id}:")
        print(error_details)
        
        return JsonResponse({
            'success': False,
            'error': f'Server error: {str(e)}'
        }, status=500)
@login_required
@require_POST
def update_announcement(request, announcement_id):
    """Update an existing announcement"""
    try:
        announcement = Announcement.objects.get(id=announcement_id)
        data = json.loads(request.body)
        
        # Update fields
        announcement.title = data.get('title', announcement.title)
        announcement.content = data.get('content', announcement.content)
        announcement.date = data.get('date', announcement.date)  
        announcement.priority = data.get('priority', announcement.priority)
        announcement.save()

        notifications_sent = 0
        send_notification = data.get('send_notification', False)
        
        # Send update notifications if enabled
        if send_notification:
            barangay_users = User.objects.filter(
                userprofile__role='barangay official',
                is_active=True
            ).distinct()
            
            notification_list = []
            for user in barangay_users:
                notification_list.append(
                    Notification(
                        user=user,
                        title=f"📢 Updated: {announcement.title}",
                        message=f"An announcement has been updated: {announcement.content[:100]}...",
                        notification_type='info',
                        created_at=timezone.now()
                    )
                )
            
            Notification.objects.bulk_create(notification_list)
            notifications_sent = len(notification_list)
        
        return JsonResponse({
            'success': True,
            'announcement': {
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'priority': announcement.priority
            },
            'notifications_sent': notifications_sent
        })
        
    except Announcement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Announcement not found'
        }, status=404)
    except Exception as e:
        import traceback
        print(f"=== UPDATE ANNOUNCEMENT ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
@login_required
@require_http_methods(["GET"])
def get_announcements(request):
    """Get all announcements"""
    try:
        announcements = Announcement.objects.all().select_related('posted_by').order_by('-posted_at')
        
        announcements_list = []
        for announcement in announcements:
            announcements_list.append({
                'id': announcement.id,
                'title': announcement.title,
                'content': announcement.content,
                'date': announcement.date.strftime('%Y-%m-%d'),  # ADD THIS LINE
                'priority': announcement.priority,
                'priority_display': announcement.get_priority_display(),
                'posted_by': announcement.posted_by.get_full_name() or announcement.posted_by.username,
                'posted_at': announcement.posted_at.strftime('%B %d, %Y'),
                'views': announcement.views,
                'sent_to_barangays': 33  
            })
        
        return JsonResponse({
            'success': True,
            'announcements': announcements_list
        })
        
    except Exception as e:
        import traceback
        print(f"=== GET ANNOUNCEMENTS ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
@login_required
def debug_users(request):
    """Debug view to see all users and their roles"""
    users = User.objects.all()
    
    user_list = []
    for user in users:
        try:
            role = user.userprofile.role if hasattr(user, 'userprofile') else 'No profile'
        except:
            role = 'Error getting role'
        
        user_list.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'is_active': user.is_active,
            'is_superuser': user.is_superuser,
            'role': role
        })
    
    return JsonResponse({
        'total_users': users.count(),
        'users': user_list
    }, json_dumps_params={'indent': 2})

#-----------------END OF ANNOUNCEMENTS AND NOTIFICATIONS--------

#---------------NOTIFICATIONS----------------------

def create_announcement_notification(announcement, send_to_all=True):
    """
    Helper function to create notifications when announcements are posted
    This is called from create_announcement view
    """
    try:
        if send_to_all:
            # Get all barangay officials
            barangay_users = User.objects.filter(
                userprofile__role='barangay official',
                is_active=True
            ).distinct()
            
            # Bulk create notifications
            notification_list = []
            for user in barangay_users:
                notification_list.append(
                    Notification(
                        user=user,
                        title=f"📢 New Announcement: {announcement.title}",
                        message=f"{announcement.content[:150]}{'...' if len(announcement.content) > 150 else ''}",
                        notification_type='announcement',
                        announcement=announcement
                    )
                )
            
            Notification.objects.bulk_create(notification_list)
            return len(notification_list)
        
        return 0
        
    except Exception as e:
        print(f"Error creating announcement notifications: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return 0


@login_required
@require_http_methods(["GET"])
def get_notifications(request):
    """
    FIXED: Get all notifications with proper error handling
    """
    try:
        from .models import Notification
        
        user = request.user
        print(f"\n📥 FETCHING NOTIFICATIONS for {user.username}")
        
        # Count unread BEFORE slicing
        unread_count = Notification.objects.filter(
            user=user,
            is_read=False
        ).count()
        
        # Get all notifications for the user (slice at the end)
        notifications = Notification.objects.filter(
            user=user
        ).select_related(
            'submission__requirement',
            'submission__barangay',
            'announcement',
            'barangay'
        ).order_by('-created_at')[:50]
        
        print(f"   Found {len(notifications)} notifications")
        print(f"   Unread: {unread_count}")
        
        # Format notifications
        notifications_list = []
        for notif in notifications:
            try:
                # Calculate time ago
                now = timezone.now()
                diff = now - notif.created_at
                seconds = diff.total_seconds()
                
                if seconds < 60:
                    time_ago = "Just now"
                elif seconds < 3600:
                    minutes = int(seconds / 60)
                    time_ago = f"{minutes}m ago"
                elif seconds < 86400:
                    hours = int(seconds / 3600)
                    time_ago = f"{hours}h ago"
                elif seconds < 604800:
                    days = int(seconds / 86400)
                    time_ago = f"{days}d ago"
                else:
                    time_ago = notif.created_at.strftime('%b %d')
                
                notification_data = {
                    'id': notif.id,
                    'title': notif.title,
                    'message': notif.message,
                    'type': notif.notification_type,
                    'is_read': notif.is_read,
                    'created_at': notif.created_at.isoformat(),
                    'time_ago': time_ago,
                }
                
                # Add related IDs safely
                if notif.submission:
                    notification_data['submission_id'] = notif.submission.id
                if notif.announcement:
                    notification_data['announcement_id'] = notif.announcement.id
                if notif.barangay:
                    notification_data['barangay_id'] = notif.barangay.id
                
                notifications_list.append(notification_data)
                
            except Exception as inner_e:
                print(f"   ⚠️ Error formatting notification {notif.id}: {str(inner_e)}")
                continue
        
        print(f"   ✅ Successfully formatted {len(notifications_list)} notifications\n")
        
        return JsonResponse({
            'success': True,
            'notifications': notifications_list,
            'unread_count': unread_count,
            'total_count': len(notifications_list)
        })
        
    except Exception as e:
        print(f"\n❌ GET NOTIFICATIONS ERROR")
        print(f"User: {request.user.username if request.user else 'Unknown'}")
        print(f"Error: {str(e)}")
        traceback.print_exc()
        print()
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'notifications': [],
            'unread_count': 0
        }, status=500)



@login_required
def debug_notifications(request):
    """
    Debug view to check notification setup
    Call this endpoint to diagnose issues
    """
    try:
        from .models import Notification
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        debug_info = {
            'current_user': {
                'username': request.user.username,
                'id': request.user.id,
                'is_superuser': request.user.is_superuser,
                'has_profile': hasattr(request.user, 'userprofile'),
            },
            'user_role': None,
            'notifications': {
                'total': 0,
                'unread': 0,
                'recent': []
            },
            'admin_users': [],
        }
        
        # Get user role
        if hasattr(request.user, 'userprofile'):
            debug_info['user_role'] = request.user.userprofile.role
            if request.user.userprofile.barangay:
                debug_info['barangay'] = request.user.userprofile.barangay.name
        
        # Get notification counts
        notifications = Notification.objects.filter(user=request.user)
        debug_info['notifications']['total'] = notifications.count()
        debug_info['notifications']['unread'] = notifications.filter(is_read=False).count()
        
        # Get recent notifications
        recent = notifications.order_by('-created_at')[:5]
        for notif in recent:
            debug_info['notifications']['recent'].append({
                'id': notif.id,
                'title': notif.title,
                'type': notif.notification_type,
                'is_read': notif.is_read,
                'created_at': str(notif.created_at)
            })
        
        # Get all admin users
        admin_users = User.objects.filter(
            Q(userprofile__role='dilg staff') | Q(is_superuser=True)
        ).distinct()
        
        for admin in admin_users:
            debug_info['admin_users'].append({
                'username': admin.username,
                'id': admin.id,
                'is_superuser': admin.is_superuser,
                'role': admin.userprofile.role if hasattr(admin, 'userprofile') else 'No profile'
            })
        
        return JsonResponse(debug_info, json_dumps_params={'indent': 2})
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

@login_required
@require_POST
def mark_notification_read(request, notification_id):
    """Mark a single notification as read"""
    try:
        notification = Notification.objects.get(
            id=notification_id,
            user=request.user
        )
        
        notification.is_read = True
        notification.save()
        
        # Get updated unread count
        unread_count = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).count()
        
        return JsonResponse({
            'success': True,
            'message': 'Notification marked as read',
            'unread_count': unread_count
        })
        
    except Notification.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Notification not found'
        }, status=404)
        
    except Exception as e:
        import traceback
        print(f"=== MARK NOTIFICATION READ ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_POST
def mark_all_notifications_read(request):
    """Mark all notifications as read for the current user"""
    try:
        updated_count = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).update(is_read=True)
        
        return JsonResponse({
            'success': True,
            'message': f'{updated_count} notifications marked as read',
            'updated_count': updated_count,
            'unread_count': 0
        })
        
    except Exception as e:
        import traceback
        print(f"=== MARK ALL NOTIFICATIONS READ ERROR ===")
        print(f"Error: {str(e)}")
        print(f"Traceback:\n{traceback.format_exc()}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)



@login_required
@require_http_methods(["GET"])
def get_unread_count(request):
    """Get count of unread notifications"""
    try:
        count = Notification.objects.filter(
            user=request.user,
            is_read=False
        ).count()
        
        return JsonResponse({
            'success': True,
            'unread_count': count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    

@login_required
@require_http_methods(["POST"])
def submit_requirement_with_notification(request, submission_id):
    """
    Handle requirement submission and create notifications
    This is the actual endpoint being called: /api/requirements/submission/<id>/submit/
    """
    try:
        from .models import RequirementSubmission
        
        print(f"\n{'='*60}")
        print(f"🚀 SUBMIT REQUIREMENT WITH NOTIFICATION")
        print(f"{'='*60}")
        print(f"👤 User: {request.user.username}")
        print(f"📋 Submission ID: {submission_id}")
        
        # Get submission
        submission = RequirementSubmission.objects.select_related(
            'requirement', 'barangay'
        ).get(
            id=submission_id,
            barangay=request.user.userprofile.barangay  # Fixed query
        )
        
        print(f"📄 Requirement: {submission.requirement.title}")
        print(f"🏘️ Barangay: {submission.barangay.name}")
        
        # Update submission status
        old_status = submission.status
        submission.status = 'accomplished'
        submission.submitted_at = timezone.now()
        submission.submitted_by = request.user
        submission.save()
        
        print(f"✅ Submission updated:")
        print(f"   Old status: {old_status}")
        print(f"   New status: {submission.status}")
        print(f"   Submitted at: {submission.submitted_at}")
        
        # ========== CREATE NOTIFICATIONS FOR ADMINS ==========
        print(f"\n🔔 Creating admin notifications...")
        notifications_sent = notify_admins(
            title=f"📋 New Submission: {submission.requirement.title}",
            message=f"{submission.barangay.name} submitted '{submission.requirement.title}' for review.",
            notification_type='new_submission',
            submission=submission
        )
        # ====================================================
        
        print(f"\n✅ SUBMISSION COMPLETE")
        print(f"   Notifications sent: {notifications_sent}")
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': True,
            'message': 'Requirement submitted successfully',
            'submission': {
                'id': submission.id,
                'status': submission.status,
                'status_display': submission.get_status_display(),
                'submitted_at': submission.submitted_at.strftime('%B %d, %Y')
            },
            'notifications_sent': notifications_sent
        })
        
    except RequirementSubmission.DoesNotExist:
        print(f"❌ ERROR: Submission {submission_id} not found")
        return JsonResponse({
            'success': False,
            'error': 'Submission not found'
        }, status=404)
        
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ SUBMIT ERROR")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        traceback.print_exc()
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    


@login_required
@require_http_methods(["POST"])
def approve_submission_with_notification(request, submission_id):
    """Approve submission and notify submitter"""
    try:
        # Check if user is admin
        if not (request.user.groups.filter(name='Admin').exists() or request.user.is_superuser):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied'
            }, status=403)
        
        submission = RequirementSubmission.objects.get(id=submission_id)
        
        # Get review notes if any
        import json
        data = json.loads(request.body) if request.body else {}
        review_notes = data.get('review_notes', '')
        
        # Update submission
        submission.status = 'approved'
        submission.reviewed_by = request.user
        submission.reviewed_at = timezone.now()
        submission.review_notes = review_notes
        submission.save()
        
        # Create notification for submitter
        message = f"Your submission for {submission.requirement.title} has been approved"
        if review_notes:
            message += f". Admin notes: {review_notes}"
        
        create_notification(
            user=submission.barangay.user,
            title="Submission Approved",
            message=message,
            notification_type='completed',
            submission=submission
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Submission approved successfully'
        })
        
    except RequirementSubmission.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Submission not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_http_methods(["POST"])
def reject_submission_with_notification(request, submission_id):
    """Reject submission and notify submitter"""
    try:
        # Check if user is admin
        if not (request.user.groups.filter(name='Admin').exists() or request.user.is_superuser):
            return JsonResponse({
                'success': False,
                'error': 'Permission denied'
            }, status=403)
        
        submission = RequirementSubmission.objects.get(id=submission_id)
        
        # Get review notes
        import json
        data = json.loads(request.body) if request.body else {}
        review_notes = data.get('review_notes', '')
        
        # Update submission
        submission.status = 'rejected'
        submission.reviewed_by = request.user
        submission.reviewed_at = timezone.now()
        submission.review_notes = review_notes
        submission.save()
        
        # Create notification for submitter
        message = f"Your submission for {submission.requirement.title} needs revision"
        if review_notes:
            message += f". Admin notes: {review_notes}"
        
        create_notification(
            user=submission.barangay.user,
            title="Submission Needs Revision",
            message=message,
            notification_type='overdue',
            submission=submission
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Submission rejected successfully'
        })
        
    except RequirementSubmission.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Submission not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def get_time_ago(datetime_obj):
    """Convert datetime to 'time ago' format"""
    from django.utils import timezone
    now = timezone.now()
    diff = now - datetime_obj
    
    seconds = diff.total_seconds()
    
    if seconds < 60:
        return "Just now"
    elif seconds < 3600:
        minutes = int(seconds / 60)
        return f"{minutes} minute{'s' if minutes > 1 else ''} ago"
    elif seconds < 86400:
        hours = int(seconds / 3600)
        return f"{hours} hour{'s' if hours > 1 else ''} ago"
    elif seconds < 604800:
        days = int(seconds / 86400)
        return f"{days} day{'s' if days > 1 else ''} ago"
    else:
        weeks = int(seconds / 604800)
        return f"{weeks} week{'s' if weeks > 1 else ''} ago"

def create_notification(user, title, message, notification_type, submission=None):
    """Helper function to create notifications"""
    try:
        notification = Notification.objects.create(
            user=user,
            title=title,
            message=message,
            notification_type=notification_type,
            submission=submission
        )
        return notification
    except Exception as e:
        print(f"Error creating notification: {e}")
        return None


#----END OF NOTIFICATIONS HELPERS----



#----CATEGORIZATION----
#----CATEGORIZATION----
@login_required
def folder_view(request):
    """Main folder view showing all categories"""
    categories = FileCategory.objects.all()
    
    # Calculate file counts for each category
    for category in categories:
        category.file_count = CategorizedFile.objects.filter(
            category=category,
            is_archived=False
        ).count()
    
    
    # Count monitoring files
    monitoring_count = CategorizedFile.objects.filter(
        category__name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually'],
        is_archived=False
    ).count()
    
    # Count certification files
    certification_count = CategorizedFile.objects.filter(
        category__name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures'],
        is_archived=False
    ).count()
    
    # Count pending users
    pending_count = User.objects.filter(
        userprofile__is_approved=False
    ).count()
    
    # Count pending applications
    pending_applications_count = EligibilityRequest.objects.filter(
        status='pending',
        archived=False
    ).count()
    
    # DEBUG - Add this temporarily
    print(f"===== FOLDER VIEW DEBUG =====")
    print(f"Monitoring count: {monitoring_count}")
    print(f"Certification count: {certification_count}")
    print(f"Pending users: {pending_count}")
    print(f"Pending applications: {pending_applications_count}")
    
    context = {
        'categories': categories,
        'monitoring_count': monitoring_count,
        'certification_count': certification_count,
        'pending_count': pending_count,
        'pending_applications_count': pending_applications_count,
    }
    
    return render(request, 'folder.html', context)

@login_required
def certification_files_view(request):
    """View for certification files"""
    # Get certificate-related categories
    categories = FileCategory.objects.filter(
        name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures']
    )
    
    for category in categories:
        category.file_count = CategorizedFile.objects.filter(
            category=category,
            is_archived=False
        ).count()
    
    # Count for monitoring files (for sidebar badge)
    monitoring_count = CategorizedFile.objects.filter(
        category__name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually'],
        is_archived=False
    ).count()
    
    # Count for certification files (for sidebar badge)
    certification_count = CategorizedFile.objects.filter(
        category__name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures'],
        is_archived=False
    ).count()
    
    # Count pending user approvals
    pending_count = User.objects.filter(is_active=False).count()
    
    # Count pending applications
    pending_applications_count = EligibilityRequest.objects.filter(
        status='pending',
        archived=False
    ).count()
    
    context = {
        'categories': categories,
        'page_title': 'Certification Files',
        'monitoring_count': monitoring_count,
        'certification_count': certification_count,
        'pending_count': pending_count,
        'pending_applications_count': pending_applications_count,
    }
    return render(request, 'certification_filess.html', context)


@login_required
def monitoring_files_view(request):
    """View for monitoring/requirements files"""
    barangays = Barangay.objects.all()
    
    # Get requirement-related categories
    categories = FileCategory.objects.filter(
        name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually']
    )
    
    for category in categories:
        category.file_count = CategorizedFile.objects.filter(
            category=category,
            is_archived=False
        ).count()
    
    # Count for monitoring files (for sidebar badge)
    monitoring_count = CategorizedFile.objects.filter(
        category__name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually'],
        is_archived=False
    ).count()
    
    # Count for certification files (for sidebar badge)
    certification_count = CategorizedFile.objects.filter(
        category__name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures'],
        is_archived=False
    ).count()
    
    # Count pending user approvals
    pending_count = User.objects.filter(is_active=False).count()
    
    # Count pending applications
    pending_applications_count = EligibilityRequest.objects.filter(
        status='pending',
        archived=False
    ).count()
    
    context = {
        'barangays': barangays,
        'categories': categories,
        'page_title': 'Monitoring Files',
        'monitoring_count': monitoring_count,
        'certification_count': certification_count,
        'pending_count': pending_count,
        'pending_applications_count': pending_applications_count,
    }
    return render(request, 'monitoring_files.html', context)


@login_required
@require_http_methods(["GET"])
def api_category_files(request, category_name):
    """API endpoint to get files by category"""
    try:
        category = get_object_or_404(FileCategory, name=category_name)
        
        # Get filter parameters
        barangay_id = request.GET.get('barangay_id')
        search_query = request.GET.get('search', '').strip()
        file_type = request.GET.get('file_type', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # Base query
        files = CategorizedFile.objects.filter(
            category=category,
            is_archived=False
        ).select_related('barangay', 'uploaded_by')
        
        # Apply filters
        if barangay_id:
            files = files.filter(barangay_id=barangay_id)
        
        if search_query:
            files = files.filter(
                Q(original_filename__icontains=search_query) |
                Q(detected_content__icontains=search_query) |
                Q(tags__icontains=search_query)
            )
        
        if file_type:
            files = files.filter(file_type=file_type)
        
        if date_from:
            files = files.filter(uploaded_at__date__gte=date_from)
        
        if date_to:
            files = files.filter(uploaded_at__date__lte=date_to)
        
        # Paginate
        paginator = Paginator(files, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # Prepare response data
        files_data = []
        for file in page_obj:
            files_data.append({
                'id': file.id,
                'filename': file.original_filename,
                'file_url': file.file.url,
                'file_type': file.file_type,
                'file_size': file.file_size_mb,
                'detected_content': file.detected_content,
                'barangay': file.barangay.name if file.barangay else None,
                'period': file.period,
                'uploaded_at': file.uploaded_at.strftime('%B %d, %Y %I:%M %p'),
                'uploaded_by': file.uploaded_by.get_full_name() if file.uploaded_by else 'System',
                'tags': file.tags,
                'thumbnail': file.get_thumbnail_url(),
            })
        
        return JsonResponse({
            'success': True,
            'files': files_data,
            'total_count': paginator.count,
            'page': page_obj.number,
            'total_pages': paginator.num_pages,
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_upload_file(request):
    """API endpoint for manual file upload"""
    try:
        category_name = request.POST.get('category')
        barangay_id = request.POST.get('barangay_id')
        period = request.POST.get('period', '')
        tags = request.POST.get('tags', '')
        
        if not category_name:
            return JsonResponse({
                'success': False,
                'error': 'Category is required'
            }, status=400)
        
        if 'file' not in request.FILES:
            return JsonResponse({
                'success': False,
                'error': 'No file uploaded'
            }, status=400)
        
        uploaded_file = request.FILES['file']
        
        # Get category
        category = get_object_or_404(FileCategory, name=category_name)
        
        # Determine file type
        mime_type = uploaded_file.content_type
        if mime_type.startswith('image/'):
            file_type = 'image'
        elif mime_type == 'application/pdf':
            file_type = 'pdf'
        elif mime_type.startswith('application/'):
            file_type = 'document'
        else:
            file_type = 'other'
        
        # Create categorized file
        categorized_file = CategorizedFile.objects.create(
            file=uploaded_file,
            original_filename=uploaded_file.name,
            file_type=file_type,
            file_size=uploaded_file.size,
            mime_type=mime_type,
            category=category,
            source='manual',
            barangay_id=barangay_id if barangay_id else None,
            period=period,
            uploaded_by=request.user,
            tags=tags,
        )
        
        # Update category file count
        category.update_file_count()
        
        # Log the upload
        try:
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                content_object=categorized_file,
                description=f"Uploaded file to {category.display_name}: {uploaded_file.name}"
            )
        except:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'File uploaded successfully',
            'file': {
                'id': categorized_file.id,
                'filename': categorized_file.original_filename,
                'file_url': categorized_file.file.url,
                'file_size': categorized_file.file_size_mb,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@require_http_methods(["DELETE"])
def api_delete_file(request, file_id):
    """
    FIXED: API endpoint to delete a file
    """
    try:
        print(f"\n{'='*50}")
        print(f"DELETE FILE - ID: {file_id}")
        print(f"{'='*50}")
        
        file = MonitoringFile.objects.get(id=file_id)
        filename = file.filename
        print(f"Found file: {filename}")
        
        # Delete the actual file
        try:
            if file.file:
                file.file.delete(save=False)
                print(f"✓ Deleted physical file")
        except Exception as file_del_error:
            print(f"✗ Could not delete physical file: {file_del_error}")
        
        # Delete database record
        file.delete()
        print(f"✓ Deleted database record")
        print(f"{'='*50}\n")
        
        return JsonResponse({
            'success': True,
            'message': f'File "{filename}" deleted successfully'
        })
        
    except MonitoringFile.DoesNotExist:
        print(f"✗ File {file_id} not found in database")
        return JsonResponse({
            'success': False,
            'error': 'File not found'
        }, status=404)
        
    except Exception as e:
        print(f"\n{'='*50}")
        print(f"ERROR IN DELETE FILE")
        print(f"{'='*50}")
        print(f"Error: {str(e)}")
        print(traceback.format_exc())
        print(f"{'='*50}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_archive_file(request, file_id):
    """API endpoint to archive a file"""
    try:
        file = get_object_or_404(CategorizedFile, id=file_id)
        
        file.archive()
        
        # Update category count
        file.category.update_file_count()
        
        return JsonResponse({
            'success': True,
            'message': 'File archived successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def api_move_file(request, file_id):
    """API endpoint to move file to different category"""
    try:
        file = get_object_or_404(CategorizedFile, id=file_id)
        
        data = json.loads(request.body)
        new_category_name = data.get('category')
        
        if not new_category_name:
            return JsonResponse({
                'success': False,
                'error': 'Category is required'
            }, status=400)
        
        new_category = get_object_or_404(FileCategory, name=new_category_name)
        old_category = file.category
        
        file.category = new_category
        file.save()
        
        # Update both category counts
        old_category.update_file_count()
        new_category.update_file_count()
        
        return JsonResponse({
            'success': True,
            'message': f'File moved to {new_category.display_name}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def api_file_statistics(request):
    """API endpoint to get file statistics"""
    try:
        stats = {
            'total_files': CategorizedFile.objects.filter(is_archived=False).count(),
            'total_size_mb': round(
                CategorizedFile.objects.filter(is_archived=False).aggregate(
                    total=Sum('file_size')
                )['total'] / (1024 * 1024), 2
            ) if CategorizedFile.objects.exists() else 0,
            'by_category': [],
            'by_type': [],
            'recent_uploads': [],
        }
        
        # Stats by category
        for category in FileCategory.objects.all():
            count = CategorizedFile.objects.filter(
                category=category,
                is_archived=False
            ).count()
            if count > 0:
                stats['by_category'].append({
                    'name': category.display_name,
                    'count': count
                })
        
        # Stats by file type
        type_counts = CategorizedFile.objects.filter(
            is_archived=False
        ).values('file_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        stats['by_type'] = list(type_counts)
        
        # Recent uploads
        recent = CategorizedFile.objects.filter(
            is_archived=False
        ).order_by('-uploaded_at')[:5]
        
        for file in recent:
            stats['recent_uploads'].append({
                'filename': file.original_filename,
                'category': file.category.display_name,
                'uploaded_at': file.uploaded_at.strftime('%B %d, %Y'),
            })
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)




@require_http_methods(["GET"])
def get_files_by_category(request, category):
    """
    FIXED: API endpoint to get files by category
    URL: /api/files/category/<category>/
    """
    try:
        print(f"\n{'='*50}")
        print(f"GET FILES BY CATEGORY - START")
        print(f"{'='*50}")
        print(f"Category: {category}")
        print(f"Request method: {request.method}")
        print(f"GET params: {dict(request.GET)}")
        
        # Get query parameters
        barangay_id = request.GET.get('barangay_id', None)
        date_from = request.GET.get('date_from', None)
        date_to = request.GET.get('date_to', None)
        
        # Check if MonitoringFile model exists
        try:
            # Test query to check if table exists
            test_count = MonitoringFile.objects.count()
            print(f"✓ MonitoringFile table exists. Total records: {test_count}")
        except Exception as model_error:
            print(f"✗ MonitoringFile table error: {model_error}")
            # Return empty result gracefully
            return JsonResponse({
                'success': True,
                'files': [],
                'total_count': 0,
                'category': category,
                'message': 'No files table found'
            })
        
        # Base query
        files = MonitoringFile.objects.filter(category=category)
        print(f"Files in category '{category}': {files.count()}")
        
        # Apply filters
        if barangay_id:
            files = files.filter(barangay_id=barangay_id)
            print(f"After barangay filter ({barangay_id}): {files.count()}")
        
        if date_from:
            files = files.filter(uploaded_at__gte=date_from)
            print(f"After date_from filter: {files.count()}")
        
        if date_to:
            files = files.filter(uploaded_at__lte=date_to)
            print(f"After date_to filter: {files.count()}")
        
        # Order by most recent
        files = files.order_by('-uploaded_at')
        total_count = files.count()
        
        # Serialize files
        files_data = []
        for file_obj in files:
            try:
                # Build file data safely
                file_data = {
                    'id': file_obj.id,
                    'filename': getattr(file_obj, 'filename', 'Unknown'),
                    'file_url': '',
                    'file_type': 'unknown',
                    'file_size': 0,
                    'barangay': 'N/A',
                    'uploaded_at': ''
                }
                
                # Get file URL
                try:
                    if hasattr(file_obj, 'file') and file_obj.file:
                        file_data['file_url'] = file_obj.file.url
                        # Get file size
                        try:
                            file_data['file_size'] = round(file_obj.file.size / (1024 * 1024), 2)
                        except:
                            pass
                except:
                    pass
                
                # Get file type
                if file_obj.filename:
                    ext = file_obj.filename.lower().split('.')[-1]
                    if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg', 'webp']:
                        file_data['file_type'] = 'image'
                    else:
                        file_data['file_type'] = 'document'
                
                # Get barangay name
                try:
                    if hasattr(file_obj, 'barangay') and file_obj.barangay:
                        file_data['barangay'] = file_obj.barangay.name
                except:
                    pass
                
                # Get upload date
                try:
                    if hasattr(file_obj, 'uploaded_at') and file_obj.uploaded_at:
                        file_data['uploaded_at'] = file_obj.uploaded_at.strftime('%Y-%m-%d %H:%M')
                except:
                    pass
                
                files_data.append(file_data)
                
            except Exception as file_error:
                print(f"✗ Error processing file {file_obj.id}: {file_error}")
                continue
        
        print(f"✓ Successfully processed {len(files_data)} files")
        print(f"{'='*50}\n")
        
        return JsonResponse({
            'success': True,
            'files': files_data,
            'total_count': total_count,
            'category': category
        })
        
    except Exception as e:
        print(f"\n{'='*50}")
        print(f"ERROR IN GET_FILES_BY_CATEGORY")
        print(f"{'='*50}")
        print(f"Error: {str(e)}")
        print(f"Error type: {type(e).__name__}")
        print(f"Traceback:")
        print(traceback.format_exc())
        print(f"{'='*50}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'category': category
        }, status=500)


def get_file_type(filename):
    """Helper function to determine file type"""
    try:
        ext = filename.lower().split('.')[-1]
        image_extensions = ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg', 'webp']
        
        if ext in image_extensions:
            return 'image'
        return 'document'
    except:
        return 'unknown'

def test_monitoring_api(request):
    """
    Temporary test endpoint to verify setup
    URL: /test-monitoring/
    """
    try:
        # Test database connection
        from .models import MonitoringFile, Barangay
        
        monitoring_count = MonitoringFile.objects.count()
        barangay_count = Barangay.objects.count()
        
        # Get sample records
        sample_files = list(MonitoringFile.objects.values(
            'id', 'filename', 'category', 'barangay__name'
        )[:5])
        
        categories = MonitoringFile.objects.values_list('category', flat=True).distinct()
        
        return JsonResponse({
            'status': 'success',
            'database': 'connected',
            'monitoring_files_count': monitoring_count,
            'barangays_count': barangay_count,
            'categories': list(categories),
            'sample_files': sample_files,
        }, json_dumps_params={'indent': 2})
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500, json_dumps_params={'indent': 2})
    
@require_http_methods(["GET"])
def get_files_by_category_simple(request, category):
    """
    Get files by category (weekly, monthly, quarterly, semestral, annually)
    Files are automatically categorized when uploaded via requirements_monitoring
    """
    try:
        print(f"\n{'='*60}")
        print(f"📁 GET FILES BY CATEGORY: {category}")
        print(f"{'='*60}")
        
        # Get query parameters
        barangay_id = request.GET.get('barangay_id')
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        page = int(request.GET.get('page', 1))
        
        print(f"Filters: barangay_id={barangay_id}, date_from={date_from}, date_to={date_to}")
        
        # Query CategorizedFile using period field
        # REMOVED is_archived filter since it doesn't exist
        files = CategorizedFile.objects.filter(
            period=category  # weekly, monthly, quarterly, semestral, annually
        ).select_related('barangay', 'uploaded_by', 'requirement_submission')
        
        print(f"Files with period '{category}': {files.count()}")
        
        # Apply filters
        if barangay_id:
            files = files.filter(barangay_id=barangay_id)
            print(f"After barangay filter: {files.count()}")
        
        if date_from:
            files = files.filter(uploaded_at__gte=date_from)
            print(f"After date_from filter: {files.count()}")
        
        if date_to:
            files = files.filter(uploaded_at__lte=date_to)
            print(f"After date_to filter: {files.count()}")
        
        # Order by newest first
        files = files.order_by('-uploaded_at')
        total_count = files.count()
        
        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(files, 20)
        page_obj = paginator.get_page(page)
        
        print(f"Total count: {total_count}, Page: {page}/{paginator.num_pages}")
        
        # Build response
        files_data = []
        for file_obj in page_obj:
            try:
                files_data.append({
                    'id': file_obj.id,
                    'filename': file_obj.original_filename,
                    'file_url': file_obj.file.url if file_obj.file else '',
                    'file_type': file_obj.file_type,
                    'file_size': file_obj.file_size_mb,
                    'barangay': file_obj.barangay.name if file_obj.barangay else 'N/A',
                    'uploaded_at': file_obj.uploaded_at.strftime('%B %d, %Y %I:%M %p'),
                    'uploaded_by': file_obj.uploaded_by.get_full_name() if file_obj.uploaded_by else 'System',
                    'detected_content': file_obj.detected_content or 'N/A',
                    'requirement_title': (
                        file_obj.requirement_submission.requirement.title 
                        if file_obj.requirement_submission 
                        else 'N/A'
                    ),
                    'tags': file_obj.tags or '',
                })
            except Exception as file_err:
                print(f"✗ Error processing file {file_obj.id}: {file_err}")
                continue
        
        print(f"✓ Successfully processed {len(files_data)} files")
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': True,
            'files': files_data,
            'total_count': total_count,
            'page': page,
            'total_pages': paginator.num_pages,
            'category': category
        })
        
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ ERROR IN get_files_by_category_simple")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc())
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'category': category
        }, status=500)
    

@require_http_methods(["DELETE"])  
def api_delete_monitoring_file(request, file_id):
    """
    Delete a categorized file
    Works for files uploaded via requirements_monitoring
    """
    try:
        print(f"\n{'='*60}")
        print(f"🗑️ DELETE FILE REQUEST")
        print(f"{'='*60}")
        print(f"File ID: {file_id}")
        print(f"User: {request.user}")
        
        # Get the file from CategorizedFile
        file = CategorizedFile.objects.get(id=file_id)
        filename = file.original_filename
        category = file.category.display_name if file.category else 'Unknown'
        
        print(f"Found file: {filename}")
        print(f"Category: {category}")
        
        # Delete physical file from storage
        try:
            if file.file:
                file.file.delete(save=False)
                print(f"✓ Deleted physical file from storage")
        except Exception as storage_err:
            print(f"⚠️ Could not delete physical file: {storage_err}")
        
        # Delete the RequirementAttachment if it exists
        if file.requirement_attachment:
            try:
                file.requirement_attachment.delete()
                print(f"✓ Deleted linked RequirementAttachment")
            except Exception as att_err:
                print(f"⚠️ Could not delete RequirementAttachment: {att_err}")
        
        # Delete database record
        file.delete()
        print(f"✓ Deleted from database: {filename}")
        
        # Update category file count
        if file.category:
            file.category.update_file_count()
            print(f"✓ Updated category file count")
        
        # Log the deletion
        try:
            AuditLog.objects.create(
                user=request.user,
                action='DELETE',
                description=f"Deleted file: {filename} from {category}"
            )
        except:
            pass
        
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': True,
            'message': f'File "{filename}" deleted successfully'
        })
        
    except CategorizedFile.DoesNotExist:
        print(f"✗ File {file_id} not found in CategorizedFile table")
        print(f"{'='*60}\n")
        return JsonResponse({
            'success': False,
            'error': 'File not found'
        }, status=404)
        
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ ERROR IN api_delete_monitoring_file")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        print(traceback.format_exc())
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

    



@require_http_methods(["GET"])
def get_category_file_counts(request):
    """
    Get file counts for all categories
    Useful for updating the monitoring_files.html folder counts
    """
    try:
        categories = ['weekly', 'monthly', 'quarterly', 'semestral', 'annually']
        barangay_id = request.GET.get('barangay_id')
        
        counts = {}
        for category in categories:
            query = CategorizedFile.objects.filter(
                period=category,
                is_archived=False
            )
            
            if barangay_id:
                query = query.filter(barangay_id=barangay_id)
            
            counts[category] = query.count()
        
        return JsonResponse({
            'success': True,
            'counts': counts
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


#-----NEW DASHBOARD MERGE USERS TEMPLATES AND MIGRATION-----

@login_required
def barangay_dashboard(request):
    """
    Unified dashboard for Barangay Users
    Shows: Eligibility, Requirements, Analytics
    """
    user_profile = request.user.userprofile
    
    # Access control
    if user_profile.role != 'barangay user':
        messages.error(request, '🚫 Access Denied')
        return redirect('dashboard')
    
    # Get assigned barangay
    assigned_barangay = user_profile.barangay
    
    # Get notification counts for sidebar
    monitoring_count = CategorizedFile.objects.filter(
        category__name__in=['weekly', 'monthly', 'quarterly', 'semestral', 'annually'],
        is_archived=False
    ).count()
    
    certification_count = CategorizedFile.objects.filter(
        category__name__in=['appointive_certificates', 'elective_certificates', 'ids', 'signatures'],
        is_archived=False
    ).count()
    
    pending_count = User.objects.filter(
        userprofile__is_approved=False
    ).count()
    
    pending_applications_count = EligibilityRequest.objects.filter(
        status='pending',
        archived=False
    ).count()
    
    # Get user's submissions
    my_submissions = []
    barangay_stats = {
        'total_requirements': 0,
        'completed': 0,
        'pending': 0,
        'overdue': 0,
        'total_change': 0,
        'completed_change': 0,
        'pending_change': 0,
        'overdue_change': 0,
        'total_chart': [0] * 12,
        'completed_chart': [0] * 12,
        'pending_chart': [0] * 12,
        'overdue_chart': [0] * 12,
    }
    
    if assigned_barangay:
        from datetime import timedelta
        
        # Current date calculations
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        
        # Get all submissions for this barangay
        submissions = RequirementSubmission.objects.filter(barangay=assigned_barangay)
        
        # CURRENT WEEK STATS
        current_total = submissions.count()
        current_completed = submissions.filter(status='approved').count()
        current_pending = submissions.filter(status='pending').count()
        current_overdue = submissions.filter(
            status__in=['pending', 'in_progress'],
            due_date__lt=today
        ).count()
        
        # LAST WEEK STATS (for comparison)
        last_week_submissions = RequirementSubmission.objects.filter(
            barangay=assigned_barangay,
            created_at__date__lte=week_ago
        )
        
        last_week_total = last_week_submissions.count()
        last_week_completed = last_week_submissions.filter(status='approved').count()
        last_week_pending = last_week_submissions.filter(status='pending').count()
        last_week_overdue = last_week_submissions.filter(
            status__in=['pending', 'in_progress'],
            due_date__lt=week_ago
        ).count()
        
        # CALCULATE PERCENTAGE CHANGES
        def calc_percentage_change(current, previous):
            if previous == 0:
                return 100.0 if current > 0 else 0.0
            change = ((current - previous) / previous) * 100
            return round(change, 1)
        
        # GENERATE CHART DATA (last 12 days)
        def get_chart_data(status_filter=None):
            chart_data = []
            for i in range(11, -1, -1):  # 12 days ago to today
                day = today - timedelta(days=i)
                
                if status_filter == 'total':
                    count = RequirementSubmission.objects.filter(
                        barangay=assigned_barangay,
                        created_at__date=day
                    ).count()
                elif status_filter == 'completed':
                    count = RequirementSubmission.objects.filter(
                        barangay=assigned_barangay,
                        status='approved',
                        updated_at__date=day
                    ).count()
                elif status_filter == 'pending':
                    count = RequirementSubmission.objects.filter(
                        barangay=assigned_barangay,
                        status='pending',
                        created_at__date=day
                    ).count()
                elif status_filter == 'overdue':
                    count = RequirementSubmission.objects.filter(
                        barangay=assigned_barangay,
                        status__in=['pending', 'in_progress'],
                        due_date__lt=day
                    ).count()
                else:
                    count = 0
                
                chart_data.append(max(count, 1))  # Minimum 1 for visibility
            
            return chart_data
        
        # Get recent submissions for table
        my_submissions = RequirementSubmission.objects.filter(
            barangay=assigned_barangay
        ).select_related('requirement').order_by('-due_date')[:5]
        
        # BUILD STATS DICTIONARY
        barangay_stats = {
            # Current counts
            'total_requirements': current_total,
            'completed': current_completed,
            'pending': current_pending,
            'overdue': current_overdue,
            
            # Percentage changes
            'total_change': calc_percentage_change(current_total, last_week_total),
            'completed_change': calc_percentage_change(current_completed, last_week_completed),
            'pending_change': calc_percentage_change(current_pending, last_week_pending),
            'overdue_change': calc_percentage_change(current_overdue, last_week_overdue),
            
            # Chart data (last 12 days)
            'total_chart': get_chart_data('total'),
            'completed_chart': get_chart_data('completed'),
            'pending_chart': get_chart_data('pending'),
            'overdue_chart': get_chart_data('overdue'),
        }
    
    # Get all barangays for monitoring view
    all_barangays = Barangay.objects.all().order_by('name')
    
    context = {
        'user_profile': user_profile,
        'assigned_barangay': assigned_barangay,
        'all_barangays': all_barangays,
        'my_submissions': my_submissions,
        'barangay_stats': barangay_stats,
        'monitoring_count': monitoring_count,
        'certification_count': certification_count,
        'pending_count': pending_count,
        'pending_applications_count': pending_applications_count,
    }
    
    return render(request, 'barangay_dashboard.html', context)





def send_test_email(request):
    """
    Test endpoint to verify email configuration
    Usage: /test-email/
    """
    try:
        from django.http import JsonResponse
        from django.core.mail import send_mail
        from django.conf import settings
        
        print(f"\n{'='*60}")
        print(f"🧪 TESTING EMAIL CONFIGURATION")
        print(f"{'='*60}")
        print(f"EMAIL_BACKEND: {settings.EMAIL_BACKEND}")
        print(f"EMAIL_HOST: {settings.EMAIL_HOST}")
        print(f"EMAIL_PORT: {settings.EMAIL_PORT}")
        print(f"EMAIL_HOST_USER: {settings.EMAIL_HOST_USER}")
        print(f"DEFAULT_FROM_EMAIL: {settings.DEFAULT_FROM_EMAIL}")
        
        test_email = "choimeramae@gmail.com"  # Your email
        
        print(f"\n📤 Sending test email to: {test_email}")
        
        send_mail(
            subject="🧪 DILG Test Email - Configuration Working",
            message="""This is a test email from the DILG Certification System.

If you received this email, your email configuration is working correctly!

Configuration Details:
- SMTP Host: smtp.gmail.com
- Port: 587
- TLS: Enabled
- From: choimeramae@gmail.com

Test sent at: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}

---
DILG Lucena - Certification System
This is an automated test message.
""",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[test_email],
            fail_silently=False,
        )
        
        print(f"✅ Test email sent successfully!")
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': True,
            'message': f'Test email sent to {test_email}',
            'config': {
                'host': settings.EMAIL_HOST,
                'port': settings.EMAIL_PORT,
                'user': settings.EMAIL_HOST_USER,
                'from': settings.DEFAULT_FROM_EMAIL,
            }
        })
        
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"❌ EMAIL TEST FAILED")
        print(f"{'='*60}")
        print(f"Error: {str(e)}")
        print(traceback.format_exc())
        print(f"{'='*60}\n")
        
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)
    
@login_required
@require_http_methods(["GET"])
def api_requirements_trend_data(request):
    try:
        from datetime import timedelta
        from django.db.models import Count, Q
        
        # Get time interval from query params (default 30 days)
        days = int(request.GET.get('days', 30))
        
        # Validate days parameter
        if days not in [7, 30, 90]:
            days = 30
        
        print(f"\n📊 REQUIREMENTS TREND DATA API")
        print(f"   Days requested: {days}")
        print(f"   User: {request.user.username}")
        
        # Calculate date range
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days-1)
        
        print(f"   Date range: {start_date} to {end_date}")
        
        # Initialize data arrays
        labels = []
        completed_data = []
        pending_data = []
        total_data = []

        for i in range(days):
            current_date = start_date + timedelta(days=i)
            labels.append(current_date.strftime('%Y-%m-%d'))
            completed_count = RequirementSubmission.objects.filter(
                status='accomplished',  
                created_at__date=current_date  
            ).count()

            pending_count = RequirementSubmission.objects.filter(
                status__in=['pending', 'in_progress'],
                created_at__date=current_date
            ).count()

            total_count = RequirementSubmission.objects.filter(
                created_at__date=current_date
            ).count()
            
            completed_data.append(completed_count)
            pending_data.append(pending_count)
            total_data.append(total_count)
        
        print(f"   ✓ Data points generated: {len(labels)}")
        print(f"   ✓ Total submissions: {sum(total_data)}")
        print(f"   ✓ Completed: {sum(completed_data)}")
        print(f"   ✓ Pending: {sum(pending_data)}")
        
        return JsonResponse({
            'success': True,
            'labels': labels,
            'completed': completed_data,
            'pending': pending_data,
            'total': total_data,
            'period': f'{days} days'
        })
        
    except Exception as e:
        print(f"❌ Error in api_requirements_trend_data: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    


from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from datetime import datetime, timedelta
import logging

logger = logging.getLogger(__name__)

@login_required
def api_eligibility_analytics(request):
    """
    FIXED: API endpoint for eligibility certifications radar chart data
    Prevents duplicate counting by using aggregation
    """
    try:
        logger.info("📊 ELIGIBILITY ANALYTICS API (FIXED)")
        logger.info(f"   User: {request.user.username}")
        
        # Get barangay
        barangay = None
        
        try:
            if hasattr(request.user, 'userprofile'):
                barangay = request.user.userprofile.barangay
                logger.info(f"   Found barangay via userprofile: {barangay}")
        except Exception as e:
            logger.warning(f"   UserProfile method failed: {str(e)}")
        
        if not barangay:
            logger.warning("   No barangay found - returning empty data")
            return JsonResponse({
                'success': True,
                'labels': ['Total'],
                'appointive': [0],
                'elective_completed': [0],
                'elective_incomplete': [0],
                'totals': {
                    'appointive': 0,
                    'elective_completed': 0,
                    'elective_incomplete': 0,
                    'grand_total': 0
                }
            })
        
        barangay_name = barangay.name
        logger.info(f"   ✓ Barangay name: {barangay_name}")
        
        from app.models import EligibilityRequest
        
        # ============================================
        # FIX: Use single aggregated query per type
        # ============================================
        
        # Get base queryset for this barangay
        base_qs = EligibilityRequest.objects.filter(barangay=barangay_name)
        
        logger.info(f"   📊 Base queryset count: {base_qs.count()}")
        
        # ========== APPOINTIVE - Single Query ==========
        appointive_qs = base_qs.filter(position_type='appointive')
        appointive_total = appointive_qs.count()
        
        logger.info(f"   📌 APPOINTIVE:")
        logger.info(f"      Total: {appointive_total}")
        logger.info(f"      IDs: {list(appointive_qs.values_list('id', flat=True))}")
        
        # ========== ELECTIVE COMPLETED - Single Query ==========
        elective_completed_qs = base_qs.filter(
            position_type='elective',
            completed_term='yes'
        )
        elective_completed_total = elective_completed_qs.count()
        
        logger.info(f"   📌 ELECTIVE - COMPLETED:")
        logger.info(f"      Total: {elective_completed_total}")
        logger.info(f"      IDs: {list(elective_completed_qs.values_list('id', flat=True))}")
        
        # ========== ELECTIVE INCOMPLETE - Single Query ==========
        elective_incomplete_qs = base_qs.filter(
            position_type='elective',
            completed_term='no'
        )
        elective_incomplete_total = elective_incomplete_qs.count()
        
        logger.info(f"   📌 ELECTIVE - INCOMPLETE:")
        logger.info(f"      Total: {elective_incomplete_total}")
        logger.info(f"      IDs: {list(elective_incomplete_qs.values_list('id', flat=True))}")
        
        # ========== VERIFY NO OVERLAP ==========
        grand_total = appointive_total + elective_completed_total + elective_incomplete_total
        
        logger.info(f"\n   🔍 VERIFICATION:")
        logger.info(f"      Appointive: {appointive_total}")
        logger.info(f"      Elective Completed: {elective_completed_total}")
        logger.info(f"      Elective Incomplete: {elective_incomplete_total}")
        logger.info(f"      Sum: {grand_total}")
        logger.info(f"      Base Total: {base_qs.count()}")
        
        if grand_total != base_qs.count():
            logger.warning(f"   ⚠️ MISMATCH: Sum ({grand_total}) != Base ({base_qs.count()})")
        else:
            logger.info(f"   ✅ COUNTS MATCH")
        
        # ========== SIMPLIFIED RESPONSE ==========
        # Just return totals (not breaking down by status)
        labels = ['Total']
        
        return JsonResponse({
            'success': True,
            'labels': labels,
            'appointive': [appointive_total],
            'elective_completed': [elective_completed_total],
            'elective_incomplete': [elective_incomplete_total],
            'totals': {
                'appointive': appointive_total,
                'elective_completed': elective_completed_total,
                'elective_incomplete': elective_incomplete_total,
                'grand_total': grand_total
            },
            'debug': {
                'barangay': barangay_name,
                'base_count': base_qs.count(),
                'appointive_ids': list(appointive_qs.values_list('id', flat=True)),
                'elective_completed_ids': list(elective_completed_qs.values_list('id', flat=True)),
                'elective_incomplete_ids': list(elective_incomplete_qs.values_list('id', flat=True)),
            }
        })
        
    except Exception as e:
        logger.error(f"❌ Error in eligibility analytics API: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        
        return JsonResponse({
            'success': True,
            'labels': ['Total'],
            'appointive': [0],
            'elective_completed': [0],
            'elective_incomplete': [0],
            'totals': {
                'appointive': 0,
                'elective_completed': 0,
                'elective_incomplete': 0,
                'grand_total': 0
            }
        })

@login_required
def debug_eligibility_data(request):
    """
    Debug endpoint to check for duplicate data
    Access at: /api/debug-eligibility/
    """
    try:
        from app.models import EligibilityRequest
        
        barangay = request.user.userprofile.barangay
        if not barangay:
            return JsonResponse({'error': 'No barangay assigned'})
        
        # Get all records
        all_records = EligibilityRequest.objects.filter(
            barangay=barangay.name
        ).values('id', 'first_name', 'last_name', 'position_type', 'completed_term', 'status')
        
        # Group by type
        appointive = []
        elective_completed = []
        elective_incomplete = []
        
        for record in all_records:
            if record['position_type'] == 'appointive':
                appointive.append(record)
            elif record['position_type'] == 'elective':
                if record['completed_term'] == 'yes':
                    elective_completed.append(record)
                elif record['completed_term'] == 'no':
                    elective_incomplete.append(record)
        
        return JsonResponse({
            'barangay': barangay.name,
            'total_records': len(all_records),
            'appointive': {
                'count': len(appointive),
                'records': appointive
            },
            'elective_completed': {
                'count': len(elective_completed),
                'records': elective_completed
            },
            'elective_incomplete': {
                'count': len(elective_incomplete),
                'records': elective_incomplete
            },
            'sum_check': len(appointive) + len(elective_completed) + len(elective_incomplete),
            'matches_total': (len(appointive) + len(elective_completed) + len(elective_incomplete)) == len(all_records)
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

@login_required
def debug_user_barangay(request):
    """Temporary debug endpoint"""
    info = {
        'username': request.user.username,
        'has_userprofile': hasattr(request.user, 'userprofile'),
        'has_barangay': hasattr(request.user, 'barangay'),
        'has_assigned_barangay': hasattr(request.user, 'assigned_barangay'),
        'user_attrs': dir(request.user),
    }
    
    if hasattr(request.user, 'userprofile'):
        info['userprofile_barangay'] = str(getattr(request.user.userprofile, 'assigned_barangay', 'None'))
    
    return JsonResponse(info, safe=False)


# CORRECTED VIEW - Add this to your views.py

from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from app.models import EligibilityRequest

@login_required
def eligibility_certifications_chart_data(request):
    """
    API endpoint for Eligibility Certifications Analytics radar chart
    Returns counts of Appointive, Elective-Completed, and Elective-Incomplete
    
    HANDLES ALL VARIATIONS of completed_term field:
    - "Yes", "yes"
    - "Yes - Full term served"
    - "Yes - Full term successfully completed"
    - "No", "no"
    - "No - Incomplete term"
    - "No - Term not completed"
    """
    try:
        user = request.user
        
        # Base queryset (adjust if you need barangay filtering)
        base_queryset = EligibilityRequest.objects.all()
        
        print(f"📊 ELIGIBILITY CERTIFICATIONS ANALYTICS")
        print(f"   User: {user.username}")
        print(f"   Total requests: {base_queryset.count()}")
        
        # Count Appointive (approved)
        appointive_count = base_queryset.filter(
            position_type='appointive',
            status='approved'
        ).count()
        
        # Count Elective - Completed
        # Use __icontains to match "Yes" in any variation
        elective_completed_count = base_queryset.filter(
            position_type='elective',
            status='approved',
            completed_term__icontains='Yes'  # Matches: "Yes", "Yes - Full term...", etc.
        ).count()
        
        # Count Elective - Incomplete
        # Matches pending status OR any "No" variation
        elective_incomplete_count = base_queryset.filter(
            position_type='elective'
        ).filter(
            Q(status='pending') | 
            Q(status='processing') |
            Q(completed_term__icontains='No')  # Matches: "No", "No - Incomplete...", etc.
        ).count()
        
        print(f"   Appointive: {appointive_count}")
        print(f"   Elective - Completed: {elective_completed_count}")
        print(f"   Elective - Incomplete: {elective_incomplete_count}")
        
        return JsonResponse({
            'success': True,
            'appointive': appointive_count,
            'elective_completed': elective_completed_count,
            'elective_incomplete': elective_incomplete_count,
            'total': appointive_count + elective_completed_count + elective_incomplete_count
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"❌ ELIGIBILITY CHART ERROR:")
        print(error_details)
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# EXPECTED RESULTS:
# ==================
# Appointive: 531
# Elective - Completed: 570 (138 + 386 + 41 + 5)
# Elective - Incomplete: 82 (21 pending + 21 processing + 59 + 15 + 3 "No" variations)
# (Numbers will be slightly different based on your exact data)

from django.http import JsonResponse
from django.db.models import Count
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
@login_required
def requirements_radar_chart_data(request):
    """
    API endpoint for radar chart data showing requirements by type over months
    """
    try:
        # Get user's barangay through UserProfile
        user = request.user
        
        if not hasattr(user, 'userprofile'):
            return JsonResponse({
                'success': False,
                'error': 'No user profile found'
            })
        
        barangay = user.userprofile.barangay
        
        if not barangay:
            return JsonResponse({
                'success': False,
                'error': 'No barangay assigned to this user'
            })
        
        print(f"📊 RADAR CHART API")
        print(f"   User: {user.username}")
        print(f"   Barangay: {barangay.name}")
        
        # Use timezone-aware datetime
        current_date = timezone.now()
        
        # Initialize data structures
        labels = []
        period_data = {
            'weekly': [],
            'monthly': [],
            'quarterly': [],
            'semestral': [],
            'annually': []
        }
        
        # Calculate data for last 8 months
        for month_offset in range(-7, 1):  # -7 to 0 = 8 months
            # Calculate target month and year
            target_month = current_date.month + month_offset
            target_year = current_date.year
            
            # Handle year wrapping
            while target_month <= 0:
                target_month += 12
                target_year -= 1
            while target_month > 12:
                target_month -= 12
                target_year += 1
            
            # Get month boundaries
            month_start = datetime(target_year, target_month, 1, 0, 0, 0)
            
            if target_month == 12:
                next_month = 1
                next_year = target_year + 1
            else:
                next_month = target_month + 1
                next_year = target_year
            
            month_end = datetime(next_year, next_month, 1, 0, 0, 0)
            
            # Make timezone aware
            month_start = timezone.make_aware(month_start)
            month_end = timezone.make_aware(month_end)
            
            # Add month label
            month_label = month_start.strftime('%b')
            labels.append(month_label)
            
            # ✅ FIX: Count submissions by SUBMITTED_AT date, not requirement created_at
            for period in period_data.keys():
                count = RequirementSubmission.objects.filter(
                    barangay=barangay,
                    requirement__period=period,
                    submitted_at__gte=month_start,  # ← Use submitted_at
                    submitted_at__lt=month_end       # ← Not requirement.created_at
                ).count()
                period_data[period].append(count)
        
        # Calculate total submissions
        total_submissions = RequirementSubmission.objects.filter(
            barangay=barangay
        ).count()
        
        # Debug output
        print(f"   Months: {labels}")
        print(f"   Total submissions: {total_submissions}")
        print(f"   Weekly: {period_data['weekly']}")
        print(f"   Monthly: {period_data['monthly']}")
        print(f"   Quarterly: {period_data['quarterly']}")
        print(f"   Semestral: {period_data['semestral']}")
        print(f"   Annually: {period_data['annually']}")
        
        # Create response
        response = JsonResponse({
            'success': True,
            'labels': labels,
            'weekly': period_data['weekly'],
            'monthly': period_data['monthly'],
            'quarterly': period_data['quarterly'],
            'semestral': period_data['semestral'],
            'annually': period_data['annually']
        })
        
        # Add no-cache headers
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        
        return response
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"❌ RADAR CHART ERROR:")
        print(error_details)
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)



@login_required
@require_http_methods(["GET"])
def list_officials(request):
    """Get all officials for the logged-in secretary"""
    try:
        officials = BarangayOfficial.objects.filter(secretary=request.user)
        
        officials_data = []
        for official in officials:
            officials_data.append({
                'id': official.id,
                'first_name': official.first_name,
                'middle_name': official.middle_name or '',
                'last_name': official.last_name,
                'suffix': official.suffix or '',
                'full_name': official.get_full_name(),
                'display_name': official.get_display_name(),
                'position': official.position,
                'position_type': official.position_type,
                'term_start': official.term_start.strftime('%Y-%m-%d'),
                'term_end': official.term_end.strftime('%Y-%m-%d'),
                'years_served': str(official.years_served) if official.years_served else '0',
                'term_status': official.term_status,
                'email': official.email or '',
                'phone': official.phone or '',
                'notes': official.notes or '',
                'is_term_active': official.is_term_active,
                'days_remaining': official.days_remaining,
                'created_at': official.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            })
        
        return JsonResponse({
            'success': True,
            'officials': officials_data,
            'total': len(officials_data)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def create_official(request):
    """Create a new official"""
    try:
        data = json.loads(request.body)
        
        # Validate required fields
        required_fields = ['first_name', 'last_name', 'position', 'term_start', 'term_end']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'success': False,
                    'error': f'Missing required field: {field}'
                }, status=400)
        
        official = BarangayOfficial.objects.create(
            secretary=request.user,
            first_name=data.get('first_name'),
            middle_name=data.get('middle_name', ''),
            last_name=data.get('last_name'),
            suffix=data.get('suffix', ''),
            position=data.get('position'),
            position_type=data.get('position_type', 'elective'),
            term_start=datetime.strptime(data.get('term_start'), '%Y-%m-%d').date(),
            term_end=datetime.strptime(data.get('term_end'), '%Y-%m-%d').date(),
            term_status=data.get('term_status', 'ongoing'),
            email=data.get('email', ''),
            phone=data.get('phone', ''),
            notes=data.get('notes', '')
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Official created successfully',
            'official': {
                'id': official.id,
                'full_name': official.get_full_name(),
                'position': official.position
            }
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except ValueError as e:
        return JsonResponse({
            'success': False,
            'error': f'Invalid date format: {str(e)}'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def update_official(request, official_id):
    """Update an existing official"""
    try:
        official = BarangayOfficial.objects.get(id=official_id, secretary=request.user)
        data = json.loads(request.body)
        
        # Update fields
        official.first_name = data.get('first_name', official.first_name)
        official.middle_name = data.get('middle_name', official.middle_name)
        official.last_name = data.get('last_name', official.last_name)
        official.suffix = data.get('suffix', official.suffix)
        official.position = data.get('position', official.position)
        official.position_type = data.get('position_type', official.position_type)
        
        # Update dates if provided
        if data.get('term_start'):
            official.term_start = datetime.strptime(data.get('term_start'), '%Y-%m-%d').date()
        if data.get('term_end'):
            official.term_end = datetime.strptime(data.get('term_end'), '%Y-%m-%d').date()
        
        official.term_status = data.get('term_status', official.term_status)
        official.email = data.get('email', official.email)
        official.phone = data.get('phone', official.phone)
        official.notes = data.get('notes', official.notes)
        
        official.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Official updated successfully',
            'official': {
                'id': official.id,
                'full_name': official.get_full_name()
            }
        })
    except BarangayOfficial.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Official not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def delete_official(request, official_id):
    """Delete an official"""
    try:
        official = BarangayOfficial.objects.get(id=official_id, secretary=request.user)
        official_name = official.get_full_name()
        official.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Official {official_name} deleted successfully'
        })
    except BarangayOfficial.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Official not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["GET"])
def get_official(request, official_id):
    """Get a single official's details"""
    try:
        official = BarangayOfficial.objects.get(id=official_id, secretary=request.user)
        
        return JsonResponse({
            'success': True,
            'official': {
                'id': official.id,
                'first_name': official.first_name,
                'middle_name': official.middle_name or '',
                'last_name': official.last_name,
                'suffix': official.suffix or '',
                'full_name': official.get_full_name(),
                'display_name': official.get_display_name(),
                'position': official.position,
                'position_type': official.position_type,
                'term_start': official.term_start.strftime('%Y-%m-%d'),
                'term_end': official.term_end.strftime('%Y-%m-%d'),
                'years_served': str(official.years_served) if official.years_served else '0',
                'term_status': official.term_status,
                'email': official.email or '',
                'phone': official.phone or '',
                'notes': official.notes or '',
                'is_term_active': official.is_term_active,
                'days_remaining': official.days_remaining,
            }
        })
    except BarangayOfficial.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Official not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def bulk_create_officials(request):
    """Bulk create officials from CSV data"""
    try:
        data = json.loads(request.body)
        officials_data = data.get('officials', [])
        
        if not officials_data:
            return JsonResponse({
                'success': False,
                'error': 'No officials data provided'
            }, status=400)
        
        created_count = 0
        errors = []
        
        for idx, official_data in enumerate(officials_data):
            try:
                # Validate required fields
                required = ['first_name', 'last_name', 'position', 'term_start', 'term_end']
                missing = [f for f in required if not official_data.get(f)]
                if missing:
                    errors.append(f"Row {idx + 1}: Missing fields: {', '.join(missing)}")
                    continue
                
                BarangayOfficial.objects.create(
                    secretary=request.user,
                    first_name=official_data.get('first_name'),
                    middle_name=official_data.get('middle_name', ''),
                    last_name=official_data.get('last_name'),
                    suffix=official_data.get('suffix', ''),
                    position=official_data.get('position'),
                    position_type=official_data.get('position_type', 'elective'),
                    term_start=datetime.strptime(official_data.get('term_start'), '%Y-%m-%d').date(),
                    term_end=datetime.strptime(official_data.get('term_end'), '%Y-%m-%d').date(),
                    term_status=official_data.get('term_status', 'ongoing'),
                    email=official_data.get('email', ''),
                    phone=official_data.get('phone', ''),
                    notes=official_data.get('notes', '')
                )
                created_count += 1
            except Exception as e:
                errors.append(f"Row {idx + 1}: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully created {created_count} official(s)',
            'created_count': created_count,
            'total_rows': len(officials_data),
            'errors': errors
        })
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# Add this to your views.py

@login_required
@require_http_methods(["GET"])
def generate_official_certificate(request, official_id):
    """Generate eligibility certificate for an official"""
    try:
        official = BarangayOfficial.objects.get(id=official_id, secretary=request.user)
        
        # Return official data formatted for eligibility form
        certificate_data = {
            'success': True,
            'official': {
                'id': official.id,
                'last_name': official.last_name,
                'first_name': official.first_name,
                'middle_initial': official.middle_name[:1] if official.middle_name else '',
                'position': official.position,
                'position_type': official.position_type,
                'term_start': official.term_start.strftime('%Y-%m-%d'),
                'term_end': official.term_end.strftime('%Y-%m-%d'),
                'years_served': str(official.years_served) if official.years_served else '0.0',
                'email': official.email or '',
            }
        }
        
        return JsonResponse(certificate_data)
        
    except BarangayOfficial.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Official not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)




@login_required
def update_profiles(request):
    """
    Update user profile information via modal
    """
    if request.method == 'POST':
        try:
            user = request.user
            profile = user.userprofile
            
            # Update user fields
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            
            if first_name:
                user.first_name = first_name
            if last_name:
                user.last_name = last_name
            if email:
                user.email = email
            
            user.save()
            
            # Update profile preferences (checkboxes)
            profile.email_notifications = 'email_notifications' in request.POST
            profile.deadline_reminders = 'deadline_reminders' in request.POST
            profile.announcements = 'announcements' in request.POST
            profile.compact_view = 'compact_view' in request.POST
            profile.save()
            
            messages.success(request, '✅ Profile updated successfully!')
            
        except Exception as e:
            messages.error(request, f'❌ Error updating profile: {str(e)}')
    
    # Redirect back to the page they came from
    return redirect(request.META.get('HTTP_REFERER', 'barangay_dashboard'))

@login_required
@require_http_methods(["POST"])
def change_passwords(request):
    try:

        current_password = request.POST.get('currentPassword', '')
        new_password = request.POST.get('newPassword', '')
        confirm_password = request.POST.get('confirmPassword', '')
        
        print(f"🔐 Password Change Request")
        print(f"   User: {request.user.username}")
        print(f"   Current password provided: {'Yes' if current_password else 'No'}")
        print(f"   New password length: {len(new_password)}")
        print(f"   Confirm password length: {len(confirm_password)}")

        if not all([current_password, new_password, confirm_password]):
            return JsonResponse({
                'success': False,
                'message': 'All fields are required'
            }, status=400)
        
        if new_password != confirm_password:
            return JsonResponse({
                'success': False,
                'message': 'New passwords do not match'
            }, status=400)
        

        if len(new_password) < 8:
            return JsonResponse({
                'success': False,
                'message': 'Password must be at least 8 characters long'
            }, status=400)
        

        user = request.user
        if not user.check_password(current_password):
            print(f"   ❌ Current password is incorrect")
            return JsonResponse({
                'success': False,
                'message': 'Current password is incorrect'
            }, status=400)
        
        print(f"    Current password verified")
        user.set_password(new_password)
        user.save()      
        print(f"    New password saved to database")
        update_session_auth_hash(request, user)
        
        print(f"    Session updated - user remains logged in")
        print(f" Password changed successfully for user: {user.username}")
        
        return JsonResponse({
            'success': True,
            'message': 'Password changed successfully'
        })
        
    except Exception as e:
        print(f" Error changing password: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while changing password'
        }, status=500)


@login_required
def profile_stats(request):
    """
    Get profile statistics (AJAX endpoint for modal)
    """
    user = request.user
    profile = user.userprofile
    
    stats = {}
    
    # Barangay User Stats
    if profile.role == 'barangay user' and profile.barangay:
        total_submissions = RequirementSubmission.objects.filter(
            barangay=profile.barangay
        ).count()
        
        completed_submissions = RequirementSubmission.objects.filter(
            barangay=profile.barangay,
            status='accomplished'
        ).count()
        
        pending_submissions = RequirementSubmission.objects.filter(
            barangay=profile.barangay,
            status='pending'
        ).count()
        
        eligibility_requests_count = EligibilityRequest.objects.filter(
            barangay=profile.barangay.name
        ).count()
        
        stats = {
            'total_submissions': total_submissions,
            'completed_submissions': completed_submissions,
            'pending_submissions': pending_submissions,
            'eligibility_requests': eligibility_requests_count,
            'login_count': profile.login_count,
            'approval_status': 'Approved' if profile.is_approved else 'Pending',
        }
    
    # DILG Staff Stats
    elif profile.role == 'dilg staff':
        total_barangays_count = Barangay.objects.count()
        total_requirements_count = RequirementSubmission.objects.count()
        
        pending_approvals_count = EligibilityRequest.objects.filter(
            status='pending'
        ).count()
        
        approved_this_month_count = EligibilityRequest.objects.filter(
            status='approved',
            date_processed__month=timezone.now().month
        ).count()
        
        stats = {
            'total_barangays': total_barangays_count,
            'total_requirements': total_requirements_count,
            'pending_approvals': pending_approvals_count,
            'approved_this_month': approved_this_month_count,
            'login_count': profile.login_count,
        }
    
    # Default stats for other roles
    else:
        stats = {
            'login_count': profile.login_count,
            'approval_status': 'Approved' if profile.is_approved else 'Pending',
            'member_since': user.date_joined.year,
        }
    
    return JsonResponse({'success': True, 'stats': stats})

@login_required
def get_sidebar_counts(request):
    """
    Get notification counts for sidebar red dots
    Reusable across all views
    """
    return {
        'pending_applications_count': EligibilityRequest.objects.filter(status='pending').count(),
        'pending_count': UserProfile.objects.filter(is_approved=False).count(),
        'monitoring_count': RequirementSubmission.objects.filter(status__in=['pending', 'in_progress']).count(),
        'certification_count': EligibilityRequest.objects.filter(status='approved').count(),
    }


def terms_conditions(request):
    """
    Display the Terms and Conditions page
    """
    context = {
        'user': request.user,
        'assigned_barangay': getattr(request.user.userprofile, 'barangay', None) if hasattr(request.user, 'userprofile') else None,
        'terms_notification': True,
    }
    
    # Add sidebar counts - PASS request here
    context.update(get_sidebar_counts(request))
    
    return render(request, 'terms_conditions.html', context)

@login_required
@require_http_methods(["POST"])
def accept_terms(request):
    """
    API endpoint to record user acceptance of terms and conditions
    """
    try:
        data = json.loads(request.body)
        
        # Get or create user profile
        user_profile = request.user.userprofile
        
        # Update terms acceptance
        user_profile.terms_accepted = True
        user_profile.terms_accepted_date = timezone.now()
        user_profile.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Terms and Conditions accepted successfully',
            'accepted_date': user_profile.terms_accepted_date.isoformat()
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def check_terms_acceptance(request):
    """
    Check if user has accepted terms and conditions
    """
    try:
        user_profile = request.user.userprofile
        
        return JsonResponse({
            'accepted': user_profile.terms_accepted,
            'date': user_profile.terms_accepted_date.isoformat() if user_profile.terms_accepted_date else None
        })
        
    except Exception as e:
        return JsonResponse({
            'accepted': False,
            'error': str(e)
        }, status=400)
    

# ===== API ENDPOINT 1: GET NOTIFICATIONS =====

@login_required
@require_http_methods(["GET"])
def api_notifications(request):
    """Get all notifications for current user"""
    try:
        notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
        unread_count = notifications.filter(is_read=False).count()
        
        notification_list = []
        for notif in notifications[:50]:  # Last 50 notifications
            notification_list.append({
                'id': notif.id,
                'title': notif.title,
                'message': notif.message,
                'type': notif.notification_type,
                'is_read': notif.is_read,
                'created_at': notif.created_at.isoformat(),
                'time_ago': notif.time_ago(),
                'target_url': get_notification_url(notif),
            })
        
        return JsonResponse({
            'success': True,
            'notifications': notification_list,
            'unread_count': unread_count,
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
# ===== API ENDPOINT 2: MARK AS READ =====
@login_required
@require_http_methods(["POST"])
def api_notification_mark_read(request, notification_id):
    """Mark single notification as read"""
    try:
        notification = Notification.objects.get(id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        
        unread_count = Notification.objects.filter(user=request.user, is_read=False).count()
        
        return JsonResponse({
            'success': True,
            'unread_count': unread_count,
        })
    
    except Notification.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Notification not found'
        }, status=404)
    

    
# ===== API ENDPOINT 3: MARK ALL AS READ =====
@login_required
@require_http_methods(["POST"])
def api_notifications_mark_all_read(request):
    """Mark all notifications as read"""
    try:
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        
        return JsonResponse({
            'success': True,
            'message': 'All notifications marked as read'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
        
    
# ===== HELPER FUNCTION =====
def get_notification_url(notification):
    """Get target URL based on notification type"""
    if notification.submission:
        return f'/requirements/'  # Or specific submission URL
    elif notification.announcement:
        return f'/announcements/'
    elif notification.barangay:
        return f'/dashboard/'
    return '#'
